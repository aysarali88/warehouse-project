import csv
import hashlib
import base64
import io
import json
import hmac
import logging
import os
import re
import secrets
import smtplib
import time
import urllib.parse
import urllib.request
import ssl
import posixpath
import zipfile
from xml.etree import ElementTree
from threading import Lock
import bcrypt
from datetime import datetime, timezone, timedelta
from email.utils import formataddr
from typing import Literal
from zoneinfo import ZoneInfo

from fastapi import Depends, FastAPI, File, Form, HTTPException, Request, Response, UploadFile
from fastapi.responses import FileResponse, JSONResponse, StreamingResponse
from fastapi.staticfiles import StaticFiles
from openpyxl import Workbook, load_workbook
from email.message import EmailMessage
from pydantic import BaseModel, Field
from sqlalchemy import func, or_, text
from sqlalchemy.orm import Session, joinedload, selectinload

from database import Base, SessionLocal, all_engines, all_sessionmakers, engine, get_sessionmaker
from models import (
    AuditLog,
    AppSession,
    AppUser,
    FiberMapArea,
    FiberMapSchematic,
    IssueOrder,
    IssueOrderItem,
    MaterialRequisition,
    MaterialRequisitionItem,
    MaterialReturn,
    MaterialReturnItem,
    MaterialScanLog,
    MaterialTransfer,
    MaterialTransferItem,
    Product,
    ProductSerial,
    ReceiveOrder,
    ReceiveOrderItem,
    RolloutEntryCounter,
    RolloutRecord,
    Site,
    StockBalance,
    StockMovement,
    Technician,
    TechnicianBalance,
    Warehouse,
)

WAREHOUSE_CACHE: dict[str, tuple[float, dict]] = {}
WAREHOUSE_CACHE_TTL = 25
ROLLOUT_CSV_CACHE: tuple[float, list[dict], str] | None = None
ROLLOUT_CSV_CACHE_TTL = 60
ROLLOUT_DB_CACHE: tuple[float, list[dict]] | None = None
ROLLOUT_DB_CACHE_TTL = 30
ROLLOUT_ENTRY_ID_CACHE: tuple[float, str] | None = None
ROLLOUT_ENTRY_ID_CACHE_TTL = 30
ROLLOUT_CODE_REFERENCE_CACHE: dict[str, list[dict]] = {}

# Hay Demashq has no uploaded fiber map yet. These approved HUB codes keep
# Hub Accessories entry available while preserving Area/XBOX validation.
ROLLOUT_MANUAL_HUBS = {
    "haydemashq": {
        "X1": [f"H{number}" for number in range(1, 11)],
        "X2": [f"H{number}" for number in range(1, 6)],
    }
}
ROLLOUT_SYNC_TTL_SECONDS = int(os.getenv("ROLLOUT_SYNC_TTL_SECONDS", "900"))
ROLLOUT_LAST_SYNC_AT = 0.0
ROLLOUT_SYNC_LOCK = Lock()
ROLLOUT_DAILY_PROGRESS_SHEET_ID = "1ZT9e9acJ9Y60J4f_DIFZiYyHa8GvNZdlTvpucHju7Ec"
ROLLOUT_DAILY_PROGRESS_GID = "440090582"
DEFAULT_ROLLOUT_DAILY_PROGRESS_LIVE_CSV_URL = f"https://docs.google.com/spreadsheets/d/{ROLLOUT_DAILY_PROGRESS_SHEET_ID}/gviz/tq?tqx=out:csv&gid={ROLLOUT_DAILY_PROGRESS_GID}"
DEFAULT_ROLLOUT_DAILY_PROGRESS_CSV_URL = "https://docs.google.com/spreadsheets/d/e/2PACX-1vRI1yMD_QsfGAQY3IpwY9X9B3VBO59X_TEGKxUSMQ2S3ciCDbf3lPPGUyXuLrR5os9NI4SBwcyOTWt7/pub?gid=440090582&single=true&output=csv"
FIBER_MAP_REFERENCE_PATH = "static/assets/fiber-map-data.json"
INITIAL_STOCK_REFERENCE_PATH = "static/assets/initial-stock-reference.json"
INITIAL_STOCK_REFERENCE_CACHE: list[dict] | None = None
logger = logging.getLogger(__name__)
SESSION_COOKIE_NAME = "warehouse_session"
SESSION_TTL_HOURS = int(os.getenv("SESSION_TTL_HOURS", "8"))
SESSION_COOKIE_SECURE = os.getenv("SESSION_COOKIE_SECURE", "1").strip().lower() not in {"0", "false", "no", "off"}
SESSION_SECRET = os.getenv("SESSION_SECRET", "").strip()
ACTIVE_USER_WINDOW_SECONDS = 120

if len(SESSION_SECRET) < 32:
    raise RuntimeError("SESSION_SECRET must be set to a value of at least 32 characters")

for program_key, db_engine in all_engines():
    try:
        Base.metadata.create_all(bind=db_engine)
    except Exception:
        logger.exception("%s database initialization failed", program_key)


def ensure_optional_columns(target_engine=engine):
    program_tables = [
        "warehouses",
        "technicians",
        "products",
        "product_serials",
        "stock_balances",
        "technician_balances",
        "stock_movements",
        "receive_orders",
        "issue_orders",
        "material_requisitions",
        "material_transfers",
        "material_returns",
        "material_scan_logs",
        "app_users",
    ]
    if target_engine.dialect.name == "postgresql":
        statements = [
            "ALTER TABLE products ADD COLUMN IF NOT EXISTS qr_code VARCHAR DEFAULT ''",
            "ALTER TABLE products ADD COLUMN IF NOT EXISTS part_number VARCHAR DEFAULT ''",
            "ALTER TABLE products ADD COLUMN IF NOT EXISTS vendor VARCHAR DEFAULT ''",
            "ALTER TABLE material_requisition_items ADD COLUMN IF NOT EXISTS vendor VARCHAR DEFAULT ''",
            "ALTER TABLE receive_orders ADD COLUMN IF NOT EXISTS receipt_date VARCHAR DEFAULT ''",
            "ALTER TABLE app_users ADD COLUMN IF NOT EXISTS email VARCHAR DEFAULT ''",
            "ALTER TABLE app_users ADD COLUMN IF NOT EXISTS warehouse_name VARCHAR DEFAULT ''",
            "ALTER TABLE material_requisitions ADD COLUMN IF NOT EXISTS return_reason TEXT DEFAULT ''",
            "ALTER TABLE rollout_records ADD COLUMN IF NOT EXISTS related_to_xbox VARCHAR DEFAULT ''",
            "ALTER TABLE rollout_records ADD COLUMN IF NOT EXISTS entry_time VARCHAR DEFAULT ''",
            "ALTER TABLE rollout_records ADD COLUMN IF NOT EXISTS cable_code VARCHAR DEFAULT ''",
            "ALTER TABLE rollout_records ADD COLUMN IF NOT EXISTS box_code VARCHAR DEFAULT ''",
            "ALTER TABLE rollout_records ADD COLUMN IF NOT EXISTS olt VARCHAR DEFAULT ''",
            "ALTER TABLE rollout_records ADD COLUMN IF NOT EXISTS cable_route VARCHAR DEFAULT ''",
            "ALTER TABLE rollout_records ADD COLUMN IF NOT EXISTS notes VARCHAR DEFAULT ''",
            "ALTER TABLE rollout_records ADD COLUMN IF NOT EXISTS submission_key VARCHAR DEFAULT ''",
            "ALTER TABLE stock_movements ADD COLUMN IF NOT EXISTS source_item_id INTEGER",
        ]
        statements.extend([f"ALTER TABLE {table} ADD COLUMN IF NOT EXISTS program VARCHAR NOT NULL DEFAULT 'FTTH'" for table in program_tables])
        statements.extend(
            [
                "ALTER TABLE warehouses DROP CONSTRAINT IF EXISTS warehouses_name_key",
                "ALTER TABLE products DROP CONSTRAINT IF EXISTS products_sku_key",
                "ALTER TABLE product_serials DROP CONSTRAINT IF EXISTS product_serials_serial_number_key",
                "ALTER TABLE receive_orders DROP CONSTRAINT IF EXISTS receive_orders_order_number_key",
                "ALTER TABLE issue_orders DROP CONSTRAINT IF EXISTS issue_orders_order_number_key",
                "ALTER TABLE material_requisitions DROP CONSTRAINT IF EXISTS material_requisitions_order_number_key",
                "ALTER TABLE material_transfers DROP CONSTRAINT IF EXISTS material_transfers_transfer_number_key",
                "ALTER TABLE material_returns DROP CONSTRAINT IF EXISTS material_returns_return_number_key",
                "DROP INDEX IF EXISTS ix_warehouses_name",
                "DROP INDEX IF EXISTS ix_products_sku",
                "DROP INDEX IF EXISTS ix_product_serials_serial_number",
                "DROP INDEX IF EXISTS ix_receive_orders_order_number",
                "DROP INDEX IF EXISTS ix_issue_orders_order_number",
                "DROP INDEX IF EXISTS ix_material_requisitions_order_number",
                "DROP INDEX IF EXISTS ix_material_transfers_transfer_number",
                "DROP INDEX IF EXISTS ix_material_returns_return_number",
                "CREATE UNIQUE INDEX IF NOT EXISTS uq_warehouse_program_name ON warehouses (program, name)",
                "CREATE UNIQUE INDEX IF NOT EXISTS uq_product_program_sku ON products (program, sku)",
                "CREATE UNIQUE INDEX IF NOT EXISTS uq_product_serial_program_serial ON product_serials (program, serial_number)",
                "CREATE UNIQUE INDEX IF NOT EXISTS uq_receive_order_program_number ON receive_orders (program, order_number)",
                "CREATE UNIQUE INDEX IF NOT EXISTS uq_issue_order_program_number ON issue_orders (program, order_number)",
                "CREATE UNIQUE INDEX IF NOT EXISTS uq_material_requisition_program_number ON material_requisitions (program, order_number)",
                "CREATE UNIQUE INDEX IF NOT EXISTS uq_material_transfer_program_number ON material_transfers (program, transfer_number)",
                "CREATE UNIQUE INDEX IF NOT EXISTS uq_material_return_program_number ON material_returns (program, return_number)",
                "CREATE UNIQUE INDEX IF NOT EXISTS uq_rollout_records_submission_key ON rollout_records (submission_key) WHERE submission_key <> ''",
                "CREATE UNIQUE INDEX IF NOT EXISTS uq_stock_return_movement_source ON stock_movements (program, movement_type, source_item_id) WHERE movement_type = 'return_in' AND source_item_id IS NOT NULL",
            ]
        )
    else:
        statements = [
            "ALTER TABLE products ADD COLUMN qr_code VARCHAR DEFAULT ''",
            "ALTER TABLE products ADD COLUMN part_number VARCHAR DEFAULT ''",
            "ALTER TABLE products ADD COLUMN vendor VARCHAR DEFAULT ''",
            "ALTER TABLE material_requisition_items ADD COLUMN vendor VARCHAR DEFAULT ''",
            "ALTER TABLE receive_orders ADD COLUMN receipt_date VARCHAR DEFAULT ''",
            "ALTER TABLE app_users ADD COLUMN email VARCHAR DEFAULT ''",
            "ALTER TABLE app_users ADD COLUMN warehouse_name VARCHAR DEFAULT ''",
            "ALTER TABLE material_requisitions ADD COLUMN return_reason TEXT DEFAULT ''",
            "ALTER TABLE rollout_records ADD COLUMN related_to_xbox VARCHAR DEFAULT ''",
            "ALTER TABLE rollout_records ADD COLUMN entry_time VARCHAR DEFAULT ''",
            "ALTER TABLE rollout_records ADD COLUMN cable_code VARCHAR DEFAULT ''",
            "ALTER TABLE rollout_records ADD COLUMN box_code VARCHAR DEFAULT ''",
            "ALTER TABLE rollout_records ADD COLUMN olt VARCHAR DEFAULT ''",
            "ALTER TABLE rollout_records ADD COLUMN cable_route VARCHAR DEFAULT ''",
            "ALTER TABLE rollout_records ADD COLUMN notes VARCHAR DEFAULT ''",
            "ALTER TABLE rollout_records ADD COLUMN submission_key VARCHAR DEFAULT ''",
            "ALTER TABLE stock_movements ADD COLUMN source_item_id INTEGER",
        ]
        statements.extend([f"ALTER TABLE {table} ADD COLUMN program VARCHAR NOT NULL DEFAULT 'FTTH'" for table in program_tables])
        statements.extend(
            [
                "DROP INDEX IF EXISTS ix_warehouses_name",
                "DROP INDEX IF EXISTS ix_products_sku",
                "DROP INDEX IF EXISTS ix_product_serials_serial_number",
                "DROP INDEX IF EXISTS ix_receive_orders_order_number",
                "DROP INDEX IF EXISTS ix_issue_orders_order_number",
                "DROP INDEX IF EXISTS ix_material_requisitions_order_number",
                "DROP INDEX IF EXISTS ix_material_transfers_transfer_number",
                "DROP INDEX IF EXISTS ix_material_returns_return_number",
                "CREATE UNIQUE INDEX IF NOT EXISTS uq_warehouse_program_name ON warehouses (program, name)",
                "CREATE UNIQUE INDEX IF NOT EXISTS uq_product_program_sku ON products (program, sku)",
                "CREATE UNIQUE INDEX IF NOT EXISTS uq_product_serial_program_serial ON product_serials (program, serial_number)",
                "CREATE UNIQUE INDEX IF NOT EXISTS uq_receive_order_program_number ON receive_orders (program, order_number)",
                "CREATE UNIQUE INDEX IF NOT EXISTS uq_issue_order_program_number ON issue_orders (program, order_number)",
                "CREATE UNIQUE INDEX IF NOT EXISTS uq_material_requisition_program_number ON material_requisitions (program, order_number)",
                "CREATE UNIQUE INDEX IF NOT EXISTS uq_material_transfer_program_number ON material_transfers (program, transfer_number)",
                "CREATE UNIQUE INDEX IF NOT EXISTS uq_material_return_program_number ON material_returns (program, return_number)",
                "CREATE UNIQUE INDEX IF NOT EXISTS uq_rollout_records_submission_key ON rollout_records (submission_key) WHERE submission_key <> ''",
                "CREATE UNIQUE INDEX IF NOT EXISTS uq_stock_return_movement_source ON stock_movements (program, movement_type, source_item_id) WHERE movement_type = 'return_in' AND source_item_id IS NOT NULL",
            ]
        )
    with target_engine.begin() as conn:
        for statement in statements:
            try:
                conn.execute(text(statement))
            except Exception:
                pass


for program_key, db_engine in all_engines():
    try:
        ensure_optional_columns(db_engine)
    except Exception:
        logger.exception("%s optional database migration failed", program_key)

app = FastAPI(title="FTTH Rollout")
app.mount("/static", StaticFiles(directory="static"), name="static")

DEFAULT_PROGRAM = "FTTH"
SINGLE_RAN_PROGRAM = "SINGLE_RAN"
PROGRAM_LABELS = {
    DEFAULT_PROGRAM: "FTTH",
    SINGLE_RAN_PROGRAM: "Single RAN",
}


def sync_single_ran_product_vendors():
    for program_key, session_factory in all_sessionmakers():
        if program_key != SINGLE_RAN_PROGRAM:
            continue
        db = session_factory()
        try:
            rows = (
                db.query(Product, ReceiveOrder.supplier)
                .join(ReceiveOrderItem, ReceiveOrderItem.product_id == Product.id)
                .join(ReceiveOrder, ReceiveOrder.id == ReceiveOrderItem.receive_order_id)
                .filter(Product.program == SINGLE_RAN_PROGRAM, ReceiveOrder.program == SINGLE_RAN_PROGRAM)
                .order_by(ReceiveOrder.id.asc())
                .all()
            )
            changed = 0
            for product, supplier in rows:
                vendor = str(supplier or "").strip()
                if vendor and not str(product.vendor or "").strip():
                    product.vendor = vendor
                    changed += 1
            requisition_items = (
                db.query(MaterialRequisitionItem)
                .join(MaterialRequisition, MaterialRequisition.id == MaterialRequisitionItem.requisition_id)
                .join(Product, Product.id == MaterialRequisitionItem.product_id)
                .filter(MaterialRequisition.program == SINGLE_RAN_PROGRAM)
                .all()
            )
            for item in requisition_items:
                vendor = str(item.product.vendor or "").strip() if item.product else ""
                if vendor and not str(item.vendor or "").strip():
                    item.vendor = vendor
                    changed += 1
            if changed:
                db.commit()
                logger.info("Single RAN product vendors populated for %s materials", changed)
        except Exception:
            db.rollback()
            logger.exception("Single RAN product vendor population failed")
        finally:
            db.close()


sync_single_ran_product_vendors()
VALID_PROGRAM_VALUES = {DEFAULT_PROGRAM, SINGLE_RAN_PROGRAM, "SR", "SINGLERAN"}
DEFAULT_SITE_IDS = ("Maqawba", "Hay Al Andalus Z2", "Ras A Tota")


def raw_program_value(value: str = "") -> str:
    return str(value or "").strip().upper().replace("-", "_").replace(" ", "_")


def normalize_program(value: str = "") -> str:
    text = raw_program_value(value)
    return SINGLE_RAN_PROGRAM if text in {"SINGLE_RAN", "SR", "SINGLERAN"} else DEFAULT_PROGRAM


def is_valid_program_value(value: str = "") -> bool:
    return raw_program_value(value) in VALID_PROGRAM_VALUES


def is_single_ran(program: str = "") -> bool:
    return normalize_program(program) == SINGLE_RAN_PROGRAM


def ensure_default_site_ids():
    db = get_sessionmaker(DEFAULT_PROGRAM)()
    try:
        existing = {
            row.name.strip().casefold()
            for row in db.query(Site).filter(Site.program == DEFAULT_PROGRAM).all()
        }
        for name in DEFAULT_SITE_IDS:
            if name.casefold() not in existing:
                db.add(Site(program=DEFAULT_PROGRAM, name=name))
        db.commit()
    except Exception:
        db.rollback()
        logger.exception("Default Site ID initialization failed")
    finally:
        db.close()


ensure_default_site_ids()


def program_filter(model, program: str = ""):
    return getattr(model, "program") == normalize_program(program)


def extract_program_from_body(body: bytes, content_type: str = "") -> str:
    if not body:
        return ""
    text_content_type = str(content_type or "").lower()
    if "application/json" in text_content_type:
        try:
            payload = json.loads(body.decode("utf-8"))
        except Exception:
            return ""
        if isinstance(payload, dict):
            return str(payload.get("program") or "")
        return ""
    text = body.decode("utf-8", errors="ignore")
    if "name=\"program\"" not in text:
        return ""
    match = re.search(r'name="program"[\s\S]*?\r?\n\r?\n([^\r\n-]+)', text)
    return match.group(1).strip() if match else ""


@app.middleware("http")
async def require_program_scope(request: Request, call_next):
    path = request.url.path
    if not path.startswith("/api/"):
        return await call_next(request)

    is_login = path == "/api/auth/login"
    token = request.cookies.get(SESSION_COOKIE_NAME, "")
    if not is_login:
        if not token:
            return JSONResponse({"detail": "Authentication required"}, status_code=401)
        session_data = None
        for _, sessionmaker in all_sessionmakers():
            with sessionmaker() as db:
                session = (
                    db.query(AppSession)
                    .options(joinedload(AppSession.user))
                    .filter(AppSession.token_hash == session_token_hash(token), AppSession.expires_at > datetime.now(timezone.utc))
                    .first()
                )
                if session is not None and session.user is not None and session.user.status == "active":
                    session_data = (session.user, normalize_program(session.program), session.csrf_token)
                    break
        if session_data is None:
            return JSONResponse({"detail": "Session expired or inactive"}, status_code=401)
        request.state.current_user, request.state.session_program, request.state.csrf_token = session_data
        request.state.program = request.state.session_program
        if request.method.upper() not in {"GET", "HEAD", "OPTIONS"}:
            supplied = request.headers.get("X-CSRF-Token", "")
            if not supplied or not hmac.compare_digest(supplied, request.state.csrf_token):
                return JSONResponse({"detail": "Invalid CSRF token"}, status_code=403)

    scoped_path = path.startswith("/api/warehouse") or path.startswith("/api/auth/users") or is_login
    if not scoped_path:
        return await call_next(request)

    program_value = request.query_params.get("program", "")
    body = b""
    if not program_value and request.method.upper() in {"POST", "PUT", "PATCH", "DELETE"}:
        body = await request.body()
        program_value = extract_program_from_body(body, request.headers.get("content-type", ""))

        async def receive():
            return {"type": "http.request", "body": body, "more_body": False}

        request._receive = receive

    if not program_value:
        return JSONResponse({"detail": "Program scope is required"}, status_code=400)
    if not is_valid_program_value(program_value):
        return JSONResponse({"detail": "Invalid program scope"}, status_code=400)
    request.state.program = normalize_program(program_value)
    if not is_login and request.state.program != request.state.session_program:
        return JSONResponse({"detail": "Program scope does not match this session"}, status_code=403)
    return await call_next(request)


def request_arrived_over_https(request: Request) -> bool:
    """Recognize HTTPS when Render terminates TLS before forwarding the request."""
    forwarded_proto = request.headers.get("X-Forwarded-Proto", "").split(",", 1)[0].strip().lower()
    return request.url.scheme == "https" or forwarded_proto == "https"


@app.middleware("http")
async def add_hsts_header(request: Request, call_next):
    response = await call_next(request)
    response.headers["X-Content-Type-Options"] = "nosniff"
    if request_arrived_over_https(request):
        response.headers["Strict-Transport-Security"] = "max-age=31536000"
    return response


def is_admin_role(role: str = "") -> bool:
    return str(role or "").strip().lower() == "admin"


def app_user_matches_program(user, program: str = DEFAULT_PROGRAM) -> bool:
    if is_admin_role(getattr(user, "role", "") if not isinstance(user, dict) else user.get("role", "")):
        return True
    user_program = user.get("program", DEFAULT_PROGRAM) if isinstance(user, dict) else getattr(user, "program", DEFAULT_PROGRAM)
    return normalize_program(user_program) == normalize_program(program)


APP_USERS = [
    {"username": "Aysar", "name": "Aysar", "role": "Admin", "password_hash": "$2b$12$3c72OtpvCsB4.CNyImCvcuQN//O7KqonCK3QuZWJs3jO/oUH6DMMO"},
    {"username": "Hamza", "name": "Hamza", "role": "Admin", "password_hash": "$2b$12$xAUkLrhxFwUXzQtTrrs86OQnEUl1a16kQhQtMlBtrfIN3HvHSZTVK"},
    {"username": "Aysar", "name": "Aysar", "role": "Requester", "password_hash": "$2b$12$xA.6M2y7h5OiplZjVzMi8.qsIoHbsjzt8.1zVBKQidm6u3UYiHYum"},
    {"username": "Hamza", "name": "Hamza", "role": "Requester", "password_hash": "$2b$12$LKYijkxCRZVxKPeY0AIA/O2I4actBdhNWVefJ7CN9c/1ituhQzt7q"},
    {"username": "Ryadh", "name": "Ryadh", "role": "Requester", "password_hash": "$2b$12$WiquFYrheGfO2LtKtswLo.J85oZ3IzlnI0md6tsQQTY4VqXR6hX12"},
    {"username": "Adel", "name": "Adel", "role": "Requester", "password_hash": "$2b$12$IaMjT3MNkht3xqCaj7jP4e2LmNwN7BpPzjs8YyS69pIHzX/6xZSPm"},
    {"username": "Nadeer", "name": "Nadeer", "role": "Requester", "password_hash": "$2b$12$DHRZmTb0yv1Qn7KtgV0L3Ongy6HznxkkLO1SN15v/3rsKSptFyOcq"},
    {"username": "Ghassan", "name": "Ghassan", "role": "Requester", "password_hash": "$2b$12$zEzqVlt6sJZgj7kQCsWY3.BElMy5nICAHRKq4dUcRvKrQba5B0Ef6"},
    {"username": "Mustafa", "name": "Mustafa", "role": "Approval", "password_hash": "$2b$12$7P79odcv0uV9bKpbOY6NWe0omDZ3ZYgCsFWp3me8i7cveBjcTU6YO"},
    {"username": "Tripoli", "name": "Tripoli", "role": "Warehouse Manager", "password_hash": "$2b$12$BqOEHrljHOmLipn6j4g1j.nKJpWWrp2SXu9fWw2.cUoZ434jKQCfy", "warehouse_name": "Tripoli"},
    {"username": "Misurata", "name": "Misurata", "role": "Warehouse Manager", "password_hash": "$2b$12$BqOEHrljHOmLipn6j4g1j.nKJpWWrp2SXu9fWw2.cUoZ434jKQCfy", "warehouse_name": "Misurata"},
    {"username": "FreeZone", "name": "FreeZone", "role": "Warehouse Manager", "password_hash": "$2b$12$dLg/7wO3.EBdDBzislnwlujcdQoMlPwrGi6H61X3OwMIiq3PgoIQS", "warehouse_name": "FreeZone"},
]

TEMP_MR_WAREHOUSE_MANAGER_OVERRIDE = ""


try:
    TRIPOLI_TZ = ZoneInfo("Africa/Tripoli")
except Exception:
    TRIPOLI_TZ = timezone(timedelta(hours=2))


def local_today() -> str:
    return datetime.now(TRIPOLI_TZ).date().isoformat()


def app_user_key(username: str, role: str) -> tuple[str, str]:
    return (username.strip().lower(), role.strip().lower())


def is_bcrypt_hash(value: str = "") -> bool:
    text = str(value or "")
    return text.startswith(("$2a$", "$2b$", "$2y$"))


def is_legacy_password_hash(value: str = "") -> bool:
    return str(value or "").startswith("pbkdf2_sha256$")


def hash_password(password: str) -> str:
    return bcrypt.hashpw(password.encode("utf-8"), bcrypt.gensalt(rounds=12)).decode("utf-8")


def verify_password(password: str, stored: str = "") -> bool:
    if not password or not stored:
        return False
    if is_bcrypt_hash(stored):
        try:
            return bcrypt.checkpw(password.encode("utf-8"), stored.encode("utf-8"))
        except ValueError:
            return False
    if is_legacy_password_hash(stored):
        try:
            _algorithm, iterations, salt, digest = stored.split("$", 3)
            derived = hashlib.pbkdf2_hmac("sha256", password.encode("utf-8"), salt.encode("utf-8"), int(iterations))
            expected = base64.b64decode(digest.encode("ascii"))
            return hmac.compare_digest(derived, expected)
        except (ValueError, TypeError, UnicodeError):
            return False
    return hmac.compare_digest(password, stored)


def serialize_app_user(row, fallback: bool = False) -> dict:
    return {
        "id": getattr(row, "id", "") or "",
        "program": normalize_program(row.get("program", DEFAULT_PROGRAM) if fallback else getattr(row, "program", DEFAULT_PROGRAM)),
        "username": row["username"] if fallback else row.username,
        "name": (row.get("name") if fallback else row.name) or (row["username"] if fallback else row.username),
        "role": row["role"] if fallback else row.role,
        "email": (row.get("email") if fallback else row.email) or "",
        "warehouse_name": (row.get("warehouse_name") if fallback else row.warehouse_name) or "",
        "is_fallback": fallback,
    }


def fallback_user_template(username: str, role: str) -> dict | None:
    key = app_user_key(username, role)
    return next((row for row in APP_USERS if app_user_key(row["username"], row["role"]) == key), None)


def migrate_app_user_password_hashes() -> None:
    for program_key, session_factory in all_sessionmakers():
        try:
            with session_factory() as db:
                rows = db.query(AppUser).all()
                changed = False
                for row in rows:
                    if row.password_hash and not is_bcrypt_hash(row.password_hash) and not is_legacy_password_hash(row.password_hash) and len(str(row.password_hash)) <= 72:
                        row.password_hash = hash_password(row.password_hash)
                        changed = True
                if changed:
                    db.commit()
        except Exception:
            logger.exception("%s app user password migration failed", program_key)


migrate_app_user_password_hashes()


def seed_legacy_app_users() -> None:
    """Seed the original FTTH legacy accounts without copying them into Single RAN."""
    for program_key, session_factory in all_sessionmakers():
        if is_single_ran(program_key):
            continue
        with session_factory() as db:
            changed = False
            for legacy in APP_USERS:
                # Production has a global unique constraint on username, so an
                # existing account must never be reseeded under another role.
                existing = db.query(AppUser).filter(
                    func.lower(AppUser.username) == legacy["username"].lower(),
                ).first()
                if existing is None:
                    db.add(AppUser(program=normalize_program(program_key), username=legacy["username"], name=legacy["name"], role=legacy["role"], warehouse_name=legacy.get("warehouse_name", ""), password_hash=legacy["password_hash"], status="active"))
                    changed = True
            if changed:
                db.commit()


seed_legacy_app_users()


def sync_admin_users_to_secondary_databases() -> None:
    sessions = all_sessionmakers()
    if len(sessions) < 2:
        return
    try:
        with sessions[0][1]() as primary_db:
            admins = (
                primary_db.query(AppUser)
                .filter(AppUser.status == "active", func.lower(AppUser.role) == "admin")
                .all()
            )
            admin_rows = [
                {
                    "username": row.username,
                    "name": row.name or row.username,
                    "role": row.role,
                    "password_hash": row.password_hash,
                    "email": row.email or "",
                    "warehouse_name": row.warehouse_name or "",
                }
                for row in admins
                if row.username and row.password_hash
            ]
    except Exception:
        logger.exception("Primary admin user read failed")
        return
    for program_key, session_factory in sessions[1:]:
        try:
            with session_factory() as db:
                changed = False
                for row in admin_rows:
                    existing = (
                        db.query(AppUser)
                        .filter(func.lower(AppUser.username) == row["username"].lower(), func.lower(AppUser.role) == "admin")
                        .first()
                    )
                    if existing:
                        continue
                    db.add(
                        AppUser(
                            program=program_key,
                            username=row["username"],
                            name=row["name"],
                            role=row["role"],
                            password_hash=row["password_hash"],
                            email=row["email"],
                            warehouse_name=row["warehouse_name"],
                            status="active",
                        )
                    )
                    changed = True
                if changed:
                    db.commit()
        except Exception:
            logger.exception("%s admin user sync failed", program_key)


# Users are intentionally managed per program. Do not mirror FTTH admins into
# Single RAN automatically, because that would recreate accounts removed there.


def sync_product_part_numbers() -> None:
    for program_key, session_factory in all_sessionmakers():
        try:
            with session_factory() as db:
                rows = db.query(Product).all()
                changed = False
                for row in rows:
                    desired = product_part_number(row.sku, row.part_number)
                    if desired != (row.part_number or ""):
                        row.part_number = desired
                        changed = True
                if changed:
                    db.commit()
        except Exception:
            logger.exception("%s product part number sync failed", program_key)



def env_bool(name: str, default: bool = False) -> bool:
    value = os.getenv(name, "").strip().lower()
    if not value:
        return default
    return value in {"1", "true", "yes", "on"}


def env_list(name: str) -> list[str]:
    raw = os.getenv(name, "")
    return [part.strip() for part in raw.split(",") if part.strip()]


def send_email(to_emails: list[str], subject: str, body: str) -> bool:
    host = os.getenv("SMTP_HOST", "").strip()
    try:
        port = int(os.getenv("SMTP_PORT", "587"))
    except ValueError:
        logger.warning("MR email skipped: SMTP_PORT is invalid")
        return False
    username = os.getenv("SMTP_USERNAME", "").strip()
    password = os.getenv("SMTP_PASSWORD", "").strip()
    from_email = os.getenv("EMAIL_FROM", username).strip()
    from_name = os.getenv("EMAIL_FROM_NAME", "Global Technology Company").strip()
    if not host or not port or not username or not password or not from_email or not to_emails:
        logger.warning("MR email skipped: SMTP settings or recipients are missing")
        return False

    msg = EmailMessage()
    msg["From"] = formataddr((from_name, from_email))
    msg["Reply-To"] = from_email
    msg["To"] = ", ".join(to_emails)
    msg["Subject"] = subject
    msg.set_content(body)

    use_ssl = env_bool("SMTP_USE_SSL", port == 465)
    use_tls = env_bool("SMTP_USE_TLS", not use_ssl and port != 465)
    context = ssl.create_default_context()

    if use_ssl:
        with smtplib.SMTP_SSL(host, port, timeout=15, context=context) as smtp:
            smtp.login(username, password)
            smtp.send_message(msg)
    else:
        with smtplib.SMTP(host, port, timeout=15) as smtp:
            if use_tls:
                smtp.starttls(context=context)
            smtp.login(username, password)
            smtp.send_message(msg)
    return True


def approval_notification_emails(db: Session) -> list[str]:
    rows = (
        db.query(AppUser)
        .filter(AppUser.status == "active", func.lower(AppUser.role) == "approval")
        .all()
    )
    emails = [row.email.strip() for row in rows if getattr(row, "email", "") and "@" in row.email]
    return emails or env_list("APPROVAL_EMAIL_RECIPIENTS")


def normalize_email_list(values: list[str]) -> list[str]:
    emails: list[str] = []
    seen: set[str] = set()
    for value in values:
        email = str(value or "").strip()
        if "@" not in email:
            continue
        key = email.lower()
        if key in seen:
            continue
        seen.add(key)
        emails.append(email)
    return emails


def active_user_emails(db: Session, role: str, identifiers: list[str] | None = None) -> list[str]:
    query = db.query(AppUser).filter(AppUser.status == "active", func.lower(AppUser.role) == role.strip().lower())
    rows = query.all()
    if identifiers:
        keys: set[str] = set()
        for value in identifiers:
            if str(value or "").strip():
                keys.update(warehouse_scope_keys(value) or {normalize_usage_key(value)})
        rows = [
            row
            for row in rows
            if normalize_usage_key(row.username) in keys
            or normalize_usage_key(row.name) in keys
            or bool(warehouse_scope_keys(getattr(row, "warehouse_name", "")).intersection(keys))
        ]
    return normalize_email_list([row.email for row in rows])


def requester_notification_emails(row: MaterialRequisition, db: Session) -> list[str]:
    identifiers = [row.requester_name, row.created_by]
    return active_user_emails(db, "requester", identifiers)


def warehouse_scope_keys(value: str) -> set[str]:
    key = normalize_usage_key(value)
    if not key:
        return set()
    keys = {key}
    if any(token in key for token in ("misurata", "misrata", "misrat")):
        keys.update({"misurata", "misrata", "misuratalnet", "misratalnet"})
    if "freezone" in key:
        keys.update({"freezone", "misuratafreezone", "misratafreezone"})
    if "tripoli" in key:
        keys.add("tripoli")
    return keys


def warehouse_scope_matches(viewer: str, warehouse_name: str) -> bool:
    viewer_keys = warehouse_scope_keys(viewer)
    warehouse_key = normalize_usage_key(warehouse_name)
    if "freezone" in warehouse_key:
        warehouse_keys = {"freezone", "misuratafreezone", "misratafreezone"}
    elif any(token in warehouse_key for token in ("misurata", "misrata", "misrat")):
        warehouse_keys = {"misurata", "misrata", "misuratalnet", "misratalnet"}
    elif "tripoli" in warehouse_key:
        warehouse_keys = {"tripoli"}
    else:
        warehouse_keys = warehouse_scope_keys(warehouse_name)
    return bool(viewer_keys and warehouse_keys and viewer_keys.intersection(warehouse_keys))


def warehouse_manager_handles_mr(viewer: str, row: MaterialRequisition) -> bool:
    override_key = normalize_usage_key(TEMP_MR_WAREHOUSE_MANAGER_OVERRIDE)
    if override_key:
        return override_key in warehouse_scope_keys(viewer)
    return warehouse_scope_matches(viewer, row.warehouse.name if row.warehouse else "")


def warehouse_manager_notification_emails(row: MaterialRequisition, db: Session) -> list[str]:
    identifiers = [row.site_address or TEMP_MR_WAREHOUSE_MANAGER_OVERRIDE or (row.warehouse.name if row.warehouse else "")]
    return active_user_emails(db, "warehouse manager", identifiers)


def transfer_source_warehouse_manager_emails(row: MaterialTransfer, db: Session) -> list[str]:
    identifiers = [row.from_warehouse.name if row.from_warehouse else ""]
    return active_user_emails(db, "warehouse manager", identifiers)


def transfer_destination_warehouse_manager_emails(row: MaterialTransfer, db: Session) -> list[str]:
    identifiers = [row.to_warehouse.name if row.to_warehouse else ""]
    return active_user_emails(db, "warehouse manager", identifiers)


def is_source_warehouse_manager(actor: str, row: MaterialTransfer, db: Session) -> bool:
    actor_key = normalize_usage_key(actor)
    warehouse_name = row.from_warehouse.name if row.from_warehouse else ""
    if not actor_key or not warehouse_name:
        return False
    rows = (
        db.query(AppUser)
        .filter(AppUser.status == "active", func.lower(AppUser.role) == "warehouse manager")
        .all()
    )
    return any(
        (normalize_usage_key(user.username) == actor_key or normalize_usage_key(user.name) == actor_key)
        and warehouse_scope_matches(getattr(user, "warehouse_name", ""), warehouse_name)
        for user in rows
    )


def is_requisition_warehouse_manager(actor: str, row: MaterialRequisition, db: Session) -> bool:
    actor_key = normalize_usage_key(actor)
    warehouse_name = row.warehouse.name if row.warehouse else ""
    if not actor_key or not warehouse_name:
        return False
    rows = (
        db.query(AppUser)
        .filter(AppUser.status == "active", func.lower(AppUser.role) == "warehouse manager")
        .all()
    )
    return any(
        (normalize_usage_key(user.username) == actor_key or normalize_usage_key(user.name) == actor_key)
        and warehouse_scope_matches(getattr(user, "warehouse_name", ""), warehouse_name)
        for user in rows
    )


def notify_mr_email(row: MaterialRequisition, db: Session, recipients: list[str], subject: str, lines: list[str], audit_prefix: str) -> None:
    recipients = normalize_email_list(recipients)
    if not recipients:
        logger.warning("MR email skipped: no recipients configured for %s", audit_prefix)
        try:
            log_audit(
                db,
                f"{audit_prefix}_skipped",
                "material_requisition",
                row.order_number,
                "system",
                {"reason": "no_recipients", "recipients": []},
            )
            db.commit()
        except Exception:
            db.rollback()
        return
    try:
        send_email(recipients, subject, "\n".join(lines))
        log_audit(
            db,
            f"{audit_prefix}_sent",
            "material_requisition",
            row.order_number,
            "system",
            {"recipients": recipients},
        )
        db.commit()
    except Exception:
        db.rollback()
        logger.exception("Failed to send MR email: %s", audit_prefix)
        try:
            log_audit(
                db,
                f"{audit_prefix}_failed",
                "material_requisition",
                row.order_number,
                "system",
                {"recipients": recipients},
            )
            db.commit()
        except Exception:
            db.rollback()


def notify_transfer_email(row: MaterialTransfer, db: Session, recipients: list[str], subject: str, lines: list[str], audit_prefix: str) -> None:
    recipients = normalize_email_list(recipients)
    if not recipients:
        logger.warning("Transfer email skipped: no recipients configured for %s", audit_prefix)
        return
    try:
        send_email(recipients, subject, "\n".join(lines))
        log_audit(
            db,
            f"{audit_prefix}_sent",
            "material_transfer",
            row.transfer_number,
            "system",
            {"recipients": recipients},
        )
        db.commit()
    except Exception:
        db.rollback()
        logger.exception("Failed to send transfer email: %s", audit_prefix)


def notify_transfer_created(row: MaterialTransfer, db: Session) -> None:
    from_name = row.from_warehouse.name if row.from_warehouse else ""
    to_name = row.to_warehouse.name if row.to_warehouse else ""
    notify_transfer_email(
        row,
        db,
        approval_notification_emails(db),
        f"Approval needed: Material Transfer {row.transfer_number}",
        [
            "Hello,",
            "",
            "A material transfer has been created and is waiting for approval in the warehouse system.",
            "",
            f"Transfer No: {row.transfer_number}",
            f"Requester: {row.requester_name or '-'}",
            f"From Warehouse: {from_name or '-'}",
            f"To Warehouse: {to_name or '-'}",
            f"Status: Pending approval",
            f"Date: {row.transfer_date or local_today()}",
            "",
            "Please sign in to the warehouse system and review this transfer when convenient.",
            "",
            "This is an automated notification from Global Technology Company.",
        ],
        "transfer_created_email",
    )


def notify_transfer_approved(row: MaterialTransfer, db: Session) -> None:
    from_name = row.from_warehouse.name if row.from_warehouse else ""
    to_name = row.to_warehouse.name if row.to_warehouse else ""
    notify_transfer_email(
        row,
        db,
        transfer_destination_warehouse_manager_emails(row, db),
        f"Receiving action needed: Material Transfer {row.transfer_number}",
        [
            "Hello,",
            "",
            "A material transfer to your warehouse has been approved and is waiting for physical receiving confirmation.",
            "",
            f"Transfer No: {row.transfer_number}",
            f"From Warehouse: {from_name or '-'}",
            f"To Warehouse: {to_name or '-'}",
            f"Approved by: {row.approver_name or '-'}",
            "Status: Approved - waiting for your confirmation",
            "",
            "Please check the delivered materials, then sign in to the warehouse system and select Confirm.",
            "Stock will move only after your confirmation.",
            "",
            "This is an automated notification from Global Technology Company.",
        ],
        "transfer_approved_destination_email",
    )


def notify_transfer_returned_for_edit(row: MaterialTransfer, db: Session) -> None:
    from_name = row.from_warehouse.name if row.from_warehouse else ""
    to_name = row.to_warehouse.name if row.to_warehouse else ""
    notify_transfer_email(
        row,
        db,
        requester_notification_emails(row, db),
        f"Material Transfer {row.transfer_number} returned for edit",
        [
            "Hello,",
            "",
            "Your material transfer was returned for update.",
            "",
            f"Transfer No: {row.transfer_number}",
            f"From Warehouse: {from_name or '-'}",
            f"To Warehouse: {to_name or '-'}",
            f"Return reason: {row.approver_comment or '-'}",
            "Status: Returned for edit",
            "",
            "Please sign in to the warehouse system, update the transfer, and submit it again.",
            "",
            "This is an automated notification from Global Technology Company.",
        ],
        "transfer_returned_for_edit_email",
    )


def notify_transfer_returned_by_destination(row: MaterialTransfer, db: Session) -> None:
    from_name = row.from_warehouse.name if row.from_warehouse else ""
    to_name = row.to_warehouse.name if row.to_warehouse else ""
    notify_transfer_email(
        row,
        db,
        requester_notification_emails(row, db),
        f"Material Transfer {row.transfer_number} returned by receiving warehouse",
        [
            "Hello,",
            "",
            "Your approved material transfer was returned by the receiving warehouse before confirmation.",
            "No stock was moved.",
            "",
            f"Transfer No: {row.transfer_number}",
            f"From Warehouse: {from_name or '-'}",
            f"To Warehouse: {to_name or '-'}",
            f"Returned by: {row.receiver_name or '-'}",
            f"Return reason: {row.receiver_comment or '-'}",
            "Status: Returned for edit",
            "",
            "Please sign in to the warehouse system, update the transfer, and submit it again for approval.",
            "",
            "This is an automated notification from Global Technology Company.",
        ],
        "transfer_returned_by_destination_email",
    )


def notify_mr_created(row: MaterialRequisition, db: Session) -> None:
    warehouse_name = row.warehouse.name if row.warehouse else ""
    lines = [
        "Hello,",
        "",
        "A new material request has been submitted and is ready for your review in the warehouse system.",
        "",
        f"Material Request No: {row.order_number}",
        f"Requester: {row.requester_name or '-'}",
        f"Warehouse: {warehouse_name or '-'}",
        f"Site: {row.site_id or row.site_address or '-'}",
        f"Status: Pending approval",
        f"Date: {row.creation_date or local_today()}",
        "",
        "Please sign in to the warehouse system and review this request when convenient.",
        "",
        "This is an automated notification from Global Technology Company.",
    ]
    notify_mr_email(
        row,
        db,
        approval_notification_emails(db),
        f"Approval needed: Material Request {row.order_number}",
        lines,
        "mr_email",
    )


def notify_mr_approved(row: MaterialRequisition, db: Session) -> None:
    warehouse_name = row.warehouse.name if row.warehouse else ""
    approver_copy = approval_notification_emails(db)
    notify_mr_email(
        row,
        db,
        warehouse_manager_notification_emails(row, db) + approver_copy,
        f"Warehouse action needed: Material Request {row.order_number}",
        [
            "Hello,",
            "",
            "A material request has been approved and is now ready for warehouse action.",
            "",
            f"Material Request No: {row.order_number}",
            f"Requester: {row.requester_name or '-'}",
            f"Warehouse: {warehouse_name or '-'}",
            f"Site: {row.site_id or row.site_address or '-'}",
            f"Approver: {row.receiver_name or '-'}",
            f"Status: Approved",
            "",
            "Please sign in to the warehouse system and continue processing this request.",
            "",
            "This is an automated notification from Global Technology Company.",
        ],
        "mr_approved_warehouse_email",
    )
    notify_mr_email(
        row,
        db,
        requester_notification_emails(row, db) + approver_copy,
        f"Your material request {row.order_number} was approved",
        [
            "Hello,",
            "",
            "Your material request has been approved.",
            "",
            f"Material Request No: {row.order_number}",
            f"Warehouse: {warehouse_name or '-'}",
            f"Site: {row.site_id or row.site_address or '-'}",
            f"Approved by: {row.receiver_name or '-'}",
            f"Status: Approved",
            "",
            "The request is now with the warehouse team for the next step.",
            "",
            "This is an automated notification from Global Technology Company.",
        ],
        "mr_approved_requester_email",
    )


def notify_mr_rejected(row: MaterialRequisition, db: Session) -> None:
    warehouse_name = row.warehouse.name if row.warehouse else ""
    notify_mr_email(
        row,
        db,
        requester_notification_emails(row, db) + approval_notification_emails(db),
        f"Your material request {row.order_number} was rejected",
        [
            "Hello,",
            "",
            "Your material request was rejected.",
            "",
            f"Material Request No: {row.order_number}",
            f"Warehouse: {warehouse_name or '-'}",
            f"Site: {row.site_id or row.site_address or '-'}",
            f"Reviewed by: {row.receiver_name or '-'}",
            f"Comment: {row.receiver_comment or '-'}",
            f"Status: Rejected",
            "",
            "Please sign in to the warehouse system for details.",
            "",
            "This is an automated notification from Global Technology Company.",
        ],
        "mr_rejected_email",
    )


def notify_mr_returned_for_edit(row: MaterialRequisition, db: Session) -> None:
    warehouse_name = row.warehouse.name if row.warehouse else ""
    notify_mr_email(
        row,
        db,
        requester_notification_emails(row, db) + approval_notification_emails(db),
        f"Material request {row.order_number} returned for edit",
        [
            "Hello,",
            "",
            "Your material request was returned for update.",
            "",
            f"Material Request No: {row.order_number}",
            f"Warehouse: {warehouse_name or '-'}",
            f"Site: {row.site_id or row.site_address or '-'}",
            f"Return reason: {row.return_reason or row.receiver_comment or '-'}",
            f"Status: Returned for edit",
            "",
            "Please sign in to the warehouse system, update the request, and submit it again.",
            "",
            "This is an automated notification from Global Technology Company.",
        ],
        "mr_returned_email",
    )


def notify_mr_issued(row: MaterialRequisition, db: Session) -> None:
    warehouse_name = row.warehouse.name if row.warehouse else ""
    notify_mr_email(
        row,
        db,
        requester_notification_emails(row, db) + approval_notification_emails(db),
        f"Material request {row.order_number} was issued",
        [
            "Hello,",
            "",
            "A material request has been issued by the warehouse team.",
            "",
            f"Material Request No: {row.order_number}",
            f"Warehouse: {warehouse_name or '-'}",
            f"Site: {row.site_id or row.site_address or '-'}",
            f"Requester: {row.requester_name or '-'}",
            f"Approver: {row.receiver_name or '-'}",
            f"Status: Issued",
            "",
            "Please sign in to the warehouse system for details.",
            "",
            "This is an automated notification from Global Technology Company.",
        ],
        "mr_issued_email",
    )


class LoginIn(BaseModel):
    username: str
    password: str
    program: str = DEFAULT_PROGRAM


class AppUserIn(BaseModel):
    username: str
    password: str = ""
    role: Literal["Admin", "Management", "Requester", "Approval", "Warehouse Manager"]
    name: str = ""
    email: str = ""
    warehouse_name: str = ""
    program: str = DEFAULT_PROGRAM


class AppUserDeleteIn(BaseModel):
    username: str
    role: str
    program: str = DEFAULT_PROGRAM


class WarehouseIn(BaseModel):
    name: str
    location: str = ""
    program: str = DEFAULT_PROGRAM


class SiteIn(BaseModel):
    name: str
    program: str = DEFAULT_PROGRAM


class TechnicianIn(BaseModel):
    name: str
    phone: str = ""
    program: str = DEFAULT_PROGRAM


class ProductIn(BaseModel):
    program: str = DEFAULT_PROGRAM
    sku: str
    part_number: str = ""
    category: str = ""
    name: str
    item_detail: str = ""
    vendor: str = ""
    qr_code: str = ""
    unit: str = "PCS"
    tracking_type: Literal["bulk", "serialized"] = "bulk"
    min_stock: float = 0


class MaterialScanIn(BaseModel):
    code: str
    actor: str = "system"
    program: str = DEFAULT_PROGRAM


class ReceiveItemIn(BaseModel):
    product_id: int
    quantity: float = Field(gt=0)
    serial_numbers: list[str] = []


class ReceiveIn(BaseModel):
    program: str = DEFAULT_PROGRAM
    warehouse_id: int
    supplier: str = ""
    receipt_number: str = ""
    receipt_date: str = ""
    created_by: str = "system"
    items: list[ReceiveItemIn]


class InventoryReceiveIn(BaseModel):
    program: str = DEFAULT_PROGRAM
    receipt_date: str = ""
    receipt_number: str = ""
    supplier: str = ""
    warehouse_id: int
    sku: str
    part_number: str = ""
    name: str
    quantity: float = Field(gt=0)
    unit: str = "PCS"
    qr_code: str = ""
    category: str = ""
    created_by: str = "manager"


class InventoryAdjustmentIn(BaseModel):
    program: str = DEFAULT_PROGRAM
    warehouse_id: int
    sku: str
    quantity: float = Field(ge=0)
    note: str = ""
    created_by: str = "manager"


class ProductPurgeIn(BaseModel):
    actor: str = "admin"
    role: str = "Admin"
    program: str = DEFAULT_PROGRAM


class IssueItemIn(BaseModel):
    product_id: int
    quantity: float = Field(gt=0)
    serial_numbers: list[str] = []


class IssueIn(BaseModel):
    program: str = DEFAULT_PROGRAM
    warehouse_id: int
    technician_id: int
    created_by: str = "system"
    items: list[IssueItemIn]


class MaterialRequisitionItemIn(BaseModel):
    product_id: int | None = None
    part_nbr: str = ""
    model: str = ""
    description: str
    vendor: str = ""
    uom: str = "PCS"
    quantity: float = Field(gt=0)
    remark: str = ""


class MaterialRequisitionIn(BaseModel):
    program: str = DEFAULT_PROGRAM
    creation_date: str = ""
    warehouse_id: int
    entity: str = "Rollout"
    project_name: str = "FTTH"
    site_id: str = ""
    site_address: str = ""
    wo_no: str = ""
    product_domain: str = "Passive"
    team_leader: str = ""
    receiver_tel: str = ""
    request_shipment_time: str = ""
    request_arrived_site_time: str = ""
    requester_name: str = ""
    requester_title: str = ""
    requester_signature: str = ""
    requester_date: str = ""
    requester_comment: str = ""
    receiver_name: str = ""
    receiver_title: str = ""
    receiver_signature: str = ""
    receiver_date: str = ""
    receiver_comment: str = ""
    return_reason: str = ""
    created_by: str = "manager"
    issue_immediately: bool = False
    items: list[MaterialRequisitionItemIn]


class MaterialRequisitionSignatureIn(BaseModel):
    program: str = DEFAULT_PROGRAM
    role: Literal["requester", "receiver"]
    name: str = ""
    title: str = ""
    date: str = ""
    signature: str
    comment: str = ""


class MaterialRequisitionActionIn(BaseModel):
    actor: str = "manager"
    title: str = ""
    comment: str = ""
    signature: str = ""
    program: str = DEFAULT_PROGRAM


class MaterialTransferItemIn(BaseModel):
    product_id: int
    quantity: float = Field(gt=0)
    remark: str = ""


class MaterialTransferIn(BaseModel):
    program: str = DEFAULT_PROGRAM
    transfer_date: str = ""
    from_warehouse_id: int
    to_warehouse_id: int
    reference_no: str = ""
    reason: str = ""
    requester_name: str = ""
    requester_title: str = ""
    approver_name: str = ""
    approver_title: str = ""
    receiver_name: str = ""
    created_by: str = "manager"
    items: list[MaterialTransferItemIn]


class MaterialReturnItemIn(BaseModel):
    product_id: int
    quantity: float = Field(gt=0)
    condition: str = "Good"
    remark: str = ""


class MaterialReturnIn(BaseModel):
    program: str = DEFAULT_PROGRAM
    return_date: str = ""
    site_id: str = ""
    site_address: str = ""
    warehouse_id: int
    returned_by: str = ""
    received_by: str = ""
    reason: str = ""
    created_by: str = "manager"
    items: list[MaterialReturnItemIn]


def db_session(request: Request):
    program_key = normalize_program(getattr(request.state, "program", request.query_params.get("program", DEFAULT_PROGRAM)))
    db = get_sessionmaker(program_key)()
    try:
        yield db
    finally:
        db.close()


def session_token_hash(token: str) -> str:
    return hmac.new(SESSION_SECRET.encode("utf-8"), token.encode("utf-8"), hashlib.sha256).hexdigest()


def session_user_payload(user: AppUser, program: str) -> dict:
    return {
        "id": user.id,
        "program": normalize_program(program),
        "username": user.username,
        "name": user.name or user.username,
        "role": user.role,
        "warehouse_name": user.warehouse_name or "",
    }


def create_app_session(db: Session, user: AppUser, program: str) -> tuple[str, str]:
    now = datetime.now(timezone.utc)
    db.query(AppSession).filter(AppSession.expires_at <= now).delete(synchronize_session=False)
    token = secrets.token_urlsafe(48)
    csrf_token = secrets.token_urlsafe(32)
    db.add(AppSession(token_hash=session_token_hash(token), csrf_token=csrf_token, user_id=user.id, program=normalize_program(program), expires_at=now + timedelta(hours=SESSION_TTL_HOURS)))
    db.commit()
    return token, csrf_token


def clear_app_session(db: Session, token: str) -> None:
    if token:
        db.query(AppSession).filter(AppSession.token_hash == session_token_hash(token)).delete(synchronize_session=False)
        db.commit()


def current_user(request: Request) -> AppUser:
    user = getattr(request.state, "current_user", None)
    if user is None:
        raise HTTPException(status_code=401, detail="Authentication required")
    return user


def require_roles(request: Request, *roles: str) -> AppUser:
    user = current_user(request)
    if user.role.strip().lower() not in {role.strip().lower() for role in roles}:
        raise HTTPException(status_code=403, detail="You do not have permission for this action")
    return user


def request_actor(request: Request) -> str:
    user = current_user(request)
    return user.name or user.username


def request_scope_viewer(request: Request) -> str:
    user = current_user(request)
    if user.role.strip().lower() == "warehouse manager":
        return user.warehouse_name or user.name or user.username
    return request_actor(request)


def user_can_access_warehouse(user: AppUser, warehouse: Warehouse | None) -> bool:
    role = user.role.strip().lower()
    if role in {"admin", "management"}:
        return True
    return role == "warehouse manager" and warehouse is not None and warehouse_scope_matches(user.warehouse_name or user.username or user.name, warehouse.name)


def require_warehouse_access(request: Request, db: Session, warehouse_id: int, program: str) -> Warehouse:
    warehouse = require_warehouse(db, warehouse_id, program)
    if not user_can_access_warehouse(current_user(request), warehouse):
        raise HTTPException(status_code=403, detail="You do not have access to this warehouse")
    return warehouse


def allowed_warehouse_ids(request: Request, db: Session, program: str) -> list[int] | None:
    user = current_user(request)
    if user.role.strip().lower() in {"admin", "management", "approval", "requester"}:
        return None
    if user.role.strip().lower() != "warehouse manager":
        return []
    scope = user.warehouse_name or user.username or user.name
    return [
        row.id
        for row in db.query(Warehouse).filter(Warehouse.program == normalize_program(program)).all()
        if warehouse_scope_matches(scope, row.name)
    ]


def require_program_record(row, program: str, label: str):
    if row is None or normalize_program(getattr(row, "program", DEFAULT_PROGRAM)) != normalize_program(program):
        raise HTTPException(status_code=404, detail=f"{label} not found")
    return row


def rollout_records_for_session(request: Request, rows: list[dict]) -> list[dict]:
    user = current_user(request)
    if user.role.strip().lower() != "warehouse manager":
        return rows
    scope = user.warehouse_name or user.name or user.username
    return [
        row for row in rows
        if warehouse_scope_matches(scope, str(row.get("city") or row.get("City") or ""))
        or warehouse_scope_matches(scope, str(row.get("Area") or row.get("area") or ""))
    ]


def next_number(db: Session, model, prefix: str) -> str:
    number_column = next(
        (
            getattr(model, name)
            for name in ("order_number", "transfer_number", "return_number")
            if hasattr(model, name)
        ),
        None,
    )
    if number_column is None:
        raise RuntimeError(f"No number column configured for {model.__name__}")
    candidate = db.query(model).count() + 1
    while db.query(model).filter(number_column == f"{prefix}-{candidate:05d}").first():
        candidate += 1
    return f"{prefix}-{candidate:05d}"


def clear_warehouse_cache():
    WAREHOUSE_CACHE.clear()


def clear_rollout_db_cache():
    global ROLLOUT_DB_CACHE, ROLLOUT_ENTRY_ID_CACHE
    ROLLOUT_DB_CACHE = None
    ROLLOUT_ENTRY_ID_CACHE = None
    ROLLOUT_CODE_REFERENCE_CACHE.clear()


def log_audit(db: Session, action: str, entity_type: str, entity_id: str, actor: str, details: dict):
    clear_warehouse_cache()
    db.add(
        AuditLog(
            action=action,
            entity_type=entity_type,
            entity_id=str(entity_id),
            actor=actor or "system",
            details=json.dumps(details, ensure_ascii=False),
        )
    )


def stock_balance(db: Session, warehouse_id: int, product_id: int, program: str = DEFAULT_PROGRAM) -> StockBalance:
    program_key = normalize_program(program)
    row = (
        db.query(StockBalance)
        .filter(StockBalance.warehouse_id == warehouse_id, StockBalance.product_id == product_id, StockBalance.program == program_key)
        .first()
    )
    if row is None:
        row = StockBalance(program=program_key, warehouse_id=warehouse_id, product_id=product_id, quantity=0)
        db.add(row)
        db.flush()
    return row


def locked_stock_balance(db: Session, warehouse_id: int, product_id: int, program: str = DEFAULT_PROGRAM) -> StockBalance:
    program_key = normalize_program(program)
    stock_balance(db, warehouse_id, product_id, program_key)
    db.flush()
    return (
        db.query(StockBalance)
        .filter(
            StockBalance.program == program_key,
            StockBalance.warehouse_id == warehouse_id,
            StockBalance.product_id == product_id,
        )
        .with_for_update()
        .one()
    )


RESERVING_REQUISITION_STATUSES = {"pending_approval", "approved", "signed"}
RESERVING_TRANSFER_STATUSES = {"pending_approval", "approved"}


def reserved_stock_quantities(
    db: Session,
    program: str = DEFAULT_PROGRAM,
    exclude_requisition_id: int | None = None,
    exclude_transfer_id: int | None = None,
) -> dict[tuple[int, int], float]:
    """Return quantities held by open MR and outbound TR workflows."""
    program_key = normalize_program(program)
    reserved: dict[tuple[int, int], float] = {}

    requisitions = (
        db.query(MaterialRequisition)
        .options(selectinload(MaterialRequisition.items))
        .filter(
            MaterialRequisition.program == program_key,
            MaterialRequisition.status.in_(RESERVING_REQUISITION_STATUSES),
        )
        .all()
    )
    for requisition in requisitions:
        if exclude_requisition_id and requisition.id == exclude_requisition_id:
            continue
        for item in requisition.items:
            if item.product_id:
                key = (requisition.warehouse_id, item.product_id)
                reserved[key] = reserved.get(key, 0) + float(item.quantity or 0)

    transfers = (
        db.query(MaterialTransfer)
        .options(selectinload(MaterialTransfer.items))
        .filter(
            MaterialTransfer.program == program_key,
            MaterialTransfer.status.in_(RESERVING_TRANSFER_STATUSES),
        )
        .all()
    )
    for transfer in transfers:
        if exclude_transfer_id and transfer.id == exclude_transfer_id:
            continue
        for item in transfer.items:
            key = (transfer.from_warehouse_id, item.product_id)
            reserved[key] = reserved.get(key, 0) + float(item.quantity or 0)

    return reserved


def validate_reservable_stock(
    db: Session,
    warehouse_id: int,
    items: list,
    program: str = DEFAULT_PROGRAM,
    exclude_requisition_id: int | None = None,
    exclude_transfer_id: int | None = None,
) -> None:
    """Prevent a new or edited workflow from consuming stock already reserved elsewhere."""
    program_key = normalize_program(program)
    requested: dict[int, float] = {}
    for item in items:
        product_id = int(getattr(item, "product_id", 0) or 0)
        if product_id:
            requested[product_id] = requested.get(product_id, 0) + float(getattr(item, "quantity", 0) or 0)

    # Lock each affected balance in a stable order so simultaneous requests cannot over-reserve it.
    balances: dict[int, StockBalance] = {}
    for product_id in sorted(requested):
        balances[product_id] = locked_stock_balance(db, warehouse_id, product_id, program_key)

    reserved = reserved_stock_quantities(
        db,
        program_key,
        exclude_requisition_id=exclude_requisition_id,
        exclude_transfer_id=exclude_transfer_id,
    )
    for product_id, requested_qty in requested.items():
        balance = balances[product_id]
        held = float(reserved.get((warehouse_id, product_id), 0) or 0)
        available = float(balance.quantity or 0) - held
        if requested_qty > available + 1e-9:
            product = require_product(db, product_id, program_key)
            raise HTTPException(
                status_code=400,
                detail=(
                    f"Insufficient available stock for {product_display_name(product)}. "
                    f"Requested {requested_qty:g}, stock {float(balance.quantity or 0):g}, "
                    f"reserved {held:g}, available {max(available, 0):g}."
                ),
            )


def technician_balance(db: Session, technician_id: int, product_id: int, program: str = DEFAULT_PROGRAM) -> TechnicianBalance:
    program_key = normalize_program(program)
    row = (
        db.query(TechnicianBalance)
        .filter(TechnicianBalance.technician_id == technician_id, TechnicianBalance.product_id == product_id, TechnicianBalance.program == program_key)
        .first()
    )
    if row is None:
        row = TechnicianBalance(program=program_key, technician_id=technician_id, product_id=product_id, quantity=0)
        db.add(row)
        db.flush()
    return row


def require_product(db: Session, product_id: int, program: str = DEFAULT_PROGRAM) -> Product:
    product = db.get(Product, product_id)
    if product is None or normalize_program(getattr(product, "program", DEFAULT_PROGRAM)) != normalize_program(program):
        raise HTTPException(status_code=404, detail=f"Product {product_id} not found")
    return product


def require_warehouse(db: Session, warehouse_id: int, program: str = DEFAULT_PROGRAM) -> Warehouse:
    warehouse = db.get(Warehouse, warehouse_id)
    if warehouse is None or normalize_program(getattr(warehouse, "program", DEFAULT_PROGRAM)) != normalize_program(program):
        raise HTTPException(status_code=404, detail=f"Warehouse {warehouse_id} not found")
    return warehouse


def require_technician(db: Session, technician_id: int, program: str = DEFAULT_PROGRAM) -> Technician:
    technician = db.get(Technician, technician_id)
    if technician is None or normalize_program(getattr(technician, "program", DEFAULT_PROGRAM)) != normalize_program(program):
        raise HTTPException(status_code=404, detail=f"Technician {technician_id} not found")
    return technician


def validate_serial_count(product: Product, quantity: float, serial_numbers: list[str]):
    if product.tracking_type != "serialized":
        return
    if quantity != int(quantity):
        raise HTTPException(status_code=400, detail="Serialized item quantity must be a whole number")
    if len(serial_numbers) != int(quantity):
        raise HTTPException(status_code=400, detail=f"{product.sku} requires one serial number per unit")
    if len(set(serial_numbers)) != len(serial_numbers):
        raise HTTPException(status_code=400, detail=f"{product.sku} has duplicate serial numbers in the request")


def row_to_record(row: RolloutRecord) -> dict:
    return {
        "ID": row.record_id,
        "Date": row.date,
        "Supervisor Name": row.supervisor_name,
        "team leader": row.team_leader,
        "Area": row.area,
        "city": row.city,
        "Activity": row.activity,
        "item": row.item,
        "material type": row.material_type,
        "mount type": row.mount_type,
        "item serial": row.item_serial,
        "planed quantity": row.planned_quantity,
        "actual": row.actual,
        "stock remaining": row.stock_remaining,
        "staus": row.status,
        "laser": row.laser,
        "acceptance": row.acceptance,
        "scan": row.scan,
        "labeling": row.labeling,
        "Related to XBOX": row.related_to_xbox,
        "entry time": row.entry_time,
        "cable code": row.cable_code,
        "box code": row.box_code,
        "OLT": row.olt,
        "Cable route": row.cable_route,
        "Notes": row.notes,
    }


def normalize_rollout_row(data: dict) -> dict:
    return {
        "ID": str(first_value(data, "ID", "id", default="") or ""),
        "Date": str(first_value(data, "Date", "date", default="") or ""),
        "Supervisor Name": str(first_value(data, "Supervisor Name", "supervisor_name", default="") or ""),
        "team leader": str(first_value(data, "team leader", "Team Leader", "team_leader", default="") or ""),
        "Area": str(first_value(data, "Area", "area", default="") or ""),
        "city": str(first_value(data, "city", "City", default="") or ""),
        "Activity": str(first_value(data, "Activity", "activity", default="") or ""),
        "Related to XBOX": str(first_value(data, "Related to XBOX", "related to xbox", "related_to_xbox", default="") or ""),
        "item": str(first_value(data, "item", "Item", default="") or ""),
        "material type": str(first_value(data, "material type", "Material Type", "material_type", default="") or ""),
        "mount type": str(first_value(data, "mount type", "Mount Type", "mount_type", default="") or ""),
        "item serial": str(first_value(data, "item serial", "Item Serial", "item_serial", default="") or ""),
        "planed quantity": safe_float(first_value(data, "planed quantity", "planned quantity", "planned_quantity", default=0)),
        "actual": safe_float(first_value(data, "actual", "Actual", default=0)),
        "stock remaining": safe_float(first_value(data, "stock remaining", "Stock Remaining", "stock_remaining", default=0)),
        "staus": str(first_value(data, "staus", "status", "Status", default="") or ""),
        "laser": str(first_value(data, "laser", "Laser", default="") or ""),
        "acceptance": str(first_value(data, "acceptance", "Acceptance", default="") or ""),
        "scan": str(first_value(data, "scan", "Scan", default="") or ""),
        "labeling": str(first_value(data, "labeling", "Labeling", default="") or ""),
        "entry time": str(first_value(data, "entry time", "Entry Time", "entry_time", default="") or ""),
        "cable code": str(first_value(data, "cable code", "Cable Code", "cable_code", default="") or ""),
        "box code": str(first_value(data, "box code", "Box Code", "box_code", default="") or ""),
        "OLT": str(first_value(data, "OLT", "olt", default="") or ""),
        "Cable route": str(first_value(data, "Cable route", "Cable Route", "cable_route", default="") or ""),
        "Notes": str(first_value(data, "Notes", "notes", default="") or ""),
    }


def first_value(data: dict, *keys: str, default=""):
    for key in keys:
        if key in data and data.get(key) is not None:
            return data.get(key)
    lowered = {str(k).strip().lower(): v for k, v in data.items()}
    for key in keys:
        value = lowered.get(key.strip().lower())
        if value is not None:
            return value
    return default


def safe_float(value) -> float:
    if value is None or value == "":
        return 0
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).replace(",", "").strip()
    if not text:
        return 0
    try:
        return float(text)
    except ValueError:
        return 0


def rollout_csv_urls() -> list[tuple[str, str]]:
    live_url = (os.getenv("ROLLOUT_DAILY_PROGRESS_LIVE_CSV_URL") or DEFAULT_ROLLOUT_DAILY_PROGRESS_LIVE_CSV_URL).strip()
    published_url = (os.getenv("ROLLOUT_DAILY_PROGRESS_CSV_URL") or DEFAULT_ROLLOUT_DAILY_PROGRESS_CSV_URL).strip()
    urls = []
    # Prefer the live gviz endpoint so a stale Publish-to-web snapshot cannot
    # silently replace the current sheet with an older, shorter export.
    canonical_live_url = DEFAULT_ROLLOUT_DAILY_PROGRESS_LIVE_CSV_URL.strip()
    if canonical_live_url:
        urls.append(("google_live_csv", canonical_live_url))
    if live_url and live_url not in {canonical_live_url, published_url}:
        urls.append(("google_live_csv_configured", live_url))
    if published_url and published_url != live_url:
        urls.append(("google_published_csv", published_url))
    return urls


def add_cache_buster(url: str) -> str:
    parts = urllib.parse.urlsplit(url)
    query = urllib.parse.parse_qsl(parts.query, keep_blank_values=True)
    query = [(key, value) for key, value in query if key != "_"]
    query.append(("_", str(int(time.time() * 1000))))
    return urllib.parse.urlunsplit((parts.scheme, parts.netloc, parts.path, urllib.parse.urlencode(query), parts.fragment))


def read_rollout_daily_progress_url(url: str, force: bool = False) -> list[dict]:
    fetch_url = add_cache_buster(url) if force else url
    request = urllib.request.Request(
        fetch_url,
        headers={
            "User-Agent": "warehouse-rollout-reader/1.0",
            "Cache-Control": "no-cache",
            "Pragma": "no-cache",
        },
    )
    with urllib.request.urlopen(request, timeout=20) as response:
        raw = response.read()
    text = raw.decode("utf-8-sig")
    rows = [normalize_rollout_row(row) for row in csv.DictReader(io.StringIO(text))]
    return [row for row in rows if any(str(value or "").strip() for value in row.values())]


def fetch_rollout_daily_progress_csv(force: bool = False) -> tuple[list[dict], str]:
    global ROLLOUT_CSV_CACHE
    urls = rollout_csv_urls()
    if not urls:
        return [], "none"
    if not force and ROLLOUT_CSV_CACHE and time.monotonic() - ROLLOUT_CSV_CACHE[0] < ROLLOUT_CSV_CACHE_TTL:
        return ROLLOUT_CSV_CACHE[1], ROLLOUT_CSV_CACHE[2]
    for source, url in urls:
        try:
            rows = read_rollout_daily_progress_url(url, force=force)
            if rows:
                ROLLOUT_CSV_CACHE = (time.monotonic(), rows, source)
                return rows, source
        except Exception:
            continue
    if ROLLOUT_CSV_CACHE:
        return ROLLOUT_CSV_CACHE[1], ROLLOUT_CSV_CACHE[2]
    return [], "none"


def db_rollout_records(db: Session) -> list[dict]:
    global ROLLOUT_DB_CACHE
    now = time.monotonic()
    if ROLLOUT_DB_CACHE and now - ROLLOUT_DB_CACHE[0] < ROLLOUT_DB_CACHE_TTL:
        return ROLLOUT_DB_CACHE[1]
    records = [row_to_record(row) for row in db.query(RolloutRecord).order_by(RolloutRecord.id.asc()).all()]
    ROLLOUT_DB_CACHE = (now, records)
    return records


def stable_rollout_record_id(data: dict) -> str:
    source_id = str(first_value(data, "ID", "id", default="") or "").strip()
    if source_id:
        return source_id
    payload = json.dumps(data, ensure_ascii=False, sort_keys=True, separators=(",", ":"))
    digest = hashlib.sha256(payload.encode("utf-8")).hexdigest()[:24]
    return f"RDP-SOURCE-{digest}"


def sync_rollout_daily_progress(db: Session, force: bool = False) -> tuple[list[dict], str]:
    # Supabase is the sole source of truth. The legacy Google Sheet is no
    # longer consulted, including for forced refreshes from the UI.
    return db_rollout_records(db), "database"


def rollout_daily_progress_records(db: Session, force: bool = False) -> tuple[list[dict], str]:
    return sync_rollout_daily_progress(db, force=force)


def rollout_norm(value) -> str:
    return re.sub(r"[^a-z0-9]+", "", str(value or "").strip().lower())


def rollout_area_key(value) -> str:
    key = rollout_norm(value)
    return "hayalandaluszone2" if key in {"hayalandalus", "hayandaluszone2", "hayalandaluszone2"} else key


def rollout_area_label(value) -> str:
    return "Hay Al Andalus Zone 2" if rollout_area_key(value) == "hayalandaluszone2" else str(value or "").strip()


def rollout_xbox_key(value) -> str:
    text_value = str(value or "").upper()
    match = re.search(r"X\s*-?\s*BOX\s*0*(\d+)|X\s*0*(\d+)", text_value)
    if match:
        return f"X{int(match.group(1) or match.group(2))}"
    return re.sub(r"[^A-Z0-9]+", "", text_value)


def rollout_code_key(value) -> str:
    text_value = str(value or "").upper().strip()
    text_value = re.sub(r"\bH0+(\d+)", r"H\1", text_value)
    text_value = re.sub(r"\bL0+(\d+)", r"L\1", text_value)
    text_value = re.sub(r"\bS0+(\d+)", r"S\1", text_value)
    text_value = re.sub(r"\bX0+(\d+)", r"X\1", text_value)
    return re.sub(r"[^A-Z0-9]+", "", text_value)


def rollout_entry_mode(data: dict) -> str:
    text_value = " ".join(
        str(first_value(data, "code_type", "type", "item", "Item", "material type", "Material Type", "material_type", default="") or "")
        for _ in [0]
    )
    norm = rollout_norm(text_value)
    if "accessor" in norm or "pigtail" in norm:
        return "accessory"
    if "cable" in norm:
        return "cable"
    if any(token in norm for token in ["box", "xbox", "hub", "sub", "end"]):
        return "box"
    if str(first_value(data, "cable code", "Cable Code", "cable_code", default="") or "").strip():
        return "cable"
    if str(first_value(data, "box code", "Box Code", "box_code", default="") or "").strip():
        return "box"
    return ""


def load_fiber_map_reference(db: Session | None = None, program: str = DEFAULT_PROGRAM) -> dict:
    data: dict = {"boxes": [], "routes": [], "area_plans": [], "schematics": []}
    try:
        with open(FIBER_MAP_REFERENCE_PATH, "r", encoding="utf-8") as handle:
            source = json.load(handle)
            if isinstance(source, dict):
                data["boxes"] = list(source.get("boxes") or [])
                data["routes"] = list(source.get("routes") or [])
    except FileNotFoundError:
        logger.warning("Fiber map reference not found: %s", FIBER_MAP_REFERENCE_PATH)
    except Exception:
        logger.exception("Could not read fiber map reference")
    if db is None:
        return data
    for saved_area in db.query(FiberMapArea).filter(FiberMapArea.program == normalize_program(program)).all():
        try:
            payload = json.loads(saved_area.design_data or "{}")
        except (TypeError, ValueError):
            logger.warning("Ignoring unreadable saved fiber map for %s", saved_area.area)
            continue
        data["boxes"].extend(payload.get("boxes") or [])
        data["routes"].extend(payload.get("routes") or [])
        data["area_plans"].append(
            {
                "city": saved_area.city,
                "area": saved_area.area,
                "start": saved_area.start_date,
                "end": saved_area.end_date,
                "targetSubEndBox": len(payload.get("boxes") or []),
                "targetHubBox": len({(rollout_xbox_key(row.get("Related to XBOX")), str(row.get("Hub") or "").strip()) for row in payload.get("boxes") or [] if row.get("Hub")}),
                "targetXbox": len({rollout_xbox_key(row.get("Related to XBOX")) for row in payload.get("boxes") or [] if row.get("Related to XBOX")}),
                "targetCableMeters": sum(safe_float(row.get("Cable length m")) for row in payload.get("boxes") or [])
                + sum(safe_float(row.get("Cable length m")) for row in payload.get("routes") or []),
                "targetUsers": saved_area.target_users,
                "dynamic": True,
            }
        )
    data["schematics"] = [
        {"area": row.area, "xbox": row.xbox, "sheet_name": row.sheet_name}
        for row in db.query(FiberMapSchematic)
        .filter(FiberMapSchematic.program == normalize_program(program))
        .order_by(FiberMapSchematic.area, FiberMapSchematic.xbox, FiberMapSchematic.sheet_name)
        .all()
    ]
    return data


def rollout_code_reference_rows(db: Session | None = None, program: str = DEFAULT_PROGRAM) -> list[dict]:
    program_key = normalize_program(program)
    if program_key in ROLLOUT_CODE_REFERENCE_CACHE:
        return ROLLOUT_CODE_REFERENCE_CACHE[program_key]

    ref = load_fiber_map_reference(db, program_key)
    rows: list[dict] = []
    seen: set[tuple[str, str, str, str]] = set()

    def add(row: dict, code: str, code_type: str, source: str, material_override: str = ""):
        if not code:
            return
        area = rollout_area_label(first_value(row, "Zone", "zone", "Area", "area", default=""))
        city = str(first_value(row, "City", "city", default="") or "").strip()
        xbox = str(first_value(row, "Related to XBOX", "related to xbox", "XBOX", "xbox", default="") or "").strip()
        key = (rollout_area_key(area), rollout_xbox_key(xbox), rollout_code_key(code), code_type)
        if key in seen:
            return
        seen.add(key)
        box_type = str(first_value(row, "Box type", "box type", default="") or "").strip()
        material = material_override or str(first_value(row, "Material type", "material type", "Item", "item", default="") or "").strip()
        rows.append(
            {
                "city": city,
                "area": area,
                "xbox": rollout_xbox_key(xbox) or xbox,
                "xbox_label": xbox,
                "code": str(code).strip(),
                "type": code_type,
                "source": source,
                "material_type": material,
                "box_type": box_type,
                "cable_length_m": safe_float(first_value(row, "Cable length m", "Real length m", "Used length m", default=0)),
                "cable_route": str(first_value(row, "Cable route", "Route", "Route type", default="") or "").strip(),
            }
        )

    for row in ref.get("boxes") or []:
        code = str(first_value(row, "Box code", "box code", default="") or "").strip()
        add(row, code, "box", "box", str(first_value(row, "Box type", "box type", default="") or "").strip())
        add(row, code, "cable", "drop")
        # An XBOX is the root node for its map rows and has no dedicated box
        # row. Add it explicitly so teams can record its installation.
        xbox_code = str(first_value(row, "Related to XBOX", "related to xbox", "XBOX", "xbox", default="") or "").strip()
        if xbox_code:
            xbox_row = {**row, "Box type": "XBOX", "Material type": "XBOX"}
            add(xbox_row, xbox_code, "box", "xbox", "XBOX")
        # Hubs are the parent node of the SUB/END box rows in the map data;
        # they do not have their own Box code row. Expose one HubBox code for
        # each area/XBOX so it can be selected during field entry.
        hub_code = str(first_value(row, "Hub", "hub", default="") or "").strip()
        if hub_code:
            hub_row = {**row, "Box type": "HUB BOX", "Material type": "HubBox"}
            add(hub_row, hub_code, "box", "hub", "HubBox")
    for row in ref.get("routes") or []:
        code = str(first_value(row, "Route code", "route code", "Cable code", "cable code", default="") or "").strip()
        add(row, code, "cable", "route")
    ROLLOUT_CODE_REFERENCE_CACHE[program_key] = rows
    return rows


def rollout_dashboard_summary_metric_rows(rows: list[dict], db: Session | None = None, program: str = DEFAULT_PROGRAM) -> list[dict]:
    """Keep Maqawba dashboard totals aligned with the active fiber-map design."""
    if not any(rollout_area_key(first_value(row, "Area", "area", default="")) == "maqawba" for row in rows):
        return rows

    refs = rollout_code_reference_rows(db, program)
    ref_sets = {
        "box": set(),
        "hub": set(),
        "xbox": set(),
        "drop_cable": set(),
        "route_cable": set(),
    }
    for ref in refs:
        if rollout_area_key(ref.get("area")) != "maqawba":
            continue
        xbox = rollout_xbox_key(ref.get("xbox"))
        code = rollout_code_key(ref.get("code"))
        if not xbox or not code:
            continue
        key = (xbox, code)
        source = rollout_norm(ref.get("source"))
        code_type = rollout_norm(ref.get("type"))
        if code_type == "box" and source == "hub":
            ref_sets["hub"].add(key)
        elif code_type == "box" and source == "xbox":
            ref_sets["xbox"].add(key)
        elif code_type == "box":
            ref_sets["box"].add(key)
        elif code_type == "cable" and source == "route":
            ref_sets["route_cable"].add(key)
        elif code_type == "cable":
            ref_sets["drop_cable"].add(key)

    def keep(row: dict) -> bool:
        if rollout_area_key(first_value(row, "Area", "area", default="")) != "maqawba":
            return True
        material = rollout_norm(f"{first_value(row, 'item', default='')} {first_value(row, 'material type', default='')}")
        xbox = rollout_xbox_key(first_value(row, "Related to XBOX", "related_to_xbox", "XBOX", "xbox", default=""))
        box_code = rollout_code_key(first_value(row, "box code", "box_code", default=""))
        cable_code = rollout_code_key(first_value(row, "cable code", "cable_code", default=""))
        if "hubbox" in material:
            return (xbox, box_code) in ref_sets["hub"]
        if "xbox" in material:
            return (xbox, box_code) in ref_sets["xbox"] and box_code == rollout_code_key(xbox)
        if "subbox" in material or "endbox" in material:
            return (xbox, box_code) in ref_sets["box"]
        if "singlecoredistributioncable" in material or "distributioncable" in material:
            return (xbox, cable_code) in ref_sets["drop_cable"]
        if "4corecable" in material:
            return (xbox, cable_code) in ref_sets["route_cable"]
        return True

    return [row for row in rows if keep(row)]


def rollout_reference_matches(area: str, xbox: str, code: str, code_type: str, db: Session | None = None, program: str = DEFAULT_PROGRAM) -> list[dict]:
    area_key = rollout_area_key(area)
    xbox_key = rollout_xbox_key(xbox)
    code_key = rollout_code_key(code)
    return [
        row
        for row in rollout_code_reference_rows(db, program)
        if row["type"] == code_type
        and rollout_area_key(row.get("area")) == area_key
        and rollout_xbox_key(row.get("xbox")) == xbox_key
        and rollout_code_key(row.get("code")) == code_key
    ]


def area_builder_header_key(value) -> str:
    return re.sub(r"[^a-z0-9]+", "", str(value or "").strip().lower())


def area_builder_sheet_rows(workbook, sheet_name: str, required_headers: set[str]) -> list[dict]:
    if sheet_name not in workbook.sheetnames:
        raise HTTPException(status_code=400, detail=f"Workbook is missing the '{sheet_name}' sheet")
    sheet = workbook[sheet_name]
    header_row = None
    header_map: dict[str, int] = {}
    for index, row in enumerate(sheet.iter_rows(values_only=True), start=1):
        candidate = {area_builder_header_key(value): column for column, value in enumerate(row) if area_builder_header_key(value)}
        if required_headers.issubset(candidate):
            header_row = index
            header_map = candidate
            break
        if index >= 25:
            break
    if header_row is None:
        raise HTTPException(status_code=400, detail=f"Could not find the required columns in '{sheet_name}'")

    rows: list[dict] = []
    for values in sheet.iter_rows(min_row=header_row + 1, values_only=True):
        item = {
            header: values[column] if column < len(values) else ""
            for header, column in header_map.items()
        }
        if any(str(value or "").strip() for value in item.values()):
            rows.append(item)
    return rows


def area_builder_hub_code(value) -> str:
    match = re.search(r"\b(H\s*\d+)\b", str(value or "").upper())
    return f"H{int(re.search(r'\d+', match.group(1)).group())}" if match else ""


def area_builder_xbox_code(value) -> str:
    match = re.search(r"\bX\s*-?\s*BOX\s*0*(\d+)\b|\bX\s*0*(\d+)\b", str(value or "").upper())
    return f"X{int(match.group(1) or match.group(2))}" if match else ""


def area_builder_number(value) -> int:
    return max(0, int(round(safe_float(value))))


def area_builder_line_number(value) -> int:
    match = re.search(r"\bL\s*(\d+)\b", str(value or ""), flags=re.I)
    return int(match.group(1)) if match else area_builder_number(value)


def area_builder_sheet_parts(workbook) -> dict[tuple[str, str], str]:
    """Use optional schematic tab names as the explicit visual grouping."""
    parts: dict[tuple[str, str], str] = {}
    for sheet_name in workbook.sheetnames:
        match = re.match(r"^\s*X\s*-?\s*(\d+)\s+H\s*(\d+)(?:\s*-\s*H\s*(\d+))?\s*$", sheet_name, flags=re.I)
        if not match:
            continue
        xbox = f"X{int(match.group(1))}"
        hubs = [f"H{int(match.group(2))}"]
        if match.group(3):
            hubs.append(f"H{int(match.group(3))}")
        label = f"{xbox} {'-'.join(hubs)}"
        for hub in hubs:
            parts[(xbox, hub)] = label
    return parts


def area_builder_schematic_xbox(sheet_name: str) -> str:
    match = re.match(r"^\s*X\s*-?\s*(\d+)\s+H\s*\d+(?:\s*-\s*H\s*\d+)?\s*$", sheet_name, flags=re.I)
    return f"X{int(match.group(1))}" if match else ""


def area_builder_zip_target(source_path: str, target: str) -> str:
    return posixpath.normpath(posixpath.join(posixpath.dirname(source_path), target)).lstrip("/")


def area_builder_relationships(archive: zipfile.ZipFile, part_path: str) -> dict[str, str]:
    rel_path = posixpath.join(posixpath.dirname(part_path), "_rels", f"{posixpath.basename(part_path)}.rels")
    if rel_path not in archive.namelist():
        return {}
    root = ElementTree.fromstring(archive.read(rel_path))
    rel_ns = "{http://schemas.openxmlformats.org/package/2006/relationships}"
    return {
        item.attrib.get("Id", ""): area_builder_zip_target(part_path, item.attrib.get("Target", ""))
        for item in root.findall(f"{rel_ns}Relationship")
        if item.attrib.get("Id") and item.attrib.get("Target")
    }


def area_builder_schematic_content_type(path: str) -> str:
    extension = posixpath.splitext(path.lower())[1]
    return {".png": "image/png", ".jpg": "image/jpeg", ".jpeg": "image/jpeg", ".gif": "image/gif"}.get(extension, "image/png")


def extract_area_builder_schematics(contents: bytes) -> list[dict]:
    """Extract the original diagram image from each named schematic tab, if present."""
    try:
        with zipfile.ZipFile(io.BytesIO(contents)) as archive:
            names = set(archive.namelist())
            workbook_path = "xl/workbook.xml"
            if workbook_path not in names:
                return []
            workbook_root = ElementTree.fromstring(archive.read(workbook_path))
            workbook_rels = area_builder_relationships(archive, workbook_path)
            main_ns = "{http://schemas.openxmlformats.org/spreadsheetml/2006/main}"
            rel_ns = "{http://schemas.openxmlformats.org/officeDocument/2006/relationships}"
            drawing_main_ns = "{http://schemas.openxmlformats.org/drawingml/2006/main}"
            found: list[dict] = []
            for sheet in workbook_root.findall(f".//{main_ns}sheet"):
                sheet_name = str(sheet.attrib.get("name") or "").strip()
                xbox = area_builder_schematic_xbox(sheet_name)
                sheet_relation = sheet.attrib.get(f"{rel_ns}id", "")
                sheet_path = workbook_rels.get(sheet_relation, "")
                if not xbox or not sheet_path or sheet_path not in names:
                    continue
                sheet_rels = area_builder_relationships(archive, sheet_path)
                drawing_path = next((path for path in sheet_rels.values() if path.startswith("xl/drawings/")), "")
                if not drawing_path or drawing_path not in names:
                    continue
                drawing_rels = area_builder_relationships(archive, drawing_path)
                drawing_root = ElementTree.fromstring(archive.read(drawing_path))
                embeds = [
                    item.attrib.get(f"{rel_ns}embed", "")
                    for item in drawing_root.findall(f".//{drawing_main_ns}blip")
                ]
                image_path = next((drawing_rels.get(embed, "") for embed in embeds if drawing_rels.get(embed, "") in names), "")
                if not image_path:
                    continue
                found.append(
                    {
                        "sheet_name": sheet_name,
                        "xbox": xbox,
                        "content_type": area_builder_schematic_content_type(image_path),
                        "image_data": archive.read(image_path),
                    }
                )
            return found
    except (OSError, ValueError, zipfile.BadZipFile, ElementTree.ParseError):
        logger.warning("Workbook diagrams could not be extracted; using the interactive map preview")
        return []


def build_area_map_from_workbook(contents: bytes, area: str, city: str) -> dict:
    if not contents:
        raise HTTPException(status_code=400, detail="Choose an Excel workbook first")
    if len(contents) > 5 * 1024 * 1024:
        raise HTTPException(status_code=400, detail="Workbook must be 5 MB or smaller")
    try:
        workbook = load_workbook(io.BytesIO(contents), data_only=True, read_only=True)
    except Exception as exc:
        raise HTTPException(status_code=400, detail="The workbook could not be read. Upload the approved .xlsx layout.") from exc

    hub_rows = area_builder_sheet_rows(
        workbook,
        "Hub Index",
        {"xbox", "hub", "from", "tohub", "qlcactualm", "qlcpreconm"},
    )
    cable_rows = area_builder_sheet_rows(
        workbook,
        "Cable Data",
        {"subbox", "hub", "xbox", "line", "sequence", "actuallengthm", "preconlengthm", "splitter"},
    )

    boxes: list[dict] = []
    routes: list[dict] = []
    sheet_parts = area_builder_sheet_parts(workbook)
    seen_boxes: set[tuple[str, str]] = set()
    seen_routes: set[tuple[str, str]] = set()
    hubs: set[tuple[str, str]] = set()
    xboxes: set[str] = set()
    distribution_meters = 0
    core_meters = 0

    for row in cable_rows:
        xbox = area_builder_xbox_code(row.get("xbox"))
        hub = area_builder_hub_code(row.get("hub"))
        line = area_builder_line_number(row.get("line"))
        sequence = area_builder_number(row.get("sequence"))
        planned = area_builder_number(row.get("preconlengthm"))
        actual = area_builder_number(row.get("actuallengthm"))
        splitter = str(row.get("splitter") or "").replace(" ", "")
        if not xbox or not hub or line < 1 or sequence < 1 or planned < 1:
            raise HTTPException(status_code=400, detail="Cable Data contains a row without XBOX, HUB, line, sequence, or planned length")
        code = f"{hub}-L{line}-S{sequence}"
        box_key = (xbox, code)
        if box_key in seen_boxes:
            raise HTTPException(status_code=400, detail=f"Cable Data contains the code {code} more than once for {xbox}")
        seen_boxes.add(box_key)
        is_end = "1:8" in splitter or "1x8" in splitter.lower()
        boxes.append(
            {
                "Area": area,
                "City": city,
                "Zone": area,
                "Related to XBOX": xbox,
                "XBOX": xbox,
                "Part": sheet_parts.get((xbox, hub), f"Part{((int(re.search(r'\d+', hub).group()) - 1) // 2) + 1:02d}"),
                "Hub": hub,
                "Line": line,
                "Splitter": sequence,
                "Box code": code,
                "Box type": "END BOX" if is_end else "SUB BOX",
                "Real length m": actual,
                "Cable length m": planned,
                "Material type": f"Single-Core Distribution Cable_{planned}m",
            }
        )
        hubs.add((xbox, hub))
        xboxes.add(xbox)
        distribution_meters += planned

    for row in hub_rows:
        xbox = area_builder_xbox_code(row.get("xbox"))
        # Each Hub Index row describes the cable that ends at its own Hub.
        # "To Hub" is a topology hint for the next node, not this cable's end.
        target = area_builder_hub_code(row.get("hub"))
        source_value = str(row.get("from") or "").strip()
        # A value such as "X1-H1" identifies H1 as the preceding hub. Prefer
        # that hub over the XBOX prefix so H1-to-H2 remains connected.
        source = area_builder_hub_code(source_value) or area_builder_xbox_code(source_value)
        planned = area_builder_number(row.get("qlcpreconm"))
        actual = area_builder_number(row.get("qlcactualm"))
        if not xbox or not target or not source or planned < 1:
            raise HTTPException(status_code=400, detail="Hub Index contains a row without XBOX, From, To Hub, or planned length")
        route_code = f"{source}-{target}"
        route_key = (xbox, route_code)
        if route_key in seen_routes:
            raise HTTPException(status_code=400, detail=f"Hub Index contains the route {route_code} more than once for {xbox}")
        seen_routes.add(route_key)
        routes.append(
            {
                "Area": area,
                "City": city,
                "Zone": area,
                "Related to XBOX": xbox,
                "XBOX": xbox,
                "Part": sheet_parts.get((xbox, target), f"Part{((int(re.search(r'\d+', target).group()) - 1) // 2) + 1:02d}"),
                "Route code": route_code,
                "Route from": source,
                "Route to": target,
                "Real length m": actual,
                "Cable length m": planned,
                "Material type": f"4-coreCable_{planned}m",
            }
        )
        xboxes.add(xbox)
        core_meters += planned

    if not boxes or not routes:
        raise HTTPException(status_code=400, detail="The workbook must contain both Cable Data and Hub Index records")
    return {
        "boxes": boxes,
        "routes": routes,
        "summary": {
            "xboxes": len(xboxes),
            "hubs": len(hubs),
            "sub_end_boxes": len(boxes),
            "distribution_meters": distribution_meters,
            "core_meters": core_meters,
            "total_cable_meters": distribution_meters + core_meters,
        },
    }


def area_builder_metadata(area: str, city: str, start_date: str, end_date: str, target_users: int) -> tuple[str, str, str, str, int]:
    area_name = str(area or "").strip()
    city_name = str(city or "").strip()
    if not area_name or not city_name:
        raise HTTPException(status_code=400, detail="Enter both Area name and City")
    try:
        start = datetime.fromisoformat(str(start_date or "")).date()
        end = datetime.fromisoformat(str(end_date or "")).date()
    except ValueError as exc:
        raise HTTPException(status_code=400, detail="Enter valid start and end dates") from exc
    if end < start:
        raise HTTPException(status_code=400, detail="End date cannot be before start date")
    return area_name, city_name, start.isoformat(), end.isoformat(), max(0, int(target_users or 0))


@app.get("/api/warehouse/fiber-map-reference")
def fiber_map_reference(request: Request, db: Session = Depends(db_session)):
    program_key = normalize_program(getattr(request.state, "program", DEFAULT_PROGRAM))
    return {"success": True, **load_fiber_map_reference(db, program_key)}


@app.get("/api/warehouse/fiber-map-schematic")
def fiber_map_schematic(
    request: Request,
    area: str = "",
    sheet: str = "",
    db: Session = Depends(db_session),
):
    current_user(request)
    program_key = normalize_program(getattr(request.state, "program", DEFAULT_PROGRAM))
    row = (
        db.query(FiberMapSchematic)
        .filter(
            FiberMapSchematic.program == program_key,
            func.lower(FiberMapSchematic.area) == str(area or "").strip().lower(),
            FiberMapSchematic.sheet_name == str(sheet or "").strip(),
        )
        .first()
    )
    if row is None:
        raise HTTPException(status_code=404, detail="Map schematic not found")
    return Response(
        content=row.image_data,
        media_type=row.content_type or "image/png",
        headers={"Cache-Control": "private, max-age=300"},
    )


@app.post("/api/warehouse/area-builder/preview")
async def preview_area_builder(
    request: Request,
    area: str = Form(""),
    city: str = Form(""),
    start_date: str = Form(""),
    end_date: str = Form(""),
    target_users: int = Form(0),
    workbook: UploadFile = File(...),
):
    require_roles(request, "Admin")
    if is_single_ran(getattr(request.state, "program", DEFAULT_PROGRAM)):
        raise HTTPException(status_code=400, detail="Area Builder is available for FTTH only")
    area_name, city_name, start, end, users = area_builder_metadata(area, city, start_date, end_date, target_users)
    if not str(workbook.filename or "").lower().endswith(".xlsx"):
        raise HTTPException(status_code=400, detail="Upload an .xlsx workbook")
    contents = await workbook.read()
    payload = build_area_map_from_workbook(contents, area_name, city_name)
    schematics = extract_area_builder_schematics(contents)
    return {
        "success": True,
        "area": {"name": area_name, "city": city_name, "start_date": start, "end_date": end, "target_users": users},
        **payload,
        "schematics": [
            {
                "sheet_name": row["sheet_name"],
                "xbox": row["xbox"],
                "image_url": f"data:{row['content_type']};base64,{base64.b64encode(row['image_data']).decode('ascii')}",
            }
            for row in schematics
        ],
        "preview": {"boxes": payload["boxes"][:12], "routes": payload["routes"][:12]},
    }


@app.post("/api/warehouse/area-builder/publish")
async def publish_area_builder(
    request: Request,
    area: str = Form(""),
    city: str = Form(""),
    start_date: str = Form(""),
    end_date: str = Form(""),
    target_users: int = Form(0),
    confirmed: str = Form(""),
    workbook: UploadFile = File(...),
    db: Session = Depends(db_session),
):
    require_roles(request, "Admin")
    program_key = normalize_program(getattr(request.state, "program", DEFAULT_PROGRAM))
    if is_single_ran(program_key):
        raise HTTPException(status_code=400, detail="Area Builder is available for FTTH only")
    if str(confirmed).strip().lower() != "publish":
        raise HTTPException(status_code=400, detail="Confirm publishing before creating the area")
    area_name, city_name, start, end, users = area_builder_metadata(area, city, start_date, end_date, target_users)
    if not str(workbook.filename or "").lower().endswith(".xlsx"):
        raise HTTPException(status_code=400, detail="Upload an .xlsx workbook")
    existing = load_fiber_map_reference(db, program_key)
    if any(rollout_area_key(first_value(row, "Area", "Zone", default="")) == rollout_area_key(area_name) for row in existing.get("boxes") or []):
        raise HTTPException(status_code=409, detail="This area already exists. Use the map editor for changes to an existing area.")
    contents = await workbook.read()
    payload = build_area_map_from_workbook(contents, area_name, city_name)
    schematics = extract_area_builder_schematics(contents)
    actor = request_actor(request)
    saved = FiberMapArea(
        program=program_key,
        area=area_name,
        city=city_name,
        start_date=start,
        end_date=end,
        target_users=users,
        design_data=json.dumps({"boxes": payload["boxes"], "routes": payload["routes"]}, ensure_ascii=False),
        created_by=actor,
    )
    db.add(saved)
    for schematic in schematics:
        db.add(
            FiberMapSchematic(
                program=program_key,
                area=area_name,
                xbox=schematic["xbox"],
                sheet_name=schematic["sheet_name"],
                content_type=schematic["content_type"],
                image_data=schematic["image_data"],
                created_by=actor,
            )
        )
    if not db.query(Site).filter(Site.program == program_key, func.lower(Site.name) == area_name.lower()).first():
        db.add(Site(program=program_key, name=area_name))
    log_audit(db, "publish_fiber_map_area", "fiber_map_area", area_name, actor, {"city": city_name, "schematics": len(schematics), **payload["summary"]})
    db.commit()
    clear_rollout_db_cache()
    return {"success": True, "message": f"{area_name} was published", "area": {"name": area_name, "city": city_name, "start_date": start, "end_date": end, "target_users": users}, "schematics": [{"sheet_name": row["sheet_name"], "xbox": row["xbox"]} for row in schematics], **payload}


def rollout_manual_hub_allowed(area: str, xbox: str, hub_code: str) -> bool:
    area_hubs = ROLLOUT_MANUAL_HUBS.get(rollout_area_key(area), {})
    allowed_hubs = area_hubs.get(rollout_xbox_key(xbox), [])
    return rollout_code_key(hub_code) in {rollout_code_key(code) for code in allowed_hubs}


def rollout_record_code_type(row: RolloutRecord) -> str:
    if str(row.cable_code or "").strip():
        return "cable"
    if str(row.box_code or "").strip():
        return "box"
    return rollout_entry_mode({"item": row.item, "material type": row.material_type})


def next_rollout_entry_id(db: Session) -> str:
    global ROLLOUT_ENTRY_ID_CACHE
    now = time.monotonic()
    if ROLLOUT_ENTRY_ID_CACHE and now - ROLLOUT_ENTRY_ID_CACHE[0] < ROLLOUT_ENTRY_ID_CACHE_TTL:
        return ROLLOUT_ENTRY_ID_CACHE[1]

    highest = 0
    for (record_id,) in db.query(RolloutRecord.record_id).filter(RolloutRecord.record_id.like("RDP-%")).all():
        match = re.match(r"^RDP-(\d+)$", str(record_id or "").strip(), flags=re.I)
        if match:
            highest = max(highest, int(match.group(1)))
    next_id = f"RDP-{highest + 1}"
    ROLLOUT_ENTRY_ID_CACHE = (now, next_id)
    return next_id


def rollout_entry_counter(db: Session) -> RolloutEntryCounter:
    """Lock the shared Field Entry counter for the current transaction."""
    counter = (
        db.query(RolloutEntryCounter)
        .filter(RolloutEntryCounter.name == "field_entry_rdp")
        .with_for_update()
        .first()
    )
    if counter is not None:
        return counter

    highest = 0
    for (record_id,) in db.query(RolloutRecord.record_id).filter(RolloutRecord.record_id.like("RDP-%")).all():
        match = re.match(r"^RDP-(\d+)$", str(record_id or "").strip(), flags=re.I)
        if match:
            highest = max(highest, int(match.group(1)))

    # PostgreSQL and supported SQLite versions both accept this conflict-safe insert.
    db.execute(
        text(
            "INSERT INTO rollout_entry_counters (name, next_value) "
            "VALUES ('field_entry_rdp', :next_value) ON CONFLICT(name) DO NOTHING"
        ),
        {"next_value": highest + 1},
    )
    return (
        db.query(RolloutEntryCounter)
        .filter(RolloutEntryCounter.name == "field_entry_rdp")
        .with_for_update()
        .one()
    )


def allocate_rollout_entry_id(db: Session, counter: RolloutEntryCounter) -> str:
    value = max(int(counter.next_value or 1), 1)
    while db.query(RolloutRecord.id).filter(RolloutRecord.record_id == f"RDP-{value}").first() is not None:
        value += 1
    counter.next_value = value + 1
    db.flush()
    return f"RDP-{value}"


def rollout_entry_summary(rows: list[dict]) -> dict:
    summary = {"cable": 0, "end_box": 0, "sub_box": 0, "hub_box": 0, "xbox": 0}
    for row in rows:
        status = rollout_norm(first_value(row, "staus", "status", default=""))
        if status and status != "done":
            continue
        qty = safe_float(first_value(row, "actual", default=0)) or 0
        if qty <= 0:
            qty = 1
        text_value = rollout_norm(f"{first_value(row, 'item', default='')} {first_value(row, 'material type', default='')}")
        if "cable" in text_value:
            summary["cable"] += qty
        elif "xbox" in text_value:
            summary["xbox"] += qty
        elif "hubbox" in text_value or text_value == "hub":
            summary["hub_box"] += qty
        elif "endbox" in text_value or text_value.startswith("end"):
            summary["end_box"] += qty
        elif "subbox" in text_value or text_value.startswith("sub"):
            summary["sub_box"] += qty
    return summary


def rollout_duplicate_details(row: RolloutRecord) -> dict:
    return {
        "ID": row.record_id,
        "Date": row.date,
        "entry time": row.entry_time,
        "Supervisor Name": row.supervisor_name,
        "team leader": row.team_leader,
        "Area": row.area,
        "Related to XBOX": row.related_to_xbox,
        "item": row.item,
        "material type": row.material_type,
        "staus": row.status,
        "cable code": row.cable_code,
        "box code": row.box_code,
    }


def upsert_rollout_record(data: dict, db: Session, existing_by_id: dict[str, RolloutRecord] | None = None) -> tuple[RolloutRecord, bool]:
    record_id = stable_rollout_record_id(data)
    row = existing_by_id.get(record_id) if existing_by_id is not None else db.query(RolloutRecord).filter(RolloutRecord.record_id == record_id).first()
    created = row is None
    if row is None:
        row = RolloutRecord(record_id=record_id)
        db.add(row)
        if existing_by_id is not None:
            existing_by_id[record_id] = row

    row.date = str(first_value(data, "Date", "date", default="") or "")
    row.supervisor_name = str(first_value(data, "Supervisor Name", "supervisor_name", default="") or "")
    row.team_leader = str(first_value(data, "team leader", "Team Leader", "team_leader", default="") or "")
    row.area = str(first_value(data, "Area", "area", default="") or "")
    row.city = str(first_value(data, "city", "City", default="") or "")
    row.activity = str(first_value(data, "Activity", "activity", default="") or "")
    row.item = str(first_value(data, "item", "Item", default="") or "")
    row.material_type = str(first_value(data, "material type", "Material Type", "material_type", default="") or "")
    row.mount_type = str(first_value(data, "mount type", "Mount Type", "mount_type", default="") or "")
    row.item_serial = str(first_value(data, "item serial", "Item Serial", "item_serial", default="") or "")
    row.planned_quantity = safe_float(first_value(data, "planed quantity", "planned quantity", "planned_quantity", default=0))
    row.actual = safe_float(first_value(data, "actual", "Actual", default=0))
    row.stock_remaining = safe_float(first_value(data, "stock remaining", "Stock Remaining", "stock_remaining", default=0))
    row.status = str(first_value(data, "staus", "status", "Status", default="") or "")
    row.laser = str(first_value(data, "laser", "Laser", default="") or "")
    row.acceptance = str(first_value(data, "acceptance", "Acceptance", default="") or "")
    row.scan = str(first_value(data, "scan", "Scan", default="") or "")
    row.labeling = str(first_value(data, "labeling", "Labeling", default="") or "")
    row.related_to_xbox = str(first_value(data, "Related to XBOX", "related to xbox", "related_to_xbox", default="") or "")
    row.entry_time = str(first_value(data, "entry time", "Entry Time", "entry_time", default="") or "")
    row.cable_code = str(first_value(data, "cable code", "Cable Code", "cable_code", default="") or "")
    row.box_code = str(first_value(data, "box code", "Box Code", "box_code", default="") or "")
    row.olt = str(first_value(data, "OLT", "olt", default="") or "")
    row.cable_route = str(first_value(data, "Cable route", "Cable Route", "cable_route", default="") or "")
    row.notes = str(first_value(data, "Notes", "notes", default="") or "")
    submission_key = str(first_value(data, "submission_key", default="") or "").strip()
    if submission_key:
        row.submission_key = submission_key
    return row, created


FOUR_CORE_CABLE_NAMES = {
    "EOSDC309I": "4-coreCable_70m",
    "EOSDC309J": "4-coreCable_100m",
    "EOSDC309K": "4-coreCable_150m",
    "EOSDC309L": "4-coreCable_200m",
    "EOSDC309M": "4-coreCable_300m",
    "EOSDC309N": "4-coreCable_500m",
}

PART_NUMBER_BY_SKU = {
    "ITC3103-A1": "52590161",
    "E0SDC309J": "14130BQC-010",
    "E0SDC309K": "#N/A",
    "E0SDC309L": "14130BQC-012",
    "E0SDC309M": "14130BQC",
    "E0SDC309N": "14130BQC-001",
    "E0SDC309I": "14130BQC-009",
    "E00ATB101": "14260372",
    "FAT2810-SE-8-A": "14261299",
    "SSC2814-TM-2": "14261384",
    "SSC2812": "#N/A",
    "FAT2811-SH-4-B": "14261785",
    "ITC3301-P1_03": "52590919",
    "E0SDC309F": "#N/A",
    "ITC2102-P2": "14261388",
    "ITC3301-P1": "52590160",
    "E00DKBA04": "21150804",
    "L05-24VDD": "#N/A",
    "E0SDC030": "14137938-002",
    "E0SDC032": "14137938-004",
    "E0SDC034": "14137938-006",
    "E0SDC035": "14137938-007",
    "E0SDC024": "14137938",
    "E0SDC038": "14137938-011",
    "E0SDC029": "14137938-001",
    "E0SDC2155": "14130ALQ-003",
    "E0SDC2157": "14130ALQ-005",
    "E0SDC2171": "14130ALQ-007",
    "E0SDC2172": "14130ALQ-008",
    "E0SDC2173": "14130ALQ-009",
    "E0SDC2147": "14130ALQ",
    "E0SDC2153": "14130ALQ-001",
    "E0SDC2154": "14130ALQ-002",
    "FAT2810-SS-8-A": "14261298",
    "SSC2814-TM-2U": "14261383",
    "SSC2802-TX-8-B": "14261816",
}


def material_display_name(name: str = "", sku: str = "") -> str:
    sku_key = (sku or "").strip().upper()
    if sku_key in FOUR_CORE_CABLE_NAMES:
        return FOUR_CORE_CABLE_NAMES[sku_key]
    text = (name or "").strip()
    match = re.fullmatch(r"coreCable_(\d+m)-4", text, flags=re.IGNORECASE)
    if match:
        return f"4-coreCable_{match.group(1)}"
    return text


def product_part_number(sku: str = "", explicit: str = "") -> str:
    explicit_value = str(explicit or "").strip()
    if explicit_value:
        return explicit_value
    return PART_NUMBER_BY_SKU.get(str(sku or "").strip().upper(), "")


sync_product_part_numbers()


def product_display_name(product: Product | None) -> str:
    return material_display_name(product.name, product.sku) if product else ""


def product_to_dict(row: Product) -> dict:
    return {
        "id": row.id,
        "program": normalize_program(getattr(row, "program", DEFAULT_PROGRAM)),
        "sku": row.sku,
        "part_number": product_part_number(row.sku, row.part_number),
        "category": row.category,
        "name": material_display_name(row.name, row.sku),
        "item_detail": row.item_detail,
        "vendor": row.vendor,
        "qr_code": row.qr_code,
        "unit": row.unit,
        "tracking_type": row.tracking_type,
        "min_stock": row.min_stock,
        "status": row.status,
    }


def balance_to_dict(row: StockBalance, reserved_quantity: float = 0) -> dict:
    quantity = float(row.quantity or 0)
    reserved = float(reserved_quantity or 0)
    return {
        "program": normalize_program(getattr(row, "program", DEFAULT_PROGRAM)),
        "warehouse_id": row.warehouse_id,
        "warehouse": row.warehouse.name if row.warehouse else "",
        "product_id": row.product_id,
        "sku": row.product.sku if row.product else "",
        "part_number": product_part_number(row.product.sku if row.product else "", row.product.part_number if row.product else ""),
        "product": product_display_name(row.product),
        "unit": row.product.unit if row.product else "",
        "quantity": quantity,
        "reserved_quantity": reserved,
        "available_quantity": max(quantity - reserved, 0),
    }


def technician_balance_to_dict(row: TechnicianBalance) -> dict:
    return {
        "program": normalize_program(getattr(row, "program", DEFAULT_PROGRAM)),
        "technician_id": row.technician_id,
        "technician": row.technician.name if row.technician else "",
        "product_id": row.product_id,
        "sku": row.product.sku if row.product else "",
        "part_number": product_part_number(row.product.sku if row.product else "", row.product.part_number if row.product else ""),
        "product": product_display_name(row.product),
        "unit": row.product.unit if row.product else "",
        "quantity": row.quantity,
    }


def movement_to_dict(row: StockMovement) -> dict:
    return {
        "id": row.id,
        "program": normalize_program(getattr(row, "program", DEFAULT_PROGRAM)),
        "type": row.movement_type,
        "reference": row.reference,
        "warehouse": row.warehouse.name if row.warehouse else "",
        "technician": row.technician.name if row.technician else "",
        "sku": row.product.sku if row.product else "",
        "product": product_display_name(row.product),
        "quantity": row.quantity,
        "serial_number": row.serial_number,
        "created_by": row.created_by,
        "created_at": row.created_at.isoformat() if row.created_at else "",
    }


def scan_log_to_dict(row: MaterialScanLog) -> dict:
    requisition = row.requisition
    product = row.product
    warehouse = row.warehouse
    return {
        "id": row.id,
        "program": normalize_program(getattr(row, "program", DEFAULT_PROGRAM)),
        "material_requisition_id": row.material_requisition_id,
        "mr_order": requisition.order_number if requisition else "Stock Scan",
        "mr_status": requisition.status if requisition else row.status,
        "site_id": requisition.site_id if requisition else "",
        "site_address": requisition.site_address if requisition else "",
        "warehouse_id": row.warehouse_id,
        "warehouse": warehouse.name if warehouse else "",
        "product_id": row.product_id,
        "sku": product.sku if product else "",
        "material": product_display_name(product),
        "scan_code": row.scan_code,
        "serial_number": row.serial_number,
        "match_type": row.match_type,
        "status": row.status,
        "scanned_by": row.scanned_by,
        "note": row.note,
        "created_at": row.created_at.isoformat() if row.created_at else "",
    }


def receive_order_to_dict(row: ReceiveOrder) -> dict:
    items = []
    for item in row.items:
        product = item.product
        items.append(
            {
                "product_id": item.product_id,
                "sku": product.sku if product else "",
                "part_number": product_part_number(product.sku if product else "", product.part_number if product else ""),
                "name": product_display_name(product),
                "unit": product.unit if product else "",
                "qr_code": product.qr_code if product else "",
                "quantity": item.quantity,
                "serial_number": item.serial_number,
            }
        )
    return {
        "id": row.id,
        "program": normalize_program(getattr(row, "program", DEFAULT_PROGRAM)),
        "order_number": row.order_number,
        "receipt_date": row.receipt_date,
        "supplier": row.supplier,
        "warehouse_id": row.warehouse_id,
        "warehouse": row.warehouse.name if row.warehouse else "",
        "status": row.status,
        "created_by": row.created_by,
        "created_at": row.created_at.isoformat() if row.created_at else "",
        "items": items,
    }


def receive_order_header_to_dict(row: ReceiveOrder) -> dict:
    return {
        "id": row.id,
        "program": normalize_program(getattr(row, "program", DEFAULT_PROGRAM)),
        "order_number": row.order_number,
        "receipt_date": row.receipt_date,
        "supplier": row.supplier,
        "warehouse_id": row.warehouse_id,
        "warehouse": row.warehouse.name if row.warehouse else "",
        "status": row.status,
        "created_by": row.created_by,
        "created_at": row.created_at.isoformat() if row.created_at else "",
        "items": [],
    }


def requisition_to_dict(row: MaterialRequisition) -> dict:
    return {
        "id": row.id,
        "program": normalize_program(getattr(row, "program", DEFAULT_PROGRAM)),
        "order_number": row.order_number,
        "creation_date": row.creation_date,
        "warehouse_id": row.warehouse_id,
        "warehouse": row.warehouse.name if row.warehouse else "",
        "entity": row.entity,
        "project_name": row.project_name,
        "site_id": row.site_id,
        "site_address": row.site_address,
        "wo_no": row.wo_no,
        "product_domain": row.product_domain,
        "team_leader": row.team_leader,
        "receiver_tel": row.receiver_tel,
        "request_shipment_time": row.request_shipment_time,
        "request_arrived_site_time": row.request_arrived_site_time,
        "requester_name": row.requester_name,
        "requester_title": row.requester_title,
        "requester_signature": row.requester_signature,
        "requester_date": row.requester_date,
        "requester_comment": row.requester_comment,
        "receiver_name": row.receiver_name,
        "receiver_title": row.receiver_title,
        "receiver_signature": row.receiver_signature,
        "receiver_date": row.receiver_date,
        "receiver_comment": row.receiver_comment,
        "return_reason": row.return_reason,
        "status": row.status,
        "created_by": row.created_by,
        "created_at": row.created_at.isoformat() if row.created_at else "",
        "items": [
            {
                "id": item.id,
                "line_no": item.line_no,
                "product_id": item.product_id,
                "part_nbr": item.part_nbr,
                "model": item.model,
                "description": item.description,
                "vendor": item.vendor,
                "uom": item.uom,
                "quantity": item.quantity,
                "remark": item.remark,
            }
            for item in row.items
        ],
    }


def issue_material_requisition_row(db: Session, row: MaterialRequisition, actor: str = "") -> str:
    program_key = normalize_program(getattr(row, "program", DEFAULT_PROGRAM))
    if row.status == "issued":
        raise HTTPException(status_code=400, detail="Material requisition is already issued")
    if row.status not in {"approved", "signed"}:
        raise HTTPException(status_code=400, detail="MR must be approved before warehouse issue")
    if not row.receiver_name.strip():
        raise HTTPException(status_code=400, detail="Approver Name is required before warehouse issue")

    technician_name = (row.team_leader or row.requester_name or row.receiver_name).strip()
    technician = db.query(Technician).filter(Technician.name == technician_name, Technician.program == program_key).first()
    if technician is None:
        technician = Technician(program=program_key, name=technician_name, phone=row.receiver_tel)
        db.add(technician)
        db.flush()

    issued_by = actor.strip() or row.created_by
    issue = IssueOrder(
        program=program_key,
        order_number=next_number(db, IssueOrder, "MR-ISS"),
        warehouse_id=row.warehouse_id,
        technician_id=technician.id,
        status="confirmed",
        created_by=issued_by,
    )
    db.add(issue)
    db.flush()

    for item in row.items:
        if not item.product_id:
            raise HTTPException(status_code=400, detail=f"MR line {item.line_no} is not linked to a product")
        product = require_product(db, item.product_id, program_key)
        balance = locked_stock_balance(db, row.warehouse_id, item.product_id, program_key)
        if balance.quantity < item.quantity:
            warehouse_name = row.warehouse.name if row.warehouse else str(row.warehouse_id)
            material_name = product_display_name(product) or product.sku or f"product {product.id}"
            raise HTTPException(
                status_code=400,
                detail=(
                    f"Insufficient stock for {material_name} in {warehouse_name}. "
                    f"Requested {item.quantity}, available {balance.quantity}."
                ),
            )
        balance.quantity -= item.quantity
        technician_balance(db, technician.id, item.product_id, program_key).quantity += item.quantity
        db.add(IssueOrderItem(issue_order_id=issue.id, product_id=item.product_id, quantity=item.quantity, serial_number=""))
        db.add(
            StockMovement(
                program=program_key,
                movement_type="issue_to_technician",
                product_id=item.product_id,
                warehouse_id=row.warehouse_id,
                technician_id=technician.id,
                quantity=-item.quantity,
                reference=row.order_number,
                note="Issued from material requisition",
                created_by=issued_by,
            )
        )

    row.status = "issued"
    log_audit(db, "issue_material_requisition", "material_requisition", row.order_number, issued_by, {"issue_order": issue.order_number})
    return issue.order_number


def delete_material_requisition_row(db: Session, row: MaterialRequisition, actor: str = "admin") -> dict:
    program_key = normalize_program(getattr(row, "program", DEFAULT_PROGRAM))
    restored = 0.0
    movements = (
        db.query(StockMovement)
        .filter(StockMovement.reference == row.order_number, StockMovement.movement_type == "issue_to_technician", StockMovement.program == program_key)
        .all()
    )
    for movement in movements:
        qty = abs(movement.quantity or 0)
        if movement.warehouse_id and movement.product_id:
            stock_balance(db, movement.warehouse_id, movement.product_id, program_key).quantity += qty
            restored += qty
        if movement.technician_id and movement.product_id:
            tech_balance = technician_balance(db, movement.technician_id, movement.product_id, program_key)
            tech_balance.quantity = max(0, (tech_balance.quantity or 0) - qty)
        db.delete(movement)

    issue_numbers: set[str] = set()
    audits = (
        db.query(AuditLog)
        .filter(
            AuditLog.action == "issue_material_requisition",
            AuditLog.entity_type == "material_requisition",
            AuditLog.entity_id == row.order_number,
        )
        .all()
    )
    for audit in audits:
        try:
            details = json.loads(audit.details or "{}")
        except Exception:
            details = {}
        issue_order = str(details.get("issue_order") or "").strip()
        if issue_order:
            issue_numbers.add(issue_order)
    if issue_numbers:
        for issue in db.query(IssueOrder).filter(IssueOrder.order_number.in_(issue_numbers)).all():
            db.delete(issue)

    db.query(MaterialScanLog).filter(MaterialScanLog.material_requisition_id == row.id).delete(synchronize_session=False)
    order_number = row.order_number
    status = row.status
    db.delete(row)
    log_audit(
        db,
        "delete_material_requisition",
        "material_requisition",
        order_number,
        actor or "admin",
        {"status": status, "restored_quantity": restored, "deleted_issue_orders": sorted(issue_numbers)},
    )
    clear_warehouse_cache()
    return {"order_number": order_number, "restored_quantity": restored, "deleted_issue_orders": sorted(issue_numbers)}


@app.get("/")
def home():
    return FileResponse("static/materials_inventory.html", headers={"Cache-Control": "no-store"})


@app.get("/rollout")
def rollout_home():
    return FileResponse("static/ftth_rollout.html", headers={"Cache-Control": "no-store"})


@app.get("/warehouse")
def warehouse_home():
    return FileResponse("static/materials_inventory.html", headers={"Cache-Control": "no-store"})


@app.post("/api/auth/login")
def login(data: LoginIn, response: Response, db: Session = Depends(db_session)):
    key = data.username.strip().lower()
    program_key = normalize_program(data.program)
    try:
        db_user = (
            db.query(AppUser)
            .filter(func.lower(AppUser.username) == key, AppUser.status == "active")
            .all()
        )
        user = next(
            (
                row
                for row in db_user
                if app_user_matches_program(row, program_key) and verify_password(data.password, row.password_hash)
            ),
            None,
        )
        if user:
            if user.password_hash and not is_bcrypt_hash(user.password_hash):
                user.password_hash = hash_password(data.password)
                db.commit()
            token, csrf_token = create_app_session(db, user, program_key)
            response.set_cookie(key=SESSION_COOKIE_NAME, value=token, max_age=SESSION_TTL_HOURS * 60 * 60, httponly=True, secure=SESSION_COOKIE_SECURE, samesite="lax", path="/")
            return {"success": True, "user": session_user_payload(user, program_key), "csrf_token": csrf_token}

        deleted_fallback = (
            db.query(AppUser)
            .filter(func.lower(AppUser.username) == key, AppUser.status == "inactive")
            .all()
        )
        if any(
            not is_admin_role(row.role)
            and app_user_matches_program(row, program_key)
            and verify_password(data.password, row.password_hash)
            for row in deleted_fallback
        ):
            raise HTTPException(status_code=401, detail="Invalid username or password")
    except HTTPException:
        raise
    except Exception:
        db.rollback()

    raise HTTPException(status_code=401, detail="Invalid username or password")


@app.get("/api/auth/me")
def current_session(request: Request):
    return {"success": True, "user": session_user_payload(current_user(request), request.state.session_program), "csrf_token": request.state.csrf_token}


@app.post("/api/auth/heartbeat")
def heartbeat(request: Request, db: Session = Depends(db_session)):
    token = request.cookies.get(SESSION_COOKIE_NAME, "")
    if not token:
        raise HTTPException(status_code=401, detail="Authentication required")
    db.query(AppSession).filter(AppSession.token_hash == session_token_hash(token)).update(
        {AppSession.last_seen_at: datetime.now(timezone.utc)},
        synchronize_session=False,
    )
    db.commit()
    return {"success": True}


@app.get("/api/auth/active-users")
def active_users(request: Request):
    require_roles(request, "Admin")
    now = datetime.now(timezone.utc)
    cutoff = now - timedelta(seconds=ACTIVE_USER_WINDOW_SECONDS)
    latest_by_user: dict[tuple[str, str], dict] = {}
    for program_key, session_factory in all_sessionmakers():
        with session_factory() as db:
            rows = (
                db.query(AppSession)
                .options(joinedload(AppSession.user))
                .filter(
                    AppSession.expires_at > now,
                    AppSession.last_seen_at >= cutoff,
                )
                .all()
            )
            for session in rows:
                user = session.user
                if user is None or user.status != "active":
                    continue
                key = (user.username.strip().lower(), normalize_program(session.program))
                current = latest_by_user.get(key)
                if current and current["last_seen_at"] >= session.last_seen_at:
                    continue
                latest_by_user[key] = {
                    "name": user.name or user.username,
                    "username": user.username,
                    "role": user.role,
                    "program": normalize_program(session.program),
                    "last_seen_at": session.last_seen_at.isoformat() if session.last_seen_at else "",
                }
    users = sorted(latest_by_user.values(), key=lambda row: (row["name"].lower(), row["program"]))
    return {"success": True, "count": len(users), "users": users}


@app.post("/api/auth/logout")
def logout(request: Request, response: Response, db: Session = Depends(db_session)):
    clear_app_session(db, request.cookies.get(SESSION_COOKIE_NAME, ""))
    response.delete_cookie(key=SESSION_COOKIE_NAME, path="/")
    return {"success": True}


def requisition_header_to_dict(row: MaterialRequisition) -> dict:
    return {
        "id": row.id,
        "program": normalize_program(getattr(row, "program", DEFAULT_PROGRAM)),
        "order_number": row.order_number,
        "creation_date": row.creation_date,
        "warehouse_id": row.warehouse_id,
        "warehouse": row.warehouse.name if row.warehouse else "",
        "entity": row.entity,
        "project_name": row.project_name,
        "site_id": row.site_id,
        "site_address": row.site_address,
        "wo_no": row.wo_no,
        "product_domain": row.product_domain,
        "team_leader": row.team_leader,
        "receiver_tel": row.receiver_tel,
        "request_shipment_time": row.request_shipment_time,
        "request_arrived_site_time": row.request_arrived_site_time,
        "requester_name": row.requester_name,
        "requester_title": row.requester_title,
        "requester_date": row.requester_date,
        "requester_comment": row.requester_comment,
        "receiver_name": row.receiver_name,
        "receiver_title": row.receiver_title,
        "receiver_date": row.receiver_date,
        "receiver_comment": row.receiver_comment,
        "return_reason": row.return_reason,
        "status": row.status,
        "created_by": row.created_by,
        "created_at": row.created_at.isoformat() if row.created_at else "",
        "items": [],
    }


def user_can_view_requisition(row: MaterialRequisition, viewer: str = "", role: str = "") -> bool:
    role_key = normalize_usage_key(role)
    viewer_key = normalize_usage_key(viewer)
    if role_key in {"admin", "management"}:
        return True
    if role_key == "warehousemanager":
        return warehouse_manager_handles_mr(viewer, row)
    if role_key in {"approval", "approver"}:
        return row.status == "pending_approval" or normalize_usage_key(row.receiver_name) == viewer_key
    if role_key == "requester":
        return normalize_usage_key(row.requester_name) == viewer_key or normalize_usage_key(row.created_by) == viewer_key
    return normalize_usage_key(row.created_by) == viewer_key


def user_can_view_transfer(row: MaterialTransfer, viewer: str = "", role: str = "") -> bool:
    role_key = normalize_usage_key(role)
    viewer_key = normalize_usage_key(viewer)
    if role_key in {"admin", "management"}:
        return True
    if role_key == "warehousemanager":
        return warehouse_scope_matches(viewer, row.from_warehouse.name if row.from_warehouse else "") or warehouse_scope_matches(
            viewer, row.to_warehouse.name if row.to_warehouse else ""
        )
    if role_key in {"approval", "approver"}:
        return row.status == "pending_approval" or normalize_usage_key(row.approver_name) == viewer_key
    if role_key == "requester":
        return normalize_usage_key(row.requester_name) == viewer_key or normalize_usage_key(row.created_by) == viewer_key
    return normalize_usage_key(row.created_by) == viewer_key


def canonical_area_name(value: str) -> str:
    text_value = str(value or "").strip()
    key = normalize_usage_key(text_value)
    if key in {"haydamascus", "haydemascus", "haydemashq"}:
        return "Hay Demascus"
    if key in {"maqawba", "magawba"}:
        return "Maqawba"
    if key in {"hayalandalusz3", "hayalandaluszone3"}:
        return "Hay Al Andalus Z3"
    if key in {"hayandalus", "hayalandalus", "hayalandalusz2", "hayalandaluszone2"}:
        return "Hay Al Andalus Z2"
    return text_value


def canonical_mr_history_area(value: str) -> str:
    return canonical_area_name(value)


def requisition_history_row_to_dict(row: MaterialRequisition) -> dict:
    item_count = len(row.items or [])
    total_quantity = sum(float(item.quantity or 0) for item in row.items)
    materials = [item.description for item in row.items if str(item.description or "").strip()]
    return {
        "id": row.id,
        "program": normalize_program(getattr(row, "program", DEFAULT_PROGRAM)),
        "order_number": row.order_number,
        "creation_date": row.creation_date,
        "warehouse": row.warehouse.name if row.warehouse else "",
        "warehouse_id": row.warehouse_id,
        "site_id": canonical_mr_history_area(row.site_id),
        "site_address": row.site_address,
        "requester_name": row.requester_name,
        "team_leader": row.team_leader,
        "receiver_name": row.receiver_name,
        "status": row.status,
        "entity": row.entity,
        "project_name": row.project_name,
        "product_domain": row.product_domain,
        "created_by": row.created_by,
        "item_count": item_count,
        "total_quantity": total_quantity,
        "materials": materials,
        "materials_text": ", ".join(materials),
    }


def requisition_history_payload(
    db: Session,
    warehouse: str = "",
    area: str = "",
    technician: str = "",
    requester: str = "",
    status: str = "",
    date_from: str = "",
    date_to: str = "",
    viewer: str = "",
    role: str = "",
    program: str = DEFAULT_PROGRAM,
) -> dict:
    program_key = normalize_program(program)
    rows = (
        db.query(MaterialRequisition)
        .options(joinedload(MaterialRequisition.warehouse), selectinload(MaterialRequisition.items))
        .filter(MaterialRequisition.program == program_key)
        .order_by(MaterialRequisition.id.desc())
        .all()
    )
    visible_rows = [row for row in rows if user_can_view_requisition(row, viewer, role)]
    options = {
        "warehouses": sorted({str(row.warehouse.name if row.warehouse else "").strip() for row in visible_rows if str(row.warehouse.name if row.warehouse else "").strip()}),
        "areas": sorted({canonical_mr_history_area(row.site_id) for row in visible_rows if str(row.site_id or "").strip()}),
        "technicians": sorted({str(row.team_leader or "").strip() for row in visible_rows if str(row.team_leader or "").strip()}),
        "requesters": sorted({str(row.requester_name or "").strip() for row in visible_rows if str(row.requester_name or "").strip()}),
        "statuses": sorted({str(row.status or "").strip() for row in visible_rows if str(row.status or "").strip()}),
    }

    def match_filter(value: str, target: str) -> bool:
        if not value:
            return True
        return normalize_usage_key(target) == normalize_usage_key(value)

    filtered_rows = []
    for row in visible_rows:
        if not match_filter(warehouse, row.warehouse.name if row.warehouse else ""):
            continue
        if not match_filter(area, canonical_mr_history_area(row.site_id)):
            continue
        if not match_filter(technician, row.team_leader):
            continue
        if not match_filter(requester, row.requester_name):
            continue
        if not match_filter(status, row.status):
            continue
        creation_date = str(row.creation_date or "").strip()
        if date_from and creation_date and creation_date < date_from:
            continue
        if date_to and creation_date and creation_date > date_to:
            continue
        if (date_from or date_to) and not creation_date:
            continue
        filtered_rows.append(row)

    items_total = sum(len(row.items or []) for row in filtered_rows)
    quantity_total = sum(sum(float(item.quantity or 0) for item in row.items) for row in filtered_rows)
    return {
        "rows": [requisition_history_row_to_dict(row) for row in filtered_rows],
        "summary": {
            "mr_count": len(filtered_rows),
            "item_count": items_total,
            "total_quantity": quantity_total,
        },
        "options": options,
    }


def transfer_to_dict(row: MaterialTransfer, include_items: bool = True) -> dict:
    items = []
    if include_items:
        items = [
            {
                "id": item.id,
                "line_no": item.line_no,
                "product_id": item.product_id,
                "part_nbr": item.part_nbr,
                "description": item.description,
                "vendor": str(item.product.vendor or "").strip() if item.product else "",
                "uom": item.uom,
                "quantity": item.quantity,
                "remark": item.remark,
            }
            for item in row.items
        ]
    return {
        "id": row.id,
        "program": normalize_program(getattr(row, "program", DEFAULT_PROGRAM)),
        "transfer_number": row.transfer_number,
        "transfer_date": row.transfer_date,
        "from_warehouse_id": row.from_warehouse_id,
        "from_warehouse": row.from_warehouse.name if row.from_warehouse else "",
        "to_warehouse_id": row.to_warehouse_id,
        "to_warehouse": row.to_warehouse.name if row.to_warehouse else "",
        "reference_no": row.reference_no,
        "reason": row.reason,
        "requester_name": row.requester_name,
        "requester_title": row.requester_title,
        "approver_name": row.approver_name,
        "approver_title": row.approver_title,
        "approver_date": row.approver_date,
        "approver_comment": row.approver_comment,
        "receiver_name": row.receiver_name,
        "receiver_date": row.receiver_date,
        "receiver_comment": row.receiver_comment,
        "status": row.status,
        "created_by": row.created_by,
        "created_at": row.created_at.isoformat() if row.created_at else "",
        "items": items,
    }


def material_return_to_dict(row: MaterialReturn, include_items: bool = True) -> dict:
    items = []
    if include_items:
        items = [
            {
                "id": item.id,
                "line_no": item.line_no,
                "product_id": item.product_id,
                "part_nbr": item.part_nbr,
                "description": item.description,
                "uom": item.uom,
                "quantity": item.quantity,
                "condition": item.condition,
                "remark": item.remark,
            }
            for item in row.items
        ]
    return {
        "id": row.id,
        "program": normalize_program(getattr(row, "program", DEFAULT_PROGRAM)),
        "return_number": row.return_number,
        "return_date": row.return_date,
        "site_id": row.site_id,
        "site_address": row.site_address,
        "warehouse_id": row.warehouse_id,
        "warehouse": row.warehouse.name if row.warehouse else "",
        "returned_by": row.returned_by,
        "received_by": row.received_by,
        "reason": row.reason,
        "status": row.status,
        "created_by": row.created_by,
        "created_at": row.created_at.isoformat() if row.created_at else "",
        "items": items,
    }


@app.get("/api/auth/users")
def list_app_users(request: Request, program: str = DEFAULT_PROGRAM, db: Session = Depends(db_session)):
    require_roles(request, "Admin")
    program_key = normalize_program(program)
    try:
        db_query = db.query(AppUser).filter(AppUser.status == "active")
        if is_single_ran(program_key):
            db_query = db_query.filter(AppUser.program == program_key, AppUser.role != "Admin")
        else:
            db_query = db_query.filter(or_(AppUser.program == program_key, AppUser.role == "Admin"))
        db_rows = db_query.order_by(AppUser.id.asc()).all()
        db_users = [serialize_app_user(row) for row in db_rows]
        seen = {app_user_key(row.username, row.role) for row in db_rows}
        deleted = {
            app_user_key(row.username, row.role)
            for row in db.query(AppUser).filter(AppUser.status == "inactive", AppUser.program == program_key).all()
        }
    except Exception:
        db.rollback()
        db_users = []
        seen = set()
        deleted = set()
    fallback_users = [
        serialize_app_user(row, fallback=True)
        for row in APP_USERS
        if not is_single_ran(program_key)
        and app_user_matches_program(row, program_key)
        if app_user_key(row["username"], row["role"]) not in seen
        and app_user_key(row["username"], row["role"]) not in deleted
    ]
    return {
        "success": True,
        "users": db_users + fallback_users,
    }


@app.post("/api/auth/users")
def create_app_user(data: AppUserIn, request: Request, db: Session = Depends(db_session)):
    require_roles(request, "Admin")
    username = data.username.strip()
    password = data.password.strip()
    role = data.role.strip()
    program_key = normalize_program(data.program)
    if not username:
        raise HTTPException(status_code=400, detail="User is required")
    filters = [
        func.lower(AppUser.username) == username.lower(),
        AppUser.role == role,
        AppUser.status == "active",
    ]
    if not is_admin_role(role):
        filters.append(AppUser.program == program_key)
    existing = (
        db.query(AppUser)
        .filter(*filters)
        .first()
    )
    if existing:
        existing.name = data.name.strip() or username
        existing.email = data.email.strip()
        existing.warehouse_name = data.warehouse_name.strip()
        if password:
            existing.password_hash = hash_password(password)
        user = existing
    else:
        fallback = fallback_user_template(username, role)
        password_hash = hash_password(password) if password else (fallback["password_hash"] if fallback else "")
        if not password_hash:
            raise HTTPException(status_code=400, detail="Password is required")
        user = AppUser(
            program=DEFAULT_PROGRAM if is_admin_role(role) else program_key,
            username=username,
            name=data.name.strip() or username,
            role=role,
            email=data.email.strip(),
            warehouse_name=data.warehouse_name.strip(),
            password_hash=password_hash,
            status="active",
        )
        db.add(user)
    db.commit()
    db.refresh(user)
    return {
        "success": True,
        "user": serialize_app_user(user),
    }


@app.post("/api/auth/users/delete")
def delete_app_user(data: AppUserDeleteIn, request: Request, db: Session = Depends(db_session)):
    require_roles(request, "Admin")
    username = data.username.strip()
    role = data.role.strip()
    program_key = normalize_program(data.program)
    if not username or not role:
        raise HTTPException(status_code=400, detail="User and role are required")
    filters = [
        func.lower(AppUser.username) == username.lower(),
        AppUser.role == role,
        AppUser.status == "active",
    ]
    if not is_admin_role(role):
        filters.append(AppUser.program == program_key)
    user = db.query(AppUser).filter(*filters).first()
    if user:
        user.status = "inactive"
    else:
        fallback = fallback_user_template(username, role)
        if not fallback or is_single_ran(program_key):
            raise HTTPException(status_code=404, detail="User not found")
        db.add(
            AppUser(
                program=DEFAULT_PROGRAM,
                username=username,
                name=fallback.get("name") or username,
                role=role,
                email=fallback.get("email", ""),
                warehouse_name=fallback.get("warehouse_name", ""),
                password_hash=fallback["password_hash"],
                status="inactive",
            )
        )
    db.commit()
    return {"success": True}


@app.get("/api/records")
def list_records(request: Request, db: Session = Depends(db_session)):
    rows = db.query(RolloutRecord).order_by(RolloutRecord.id.desc()).all()
    rows = [row for row in rows if rollout_records_for_session(request, [row_to_record(row)])]
    return {"success": True, "records": [row_to_record(row) for row in rows]}


@app.post("/api/records")
def save_record(data: dict, request: Request, db: Session = Depends(db_session)):
    require_roles(request, "Admin", "Management", "Requester")
    data["program"] = request.state.program
    row, _ = upsert_rollout_record(data, db)
    db.commit()
    clear_warehouse_cache()
    clear_rollout_db_cache()
    db.refresh(row)

    return {
        "success": True,
        "message": "Progress saved",
        "record": row_to_record(row),
    }


@app.get("/api/warehouse/rollout-daily-progress")
def list_rollout_daily_progress(request: Request, limit: int = 500, refresh: str = "", program: str = DEFAULT_PROGRAM, db: Session = Depends(db_session)):
    require_roles(request, "Admin", "Management", "Requester", "Approval", "Warehouse Manager")
    if is_single_ran(program):
        return {"success": True, "source": "disabled", "count": 0, "records": []}
    force_refresh = str(refresh or "").strip().lower() in {"1", "true", "yes", "now"} or str(refresh or "").strip().isdigit()
    rows, source = rollout_daily_progress_records(db, force=force_refresh)
    rows = rollout_records_for_session(request, rows)
    limited = list(reversed(rows))[: min(max(limit, 1), 10000)]
    return {
        "success": True,
        "name": "Rollout Daily Progress",
        "source": source,
        "read_only": source == "google_csv",
        "count": len(rows),
        "fetched_at": datetime.now(TRIPOLI_TZ).isoformat(),
        "records": limited,
    }


@app.get("/api/warehouse/rollout-dashboard-summary")
def rollout_dashboard_summary(request: Request, program: str = DEFAULT_PROGRAM, db: Session = Depends(db_session)):
    """Return only the grouped values used by the Rollout KPI dashboard."""
    require_roles(request, "Admin", "Management", "Requester", "Approval", "Warehouse Manager")
    if is_single_ran(program):
        return {"success": True, "source": "disabled", "count": 0, "records": [], "summary_only": True}

    grouped_rows = (
        db.query(
            RolloutRecord.date.label("date"),
            RolloutRecord.city.label("city"),
            RolloutRecord.area.label("area"),
            RolloutRecord.item.label("item"),
            RolloutRecord.material_type.label("material_type"),
            RolloutRecord.team_leader.label("team_leader"),
            RolloutRecord.status.label("status"),
            RolloutRecord.related_to_xbox.label("related_to_xbox"),
            RolloutRecord.cable_code.label("cable_code"),
            RolloutRecord.box_code.label("box_code"),
            func.sum(RolloutRecord.actual).label("actual"),
            func.count(RolloutRecord.id).label("record_count"),
        )
        .group_by(
            RolloutRecord.date,
            RolloutRecord.city,
            RolloutRecord.area,
            RolloutRecord.item,
            RolloutRecord.material_type,
            RolloutRecord.team_leader,
            RolloutRecord.status,
            RolloutRecord.related_to_xbox,
            RolloutRecord.cable_code,
            RolloutRecord.box_code,
        )
        .order_by(RolloutRecord.date.asc(), RolloutRecord.city.asc(), RolloutRecord.area.asc())
        .all()
    )
    records = [
        {
            "Date": str(row.date) if row.date else "",
            "city": row.city or "",
            "Area": row.area or "",
            "item": row.item or "",
            "material type": row.material_type or "",
            "team leader": row.team_leader or "",
            "staus": row.status or "Done",
            "Related to XBOX": row.related_to_xbox or "",
            "cable code": row.cable_code or "",
            "box code": row.box_code or "",
            "actual": float(row.actual or 0),
            "__rollout_record_count": int(row.record_count or 0),
        }
        for row in grouped_rows
    ]
    records = rollout_records_for_session(request, records)
    records = rollout_dashboard_summary_metric_rows(records, db, program)
    response = {
        "success": True,
        "name": "Rollout Dashboard Summary",
        "source": "database",
        "summary_only": True,
        "count": sum(int(row.get("__rollout_record_count") or 0) for row in records),
        "fetched_at": datetime.now(TRIPOLI_TZ).isoformat(),
        "records": records,
        "metrics": {
            "database_rows_loaded": len(grouped_rows),
            "rows_returned": len(records),
            "full_record_rows_avoided": sum(int(row.get("__rollout_record_count") or 0) for row in records),
        },
    }
    response["metrics"]["estimated_payload_bytes"] = len(json.dumps(response, ensure_ascii=False, separators=(",", ":")).encode("utf-8"))
    return response


@app.get("/api/warehouse/rollout-daily-progress/export")
def export_rollout_daily_progress(request: Request, program: str = DEFAULT_PROGRAM, db: Session = Depends(db_session)):
    require_roles(request, "Admin", "Management", "Requester", "Approval", "Warehouse Manager")
    if is_single_ran(program):
        rows: list[dict] = []
    else:
        rows, _source = rollout_daily_progress_records(db, force=False)
        rows = rollout_records_for_session(request, rows)

    headers = [
        "ID", "Date", "entry time", "Supervisor Name", "team leader", "city", "Area", "Activity",
        "Related to XBOX", "item", "material type", "mount type", "Cable route", "cable code", "box code",
        "item serial", "planed quantity", "actual", "stock remaining", "staus", "laser", "acceptance", "scan", "labeling", "Notes",
    ]
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Rollout Progress"
    sheet.append(headers)
    for row in rows:
        sheet.append([row.get(header, "") for header in headers])
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions

    # Keep the main sheet complete, and provide a focused view for reviewing
    # Hub-by-Hub accessory entries without changing the underlying records.
    accessory_rows = [
        row for row in rows
        if rollout_norm(row.get("item")) in {"hubaccessory", "hubaccessories"}
    ]
    if accessory_rows:
        accessories_sheet = workbook.create_sheet("Hub Accessories")
        accessories_sheet.append(headers)
        for row in accessory_rows:
            accessories_sheet.append([row.get(header, "") for header in headers])
        accessories_sheet.freeze_panes = "A2"
        accessories_sheet.auto_filter.ref = accessories_sheet.dimensions

    for index, header in enumerate(headers, start=1):
        values = [header] + [str(row.get(header, "") or "") for row in rows[:250]]
        sheet.column_dimensions[chr(64 + index) if index <= 26 else "A"].width = min(42, max(12, max(len(value) for value in values) + 2))
    stream = io.BytesIO()
    workbook.save(stream)
    stream.seek(0)
    filename = f"rollout-progress-{datetime.now(TRIPOLI_TZ).strftime('%Y%m%d-%H%M%S')}.xlsx"
    return StreamingResponse(
        stream,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@app.get("/api/warehouse/rollout-entry-reference")
def rollout_entry_reference(
    request: Request,
    area: str = "",
    xbox: str = "",
    query: str = "",
    limit: int = 50,
    db: Session = Depends(db_session),
):
    records, source = rollout_daily_progress_records(db, force=False)
    records = rollout_records_for_session(request, records)
    latest_records = records
    if query:
        needle = rollout_norm(query)
        latest_records = [
            row
            for row in latest_records
            if needle in rollout_norm(" ".join(str(v or "") for v in row.values()))
        ]
    latest = list(reversed(latest_records))[: min(max(limit, 1), 200)]

    refs = rollout_code_reference_rows(db, getattr(request.state, "program", DEFAULT_PROGRAM))
    areas_map: dict[str, dict] = {}
    xboxes_map: dict[str, set[str]] = {}
    for row in refs:
        area_name = row.get("area") or ""
        if not area_name:
            continue
        area_key = rollout_area_key(area_name)
        areas_map.setdefault(area_key, {"area": area_name, "city": row.get("city") or ""})
        xboxes_map.setdefault(area_key, set()).add(row.get("xbox") or "")
    code_states: dict[tuple[str, str, str, str], str] = {}
    for record in records:
        code_type = rollout_entry_mode(record)
        if code_type not in {"cable", "box"}:
            continue
        code_value = first_value(record, "cable code", "Cable Code", default="") if code_type == "cable" else first_value(record, "box code", "Box Code", default="")
        code = str(code_value or "").strip()
        if not code:
            continue
        key = (
            rollout_area_key(first_value(record, "Area", "area", default="")),
            rollout_xbox_key(first_value(record, "Related to XBOX", "related_to_xbox", default="")),
            code_type,
            rollout_code_key(code),
        )
        status = "done" if rollout_norm(first_value(record, "staus", "status", default="")) == "done" else "in_progress"
        if status == "done" or key not in code_states:
            code_states[key] = status

    scoped_refs = refs
    if area:
        scoped_refs = [r for r in scoped_refs if rollout_area_key(r.get("area")) == rollout_area_key(area)]
    if xbox:
        scoped_refs = [r for r in scoped_refs if rollout_xbox_key(r.get("xbox")) == rollout_xbox_key(xbox)]
    scoped_refs = [
        {
            **row,
            "entry_status": code_states.get(
                (rollout_area_key(row.get("area")), rollout_xbox_key(row.get("xbox")), row.get("type") or "", rollout_code_key(row.get("code"))),
                "available",
            ),
        }
        for row in scoped_refs
    ]

    return {
        "success": True,
        "source": source,
        "next_id": next_rollout_entry_id(db),
        # The entry form is scoped to an Area/XBOX, but its headline cards are
        # a project-wide installation total and must not change with that scope.
        "summary": rollout_entry_summary(records),
        "areas": sorted(areas_map.values(), key=lambda r: (r.get("city") or "", r.get("area") or "")),
        "xboxes_by_area": {areas_map[k]["area"]: sorted(v) for k, v in xboxes_map.items()},
        "codes": scoped_refs,
        "latest": latest,
        "count": len(latest_records),
    }


@app.post("/api/warehouse/rollout-field-entry")
def save_rollout_field_entry(data: dict, request: Request, db: Session = Depends(db_session)):
    require_roles(request, "Requester", "Admin")

    submission_key = str(first_value(data, "submission_key", default="") or "").strip()
    if not re.fullmatch(r"[A-Za-z0-9-]{16,128}", submission_key):
        raise HTTPException(status_code=400, detail="Invalid field entry submission key")
    existing_submission = db.query(RolloutRecord).filter(RolloutRecord.submission_key == submission_key).first()
    if existing_submission is not None:
        records, _ = rollout_daily_progress_records(db, force=False)
        return {
            "success": True,
            "message": "Field entry already saved",
            "record": row_to_record(existing_submission),
            "summary": rollout_entry_summary(records),
            "warnings": [],
            "replayed": True,
        }

    area = str(first_value(data, "Area", "area", default="") or "").strip()
    xbox = str(first_value(data, "Related to XBOX", "related_to_xbox", default="") or "").strip()
    code_type = rollout_entry_mode(data)
    raw_code = str(first_value(data, "code", "Cable Code", "cable code", "cable_code", "Box Code", "box code", "box_code", default="") or "").strip()
    if not area or not xbox:
        raise HTTPException(status_code=400, detail="Select Area and Related to XBOX")
    if code_type not in {"cable", "box", "accessory"}:
        raise HTTPException(status_code=400, detail="Select a material type")
    if safe_float(first_value(data, "actual", "Actual", default=0)) <= 0:
        raise HTTPException(status_code=400, detail="Actual quantity must be greater than zero")

    matches: list[dict] = []
    duplicates: list[RolloutRecord] = []
    if code_type in {"cable", "box"}:
        if not raw_code:
            raise HTTPException(status_code=400, detail="Select a code from the list")
        matches = rollout_reference_matches(area, xbox, raw_code, code_type, db, getattr(request.state, "program", DEFAULT_PROGRAM))
        if not matches:
            raise HTTPException(status_code=400, detail="Code is not listed for this Area / XBOX / Type")

        code_attr = RolloutRecord.cable_code if code_type == "cable" else RolloutRecord.box_code
        existing_candidates = db.query(RolloutRecord).filter(code_attr != "").all()
        for row in existing_candidates:
            if rollout_record_code_type(row) != code_type:
                continue
            if rollout_area_key(row.area) != rollout_area_key(area):
                continue
            if rollout_xbox_key(row.related_to_xbox) != rollout_xbox_key(xbox):
                continue
            saved_code = row.cable_code if code_type == "cable" else row.box_code
            if rollout_code_key(saved_code) == rollout_code_key(raw_code):
                duplicates.append(row)

        done_duplicate = next((row for row in duplicates if rollout_norm(row.status) == "done"), None)
        if done_duplicate:
            raise HTTPException(
                status_code=400,
                detail=f"Code already saved as Done in {done_duplicate.record_id}",
            )

    warnings: list[dict] = []
    in_progress = next((row for row in duplicates if rollout_norm(row.status) in {"inprogress", "planned", "blocked"}), None)
    if in_progress:
        warnings.append({"type": "duplicate_in_progress", "message": "Code exists in another non-Done record", "record": rollout_duplicate_details(in_progress)})

    material_type = str(first_value(data, "material type", "Material Type", "material_type", default="") or "").strip()
    material_norm = rollout_norm(material_type)
    ref = matches[0] if matches else {}
    expected_len = int(ref.get("cable_length_m") or 0)
    if code_type == "cable" and expected_len:
        length_match = re.search(r"(\d+)\s*m", material_type, flags=re.I)
        entered_len = int(length_match.group(1)) if length_match else 0
        if entered_len and entered_len != expected_len:
            warnings.append({"type": "cable_length_mismatch", "message": f"Planned length is {expected_len}m, selected material looks {entered_len}m"})
    if code_type == "box":
        expected_box = rollout_norm(ref.get("box_type") or "")
        if expected_box:
            if "end" in expected_box and "end" not in material_norm:
                warnings.append({"type": "box_type_mismatch", "message": "Map reference is END BOX, selected material is different"})
            if "sub" in expected_box and "sub" not in material_norm:
                warnings.append({"type": "box_type_mismatch", "message": "Map reference is SUB BOX, selected material is different"})

    if warnings and not first_value(data, "confirm_warnings", "confirmed", default=False):
        return {"success": False, "needs_confirmation": True, "warnings": warnings, "message": "Confirm warnings before saving"}

    with ROLLOUT_SYNC_LOCK:
        counter = rollout_entry_counter(db)
        # A retry can arrive while the first request is committing. Check again
        # after the database lock so the same submission is never saved twice.
        existing_submission = db.query(RolloutRecord).filter(RolloutRecord.submission_key == submission_key).first()
        if existing_submission is not None:
            db.rollback()
            records, _ = rollout_daily_progress_records(db, force=False)
            return {
                "success": True,
                "message": "Field entry already saved",
                "record": row_to_record(existing_submission),
                "summary": rollout_entry_summary(records),
                "warnings": [],
                "replayed": True,
            }
        record_id = allocate_rollout_entry_id(db, counter)
        payload = normalize_rollout_row(data)
        payload["ID"] = record_id
        payload["submission_key"] = submission_key
        payload["Related to XBOX"] = rollout_xbox_key(xbox)
        payload["entry time"] = datetime.now(TRIPOLI_TZ).strftime("%Y-%m-%d %H:%M:%S")
        payload["cable code"] = raw_code if code_type == "cable" else ""
        payload["box code"] = raw_code if code_type == "box" else ""
        payload["staus"] = payload["staus"] or "Done"
        row, _ = upsert_rollout_record(payload, db)
        db.commit()
        clear_warehouse_cache()
        clear_rollout_db_cache()
        db.refresh(row)
    records, _ = rollout_daily_progress_records(db, force=False)
    return {
        "success": True,
        "message": "Field entry saved",
        "record": row_to_record(row),
        "summary": rollout_entry_summary(records),
        "warnings": warnings,
    }


@app.post("/api/warehouse/rollout-field-entry/hub-accessories")
def save_rollout_hub_accessories(data: dict, request: Request, db: Session = Depends(db_session)):
    require_roles(request, "Requester", "Admin")

    submission_key = str(first_value(data, "submission_key", default="") or "").strip()
    if not re.fullmatch(r"[A-Za-z0-9-]{16,128}", submission_key):
        raise HTTPException(status_code=400, detail="Invalid accessory submission key")

    area = str(first_value(data, "Area", "area", default="") or "").strip()
    xbox = str(first_value(data, "Related to XBOX", "related_to_xbox", default="") or "").strip()
    hub_code = str(first_value(data, "hub_code", "hub", default="") or "").strip()
    if not area or not xbox or not hub_code:
        raise HTTPException(status_code=400, detail="Select Area, Related to XBOX, and Hub")

    hub_matches = [
        row
        for row in rollout_reference_matches(area, xbox, hub_code, "box", db, getattr(request.state, "program", DEFAULT_PROGRAM))
        if "hub" in rollout_norm(row.get("box_type") or "")
    ]
    if not hub_matches and not rollout_manual_hub_allowed(area, xbox, hub_code):
        raise HTTPException(status_code=400, detail="Selected Hub is not listed for this Area / XBOX")

    allowed_materials = {
        rollout_norm("S' type clamp"): "S' type clamp",
        rollout_norm("Metal wedge clamping"): "Metal wedge clamping",
        rollout_norm("Plastic Cable Storing Assembly"): "Plastic Cable Storing Assembly",
        rollout_norm("Plum ring hook"): "Plum ring hook",
        rollout_norm("Pole mounting assembly"): "Pole mounting assembly",
    }
    requested = first_value(data, "accessories", default=[])
    if not isinstance(requested, list):
        raise HTTPException(status_code=400, detail="Accessories must be a list")
    accessories: list[tuple[str, float]] = []
    seen_materials: set[str] = set()
    for entry in requested:
        if not isinstance(entry, dict):
            continue
        material = allowed_materials.get(rollout_norm(first_value(entry, "material", default="")))
        quantity = safe_float(first_value(entry, "quantity", "actual", default=0))
        if not material or quantity <= 0 or material in seen_materials:
            continue
        seen_materials.add(material)
        accessories.append((material, quantity))
    if not accessories:
        raise HTTPException(status_code=400, detail="Enter at least one accessory quantity")

    saved_rows: list[RolloutRecord] = []
    entry_time = datetime.now(TRIPOLI_TZ).strftime("%Y-%m-%d %H:%M:%S")
    user_notes = str(first_value(data, "Notes", "notes", default="") or "").strip()
    hub_note = f"Hub: {hub_code}"
    notes = f"{hub_note} | {user_notes}" if user_notes else hub_note
    with ROLLOUT_SYNC_LOCK:
        counter = rollout_entry_counter(db)
        for index, (material, quantity) in enumerate(accessories, start=1):
            item_submission_key = f"{submission_key}-ha-{index}"
            existing = db.query(RolloutRecord).filter(RolloutRecord.submission_key == item_submission_key).first()
            if existing is not None:
                saved_rows.append(existing)
                continue
            payload = normalize_rollout_row(data)
            payload.update(
                {
                    "ID": allocate_rollout_entry_id(db, counter),
                    "submission_key": item_submission_key,
                    "Related to XBOX": rollout_xbox_key(xbox),
                    "item": "Hub Accessories",
                    "material type": material,
                    "actual": quantity,
                    "stock remaining": 0,
                    "entry time": entry_time,
                    "cable code": "",
                    "box code": "",
                    "staus": str(first_value(data, "staus", "status", default="Done") or "Done"),
                    "Notes": notes,
                }
            )
            row, _ = upsert_rollout_record(payload, db)
            saved_rows.append(row)
        db.commit()
        clear_warehouse_cache()
        clear_rollout_db_cache()
        for row in saved_rows:
            db.refresh(row)

    records, _ = rollout_daily_progress_records(db, force=False)
    return {
        "success": True,
        "message": "Hub accessories saved",
        "records": [row_to_record(row) for row in saved_rows],
        "summary": rollout_entry_summary(records),
    }


@app.patch("/api/warehouse/rollout-field-entry/{record_id}")
def edit_rollout_field_entry(record_id: str, data: dict, request: Request, db: Session = Depends(db_session)):
    require_roles(request, "Admin")

    row = db.query(RolloutRecord).filter(RolloutRecord.record_id == str(record_id).strip()).first()
    if row is None:
        raise HTTPException(status_code=404, detail="Field Entry record was not found")

    date = str(first_value(data, "Date", "date", default=row.date) or "").strip()
    supervisor_name = str(first_value(data, "Supervisor Name", "supervisor_name", default=row.supervisor_name) or "").strip()
    team_leader = str(first_value(data, "team leader", "team_leader", default=row.team_leader) or "").strip()
    city = str(first_value(data, "city", "City", default=row.city) or "").strip()
    area = str(first_value(data, "Area", "area", default=row.area) or "").strip()
    activity = str(first_value(data, "Activity", "activity", default=row.activity) or "").strip()
    xbox = str(first_value(data, "Related to XBOX", "related_to_xbox", default=row.related_to_xbox) or "").strip()
    item = str(first_value(data, "item", "Item", default=row.item) or "").strip()
    material_type = str(first_value(data, "material type", "Material Type", "material_type", default=row.material_type) or "").strip()
    mount_type = str(first_value(data, "mount type", "Mount Type", "mount_type", default=row.mount_type) or "").strip()
    item_serial = str(first_value(data, "item serial", "Item Serial", "item_serial", default=row.item_serial) or "").strip()
    planned_quantity = safe_float(first_value(data, "planed quantity", "planned quantity", "planned_quantity", default=row.planned_quantity))
    actual = safe_float(first_value(data, "actual", "Actual", default=row.actual))
    stock_remaining = safe_float(first_value(data, "stock remaining", "stock_remaining", default=row.stock_remaining))
    status = str(first_value(data, "staus", "status", "Status", default=row.status) or "").strip()
    laser = str(first_value(data, "laser", "Laser", default=row.laser) or "").strip()
    acceptance = str(first_value(data, "acceptance", "Acceptance", default=row.acceptance) or "").strip()
    scan = str(first_value(data, "scan", "Scan", default=row.scan) or "").strip()
    labeling = str(first_value(data, "labeling", "Labeling", default=row.labeling) or "").strip()
    olt = str(first_value(data, "OLT", "olt", default=row.olt) or "").strip()
    cable_route = str(first_value(data, "Cable route", "cable route", "cable_route", default=row.cable_route) or "").strip()
    notes = str(first_value(data, "Notes", "notes", default=row.notes) or "").strip()
    code_type = rollout_entry_mode({"code_type": first_value(data, "code_type", "type", default=""), "item": item, "material_type": material_type})
    raw_code = str(first_value(data, "code", "Cable Code", "cable code", "cable_code", "Box Code", "box code", "box_code", default="") or "").strip()

    if not area or not xbox:
        raise HTTPException(status_code=400, detail="This record has no Area or Related to XBOX")
    if code_type not in {"cable", "box", "accessory"}:
        raise HTTPException(status_code=400, detail="Select a material type")
    if actual <= 0:
        raise HTTPException(status_code=400, detail="Actual quantity must be greater than zero")

    matches: list[dict] = []
    duplicates: list[RolloutRecord] = []
    if code_type in {"cable", "box"}:
        if not raw_code:
            raise HTTPException(status_code=400, detail="Select a code from the list")
        matches = rollout_reference_matches(area, xbox, raw_code, code_type, db, getattr(request.state, "program", DEFAULT_PROGRAM))
        if not matches:
            raise HTTPException(status_code=400, detail="Code is not listed for this Area / XBOX / Type")

        code_attr = RolloutRecord.cable_code if code_type == "cable" else RolloutRecord.box_code
        for candidate in db.query(RolloutRecord).filter(code_attr != "").all():
            if candidate.record_id == row.record_id or rollout_record_code_type(candidate) != code_type:
                continue
            if rollout_area_key(candidate.area) != rollout_area_key(area):
                continue
            if rollout_xbox_key(candidate.related_to_xbox) != rollout_xbox_key(xbox):
                continue
            saved_code = candidate.cable_code if code_type == "cable" else candidate.box_code
            if rollout_code_key(saved_code) == rollout_code_key(raw_code):
                duplicates.append(candidate)

        done_duplicate = next((candidate for candidate in duplicates if rollout_norm(candidate.status) == "done"), None)
        if done_duplicate:
            raise HTTPException(status_code=400, detail=f"Code already saved as Done in {done_duplicate.record_id}")

    warnings: list[dict] = []
    in_progress = next((candidate for candidate in duplicates if rollout_norm(candidate.status) in {"inprogress", "planned", "blocked"}), None)
    if in_progress:
        warnings.append({"type": "duplicate_in_progress", "message": "Code exists in another non-Done record", "record": rollout_duplicate_details(in_progress)})

    material_norm = rollout_norm(material_type)
    ref = matches[0] if matches else {}
    expected_len = int(ref.get("cable_length_m") or 0)
    if code_type == "cable" and expected_len:
        length_match = re.search(r"(\d+)\s*m", material_type, flags=re.I)
        entered_len = int(length_match.group(1)) if length_match else 0
        if entered_len and entered_len != expected_len:
            warnings.append({"type": "cable_length_mismatch", "message": f"Planned length is {expected_len}m, selected material looks {entered_len}m"})
    if code_type == "box":
        expected_box = rollout_norm(ref.get("box_type") or "")
        if expected_box:
            if "end" in expected_box and "end" not in material_norm:
                warnings.append({"type": "box_type_mismatch", "message": "Map reference is END BOX, selected material is different"})
            if "sub" in expected_box and "sub" not in material_norm:
                warnings.append({"type": "box_type_mismatch", "message": "Map reference is SUB BOX, selected material is different"})

    if warnings and not first_value(data, "confirm_warnings", "confirmed", default=False):
        return {"success": False, "needs_confirmation": True, "warnings": warnings, "message": "Confirm warnings before saving"}

    before = row_to_record(row)
    row.date = date
    row.supervisor_name = supervisor_name
    row.team_leader = team_leader
    row.city = city
    row.area = area
    row.activity = activity
    row.related_to_xbox = xbox
    row.item = item
    row.material_type = material_type
    row.mount_type = mount_type
    row.item_serial = item_serial
    row.planned_quantity = planned_quantity
    row.actual = actual
    row.stock_remaining = stock_remaining
    row.status = status
    row.laser = laser
    row.acceptance = acceptance
    row.scan = scan
    row.labeling = labeling
    row.olt = olt
    row.cable_route = cable_route
    row.notes = notes
    row.cable_code = raw_code if code_type == "cable" else ""
    row.box_code = raw_code if code_type == "box" else ""
    after = row_to_record(row)
    log_audit(
        db,
        "edit_rollout_field_entry",
        "rollout_record",
        row.record_id,
        request_actor(request),
        {"program": request.state.program, "before": before, "after": after},
    )
    db.commit()
    clear_warehouse_cache()
    clear_rollout_db_cache()
    db.refresh(row)
    records, _ = rollout_daily_progress_records(db, force=False)
    return {"success": True, "message": "Field Entry updated", "record": row_to_record(row), "summary": rollout_entry_summary(records), "warnings": warnings}


@app.delete("/api/warehouse/rollout-field-entry/{record_id}")
def delete_rollout_field_entry(record_id: str, data: dict, request: Request, db: Session = Depends(db_session)):
    require_roles(request, "Admin")

    row = db.query(RolloutRecord).filter(RolloutRecord.record_id == str(record_id).strip()).first()
    if row is None:
        raise HTTPException(status_code=404, detail="Field Entry record was not found")

    deleted = row_to_record(row)
    log_audit(
        db,
        "delete_rollout_field_entry",
        "rollout_record",
        row.record_id,
        request_actor(request),
        {"program": request.state.program, "deleted": deleted},
    )
    db.delete(row)
    db.commit()
    clear_warehouse_cache()
    clear_rollout_db_cache()
    records, _ = rollout_daily_progress_records(db, force=False)
    return {"success": True, "message": "Field Entry deleted", "record_id": str(record_id).strip(), "summary": rollout_entry_summary(records)}


@app.get("/api/warehouse/rollout-source-check")
def rollout_source_check(request: Request, program: str = DEFAULT_PROGRAM, db: Session = Depends(db_session)):
    require_roles(request, "Admin", "Management")
    if is_single_ran(program):
        return {
            "success": True,
            "checked_at": datetime.now(TRIPOLI_TZ).isoformat(),
            "sources": [],
            "source": "disabled",
        }
    rows, _source = rollout_daily_progress_records(db, force=False)
    rows = rollout_records_for_session(request, rows)
    latest = max(rows, key=lambda row: str(row.get("entry time") or row.get("Date") or ""), default={})
    return {
        "success": True,
        "checked_at": datetime.now(TRIPOLI_TZ).isoformat(),
        "source": "database",
        "sources": [
            {
                "source": "supabase",
                "ok": True,
                "count": len(rows),
                "latest_id": latest.get("ID") or latest.get("id") or "",
                "latest_date": str(latest.get("Date") or latest.get("date") or "")[:10],
            }
        ],
    }


@app.get("/api/warehouse/summary")
def warehouse_summary(program: str = DEFAULT_PROGRAM, db: Session = Depends(db_session)):
    program_key = normalize_program(program)
    inventory_receive_total = (
        db.query(func.coalesce(func.sum(StockMovement.quantity), 0))
        .filter(
            StockMovement.program == program_key,
            StockMovement.movement_type == "receive",
            StockMovement.note.like("Inventory receive:%"),
        )
        .scalar()
    )
    return {
        "success": True,
        "program": program_key,
        "warehouses": db.query(Warehouse).filter(Warehouse.program == program_key).count(),
        "technicians": db.query(Technician).filter(Technician.program == program_key).count(),
        "products": db.query(Product).filter(Product.program == program_key).count(),
        "stock_movements": db.query(StockMovement).filter(StockMovement.program == program_key).count(),
        "open_serials": db.query(ProductSerial).filter(ProductSerial.program == program_key, ProductSerial.status.in_(["in_warehouse", "with_technician"])).count(),
        "inventory_receive_total": float(inventory_receive_total or 0),
    }


@app.get("/api/warehouse/bootstrap")
def warehouse_bootstrap(request: Request, light: bool = False, viewer: str = "", role: str = "", program: str = DEFAULT_PROGRAM, db: Session = Depends(db_session)):
    program_key = normalize_program(program)
    viewer = request_scope_viewer(request)
    role = current_user(request).role
    cache_key = f"{program_key}:{'light' if light else 'full'}:{normalize_usage_key(role)}:{normalize_usage_key(viewer)}"
    cached = WAREHOUSE_CACHE.get(cache_key)
    if cached and time.monotonic() - cached[0] < WAREHOUSE_CACHE_TTL:
        return {**cached[1], "cached": True}

    stock = list_stock_balances(request, program_key, db)
    usage = list_stock_usage(request, program_key, db)
    movements = list_stock_movements(request, 12 if light else 40, program_key, db)
    mrs = (
        list_material_requisition_headers(80, db=db, viewer=viewer, role=role, program=program_key)
        if light
        else list_material_requisitions(request, 200, viewer=viewer, role=role, program=program_key, db=db)
    )
    receipts = list_receive_order_headers(12, program_key, db) if light else list_receive_orders(request, 60, program_key, db)
    transfers = (
        list_material_transfer_headers(120, db=db, viewer=viewer, role=role, program=program_key)
        if light
        else list_material_transfers(request, 200, viewer=viewer, role=role, program=program_key, db=db)
    )
    returns = list_material_return_headers(120, program_key, db) if light else list_material_returns(request, 200, program_key, db)
    scans = list_material_scans(request, 80 if light else 300, program_key, db)
    payload = {
        "success": True,
        "program": program_key,
        "programLabel": PROGRAM_LABELS[program_key],
        "partial": light,
        "summary": warehouse_summary(program_key, db),
        "warehouses": list_warehouses(program_key, db)["warehouses"],
        "sites": list_sites(program_key, db)["sites"],
        "technicians": list_technicians(program_key, db)["technicians"],
        "products": list_products(program_key, db)["products"],
        "stockBalances": stock["balances"],
        "stockUsage": usage["usage"],
        "movements": movements["movements"],
        "mrs": mrs["requisitions"],
        "receipts": receipts["receipts"],
        "transfers": transfers["transfers"],
        "returns": returns["returns"],
        "scanLogs": scans["scans"],
    }
    if not light:
        tech = list_technician_balances(request, program_key, db)
        audit = list_audit_logs(request, 40, program_key, db) if current_user(request).role.strip().lower() == "admin" else {"logs": []}
        payload.update(
            {
                "technicianBalances": tech["balances"],
                "audit": audit["logs"],
            }
        )
    WAREHOUSE_CACHE[cache_key] = (time.monotonic(), payload)
    return payload


@app.get("/api/warehouse/warehouses")
def list_warehouses(program: str = DEFAULT_PROGRAM, db: Session = Depends(db_session)):
    program_key = normalize_program(program)
    rows = db.query(Warehouse).filter(Warehouse.program == program_key).order_by(Warehouse.name).all()
    return {"success": True, "warehouses": [{"id": r.id, "program": program_key, "name": r.name, "location": r.location, "status": r.status} for r in rows]}


@app.get("/api/warehouse/sites")
def list_sites(program: str = DEFAULT_PROGRAM, db: Session = Depends(db_session)):
    program_key = normalize_program(program)
    rows = db.query(Site).filter(Site.program == program_key).order_by(Site.name).all()
    return {"success": True, "sites": [{"id": r.id, "program": program_key, "name": r.name} for r in rows]}


@app.post("/api/warehouse/sites")
def create_site(data: SiteIn, request: Request, db: Session = Depends(db_session)):
    require_roles(request, "Admin")
    program_key = normalize_program(data.program)
    name = data.name.strip()
    if not name:
        raise HTTPException(status_code=400, detail="Site ID is required")
    row = Site(program=program_key, name=name)
    db.add(row)
    try:
        db.commit()
    except Exception as exc:
        db.rollback()
        raise HTTPException(status_code=400, detail="Site ID already exists") from exc
    db.refresh(row)
    clear_warehouse_cache()
    return {"success": True, "site": {"id": row.id, "program": program_key, "name": row.name}}


def resolved_site_name(db: Session, program: str, site_value: str) -> str:
    value = str(site_value or "").strip()
    if not value.isdigit():
        return value
    site = db.query(Site).filter(Site.id == int(value), Site.program == normalize_program(program)).first()
    return site.name if site else value


@app.post("/api/warehouse/warehouses")
def create_warehouse(data: WarehouseIn, db: Session = Depends(db_session)):
    program_key = normalize_program(data.program)
    row = Warehouse(program=program_key, name=data.name.strip(), location=data.location.strip())
    db.add(row)
    try:
        db.commit()
    except Exception as exc:
        db.rollback()
        raise HTTPException(status_code=400, detail="Warehouse name already exists") from exc
    db.refresh(row)
    clear_warehouse_cache()
    return {"success": True, "warehouse": {"id": row.id, "program": program_key, "name": row.name, "location": row.location, "status": row.status}}


@app.get("/api/warehouse/technicians")
def list_technicians(program: str = DEFAULT_PROGRAM, db: Session = Depends(db_session)):
    program_key = normalize_program(program)
    rows = db.query(Technician).filter(Technician.program == program_key).order_by(Technician.name).all()
    return {"success": True, "technicians": [{"id": r.id, "program": program_key, "name": r.name, "phone": r.phone, "status": r.status} for r in rows]}


@app.post("/api/warehouse/technicians")
def create_technician(data: TechnicianIn, db: Session = Depends(db_session)):
    program_key = normalize_program(data.program)
    row = Technician(program=program_key, name=data.name.strip(), phone=data.phone.strip())
    db.add(row)
    db.commit()
    db.refresh(row)
    clear_warehouse_cache()
    return {"success": True, "technician": {"id": row.id, "program": program_key, "name": row.name, "phone": row.phone, "status": row.status}}


@app.get("/api/warehouse/products")
def list_products(program: str = DEFAULT_PROGRAM, db: Session = Depends(db_session)):
    program_key = normalize_program(program)
    rows = db.query(Product).filter(Product.program == program_key).order_by(Product.sku).all()
    return {"success": True, "products": [product_to_dict(r) for r in rows]}


def scan_code_candidates(scanned: str) -> list[str]:
    candidates = [scanned]
    if re.fullmatch(r"\d+\.0+", scanned):
        candidates.append(scanned.split(".", 1)[0])
    elif re.fullmatch(r"\d+", scanned):
        candidates.append(f"{scanned}.00")
    return list(dict.fromkeys(candidates))


@app.get("/api/warehouse/scan-material")
def scan_material(code: str, warehouse_id: int | None = None, program: str = DEFAULT_PROGRAM, db: Session = Depends(db_session)):
    program_key = normalize_program(program)
    scanned = code.strip()
    if not scanned:
        raise HTTPException(status_code=400, detail="Scan code is required")
    candidates = scan_code_candidates(scanned)

    serial_row = (
        db.query(ProductSerial)
        .options(joinedload(ProductSerial.product))
        .filter(ProductSerial.program == program_key, ProductSerial.serial_number.in_(candidates))
        .first()
    )
    product = serial_row.product if serial_row else None
    match_type = "serial_number" if serial_row else ""

    if product is None:
        product = (
            db.query(Product)
            .filter(
                Product.program == program_key,
                or_(
                    Product.qr_code.in_(candidates),
                    Product.sku.in_(candidates),
                    Product.name.in_(candidates),
                    Product.item_detail.in_(candidates),
                )
            )
            .first()
        )
        if product:
            if product.qr_code in candidates:
                match_type = "qr_code"
            elif product.sku in candidates:
                match_type = "sku"
            elif product.name in candidates:
                match_type = "name"
            else:
                match_type = "item_detail"

    if product is None:
        raise HTTPException(status_code=404, detail=f"Material not found for scan: {scanned}")

    balance = None
    balances = []
    if warehouse_id:
        balance = (
            db.query(StockBalance)
            .filter(StockBalance.warehouse_id == warehouse_id, StockBalance.product_id == product.id, StockBalance.program == program_key)
            .first()
        )
    elif serial_row and serial_row.warehouse_id:
        warehouse_id = serial_row.warehouse_id
        balance = (
            db.query(StockBalance)
            .filter(StockBalance.warehouse_id == warehouse_id, StockBalance.product_id == product.id, StockBalance.program == program_key)
            .first()
        )
    else:
        balances = (
            db.query(StockBalance)
            .options(joinedload(StockBalance.warehouse))
            .filter(StockBalance.product_id == product.id, StockBalance.program == program_key, StockBalance.quantity > 0)
            .order_by(StockBalance.quantity.desc())
            .all()
        )
        if balances:
            balance = balances[0]
            warehouse_id = balance.warehouse_id

    return {
        "success": True,
        "scan": scanned,
        "match_type": match_type,
        "product": product_to_dict(product),
        "serial": {
            "serial_number": serial_row.serial_number,
            "status": serial_row.status,
            "warehouse_id": serial_row.warehouse_id,
            "technician_id": serial_row.technician_id,
        }
        if serial_row
        else None,
        "balance": balance.quantity if balance else 0,
        "warehouse_id": warehouse_id,
        "warehouse": balance.warehouse.name if balance and balance.warehouse else "",
        "balances": [
            {
                "warehouse_id": b.warehouse_id,
                "warehouse": b.warehouse.name if b.warehouse else "",
                "quantity": b.quantity,
            }
            for b in balances
        ],
    }


@app.get("/api/warehouse/material-scans")
def list_material_scans(request: Request, limit: int = 300, program: str = DEFAULT_PROGRAM, db: Session = Depends(db_session)):
    program_key = normalize_program(program)
    query = (
        db.query(MaterialScanLog)
        .options(
            joinedload(MaterialScanLog.requisition),
            joinedload(MaterialScanLog.product),
            joinedload(MaterialScanLog.warehouse),
        )
        .filter(MaterialScanLog.program == program_key)
        .order_by(MaterialScanLog.id.desc())
    )
    allowed = allowed_warehouse_ids(request, db, program_key)
    if allowed is not None:
        query = query.filter(MaterialScanLog.warehouse_id.in_(allowed))
    rows = query.limit(min(limit, 500)).all()
    return {"success": True, "scans": [scan_log_to_dict(r) for r in rows]}


@app.post("/api/warehouse/material-scans/scan-record")
def record_general_material_scan(data: MaterialScanIn, request: Request, db: Session = Depends(db_session)):
    require_roles(request, "Admin", "Management", "Warehouse Manager")
    data.actor = request_actor(request)
    program_key = normalize_program(data.program)
    scanned = data.code.strip()
    found = scan_material(scanned, None, program_key, db)
    product = require_product(db, found["product"]["id"], program_key)

    warehouse_id = found.get("warehouse_id")
    balance = float(found.get("balance") or 0)
    if not warehouse_id or balance <= 0:
        raise HTTPException(status_code=404, detail=f"Material found but not available in warehouse stock: {product.sku}")
    require_warehouse_access(request, db, warehouse_id, program_key)

    serial_number = found["serial"]["serial_number"] if found.get("serial") else ""
    if serial_number:
        duplicate = (
            db.query(MaterialScanLog)
            .filter(
                MaterialScanLog.program == program_key,
                MaterialScanLog.material_requisition_id.is_(None),
                MaterialScanLog.serial_number == serial_number,
            )
            .first()
        )
        if duplicate:
            raise HTTPException(status_code=400, detail=f"Serial already scanned in Scan History: {serial_number}")

    log = MaterialScanLog(
        material_requisition_id=None,
        program=program_key,
        product_id=product.id,
        warehouse_id=warehouse_id,
        scan_code=scanned,
        serial_number=serial_number,
        match_type=found["match_type"],
        status="in_stock",
        scanned_by=data.actor.strip() or "system",
        note="Warehouse stock scan",
    )
    db.add(log)
    log_audit(db, "scan_stock_lookup", "product", product.sku, data.actor, {"scan": scanned, "warehouse_id": warehouse_id})
    db.commit()
    db.refresh(log)
    clear_warehouse_cache()
    return {
        "success": True,
        "scan": scan_log_to_dict(log),
        "product": product_to_dict(product),
        "serial": found.get("serial"),
        "match_type": found["match_type"],
        "balance": balance,
        "warehouse": found.get("warehouse", ""),
    }


@app.post("/api/warehouse/material-requisitions/{requisition_id}/scan-record")
def record_material_scan(requisition_id: int, data: MaterialScanIn, request: Request, db: Session = Depends(db_session)):
    require_roles(request, "Admin", "Management", "Warehouse Manager")
    data.actor = request_actor(request)
    program_key = normalize_program(data.program)
    row = (
        db.query(MaterialRequisition)
        .options(selectinload(MaterialRequisition.items).joinedload(MaterialRequisitionItem.product), joinedload(MaterialRequisition.warehouse))
        .filter(MaterialRequisition.id == requisition_id, MaterialRequisition.program == program_key)
        .first()
    )
    if row is None:
        raise HTTPException(status_code=404, detail="MR not found")
    require_warehouse_access(request, db, row.warehouse_id, program_key)

    scanned = data.code.strip()
    found = scan_material(scanned, row.warehouse_id, program_key, db)
    product = require_product(db, found["product"]["id"], program_key)

    requested_qty = sum((item.quantity or 0) for item in row.items if item.product_id == product.id)
    if requested_qty <= 0:
        raise HTTPException(status_code=400, detail=f"Scanned material is not part of MR {row.order_number}")

    serial_number = found["serial"]["serial_number"] if found.get("serial") else ""
    if serial_number:
        duplicate = (
            db.query(MaterialScanLog)
            .filter(
                MaterialScanLog.program == program_key,
                MaterialScanLog.material_requisition_id == row.id,
                MaterialScanLog.serial_number == serial_number,
            )
            .first()
        )
        if duplicate:
            raise HTTPException(status_code=400, detail=f"Serial already scanned for MR {row.order_number}: {serial_number}")

    scanned_qty = (
        db.query(func.count(MaterialScanLog.id))
        .filter(
            MaterialScanLog.program == program_key,
            MaterialScanLog.material_requisition_id == row.id,
            MaterialScanLog.product_id == product.id,
            MaterialScanLog.status == "matched",
        )
        .scalar()
        or 0
    )
    if scanned_qty >= requested_qty:
        raise HTTPException(status_code=400, detail=f"Requested quantity already scanned for {product.sku}")

    log = MaterialScanLog(
        material_requisition_id=row.id,
        program=program_key,
        product_id=product.id,
        warehouse_id=row.warehouse_id,
        scan_code=scanned,
        serial_number=serial_number,
        match_type=found["match_type"],
        status="matched",
        scanned_by=data.actor.strip() or "system",
        note=f"MR {row.order_number}",
    )
    db.add(log)
    log_audit(db, "scan_material", "material_requisition", row.order_number, data.actor, {"scan": scanned, "sku": product.sku})
    db.commit()
    db.refresh(log)
    clear_warehouse_cache()
    return {
        "success": True,
        "scan": scan_log_to_dict(log),
        "product": product_to_dict(product),
        "serial": found.get("serial"),
        "match_type": found["match_type"],
        "balance": found["balance"],
        "scan_count": scanned_qty + 1,
        "requested_qty": requested_qty,
    }


@app.post("/api/warehouse/products")
def create_product(data: ProductIn, request: Request, db: Session = Depends(db_session)):
    require_roles(request, "Admin", "Management", "Warehouse Manager")
    program_key = normalize_program(data.program)
    sku = data.sku.strip()
    name = material_display_name(data.name.strip(), sku)
    row = Product(
        program=program_key,
        sku=sku,
        part_number=product_part_number(sku, data.part_number),
        category=data.category.strip(),
        name=name,
        item_detail=data.item_detail.strip(),
        vendor=data.vendor.strip() if program_key == SINGLE_RAN_PROGRAM else "",
        qr_code=data.qr_code.strip(),
        unit=data.unit.strip() or "PCS",
        tracking_type=data.tracking_type,
        min_stock=data.min_stock,
    )
    db.add(row)
    try:
        db.commit()
    except Exception as exc:
        db.rollback()
        raise HTTPException(status_code=400, detail="Product SKU already exists") from exc
    db.refresh(row)
    clear_warehouse_cache()
    return {"success": True, "product": product_to_dict(row)}


@app.post("/api/warehouse/products/{product_id}/purge")
def purge_product(product_id: int, data: ProductPurgeIn, request: Request, db: Session = Depends(db_session)):
    require_roles(request, "Admin")
    actor = request_actor(request)
    product = db.get(Product, product_id)
    program_key = normalize_program(data.program)
    if product is None or normalize_program(getattr(product, "program", DEFAULT_PROGRAM)) != program_key:
        raise HTTPException(status_code=404, detail="Product not found")

    receive_order_ids = [row[0] for row in db.query(ReceiveOrderItem.receive_order_id).filter(ReceiveOrderItem.product_id == product.id).distinct().all()]
    issue_order_ids = [row[0] for row in db.query(IssueOrderItem.issue_order_id).filter(IssueOrderItem.product_id == product.id).distinct().all()]
    return_order_ids = [row[0] for row in db.query(MaterialReturnItem.return_id).filter(MaterialReturnItem.product_id == product.id).distinct().all()]
    transfer_order_ids = [row[0] for row in db.query(MaterialTransferItem.transfer_id).filter(MaterialTransferItem.product_id == product.id).distinct().all()]
    snapshot = product_to_dict(product)

    db.query(ProductSerial).filter(ProductSerial.product_id == product.id, ProductSerial.program == program_key).delete(synchronize_session=False)
    db.query(StockBalance).filter(StockBalance.product_id == product.id, StockBalance.program == program_key).delete(synchronize_session=False)
    db.query(TechnicianBalance).filter(TechnicianBalance.product_id == product.id, TechnicianBalance.program == program_key).delete(synchronize_session=False)
    db.query(StockMovement).filter(StockMovement.product_id == product.id, StockMovement.program == program_key).delete(synchronize_session=False)
    db.query(MaterialScanLog).filter(MaterialScanLog.product_id == product.id, MaterialScanLog.program == program_key).delete(synchronize_session=False)
    db.query(ReceiveOrderItem).filter(ReceiveOrderItem.product_id == product.id).delete(synchronize_session=False)
    db.query(IssueOrderItem).filter(IssueOrderItem.product_id == product.id).delete(synchronize_session=False)
    db.query(MaterialRequisitionItem).filter(MaterialRequisitionItem.product_id == product.id).delete(synchronize_session=False)
    db.query(MaterialReturnItem).filter(MaterialReturnItem.product_id == product.id).delete(synchronize_session=False)
    db.query(MaterialTransferItem).filter(MaterialTransferItem.product_id == product.id).delete(synchronize_session=False)

    for order_id in receive_order_ids:
        has_items = db.query(ReceiveOrderItem).filter(ReceiveOrderItem.receive_order_id == order_id).first()
        if has_items is None:
            db.query(ReceiveOrder).filter(ReceiveOrder.id == order_id).delete(synchronize_session=False)
    for order_id in issue_order_ids:
        has_items = db.query(IssueOrderItem).filter(IssueOrderItem.issue_order_id == order_id).first()
        if has_items is None:
            db.query(IssueOrder).filter(IssueOrder.id == order_id).delete(synchronize_session=False)
    for order_id in return_order_ids:
        has_items = db.query(MaterialReturnItem).filter(MaterialReturnItem.return_id == order_id).first()
        if has_items is None:
            db.query(MaterialReturn).filter(MaterialReturn.id == order_id).delete(synchronize_session=False)
    for order_id in transfer_order_ids:
        has_items = db.query(MaterialTransferItem).filter(MaterialTransferItem.transfer_id == order_id).first()
        if has_items is None:
            db.query(MaterialTransfer).filter(MaterialTransfer.id == order_id).delete(synchronize_session=False)

    db.delete(product)
    log_audit(db, "purge_product", "product", snapshot["sku"], actor, snapshot)
    db.commit()
    return {"success": True, "deleted": snapshot}


@app.get("/api/warehouse/stock-balances")
def list_stock_balances(request: Request, program: str = DEFAULT_PROGRAM, db: Session = Depends(db_session)):
    program_key = normalize_program(program)
    query = (
        db.query(StockBalance)
        .options(joinedload(StockBalance.warehouse), joinedload(StockBalance.product))
        .filter(StockBalance.program == program_key)
        .order_by(StockBalance.warehouse_id, StockBalance.product_id)
    )
    allowed = allowed_warehouse_ids(request, db, program_key)
    if allowed is not None:
        query = query.filter(StockBalance.warehouse_id.in_(allowed))
    rows = query.all()
    reserved = reserved_stock_quantities(db, program_key)
    return {
        "success": True,
        "balances": [balance_to_dict(r, reserved.get((r.warehouse_id, r.product_id), 0)) for r in rows],
    }


@app.get("/api/warehouse/stock-usage")
def list_stock_usage(request: Request, program: str = DEFAULT_PROGRAM, db: Session = Depends(db_session)):
    program_key = normalize_program(program)
    balances_query = (
        db.query(StockBalance)
        .options(joinedload(StockBalance.warehouse), joinedload(StockBalance.product))
        .filter(StockBalance.program == program_key)
        .order_by(StockBalance.warehouse_id, StockBalance.product_id)
    )
    allowed = allowed_warehouse_ids(request, db, program_key)
    if allowed is not None:
        balances_query = balances_query.filter(StockBalance.warehouse_id.in_(allowed))
    balances = balances_query.all()
    reserved_quantities = reserved_stock_quantities(db, program_key)
    received_totals = {
        (warehouse_id, product_id): total or 0
        for warehouse_id, product_id, total in (
            db.query(StockMovement.warehouse_id, StockMovement.product_id, func.sum(StockMovement.quantity))
            .filter(StockMovement.program == program_key, StockMovement.warehouse_id.isnot(None), StockMovement.movement_type.in_(["receive", "return_in", "transfer_in"]))
            .group_by(StockMovement.warehouse_id, StockMovement.product_id)
            .all()
        )
    }
    consumed_query = (
        db.query(StockMovement.warehouse_id, StockMovement.product_id, func.sum(-StockMovement.quantity))
        .filter(
            StockMovement.warehouse_id.isnot(None),
            StockMovement.program == program_key,
            StockMovement.movement_type == "issue_to_technician",
        )
    )
    consumed_totals = {
        (warehouse_id, product_id): total or 0
        for warehouse_id, product_id, total in consumed_query.group_by(StockMovement.warehouse_id, StockMovement.product_id).all()
    }
    adjustment_totals = {
        (warehouse_id, product_id): total or 0
        for warehouse_id, product_id, total in (
            db.query(StockMovement.warehouse_id, StockMovement.product_id, func.sum(StockMovement.quantity))
            .filter(StockMovement.program == program_key, StockMovement.warehouse_id.isnot(None), StockMovement.movement_type == "adjustment")
            .group_by(StockMovement.warehouse_id, StockMovement.product_id)
            .all()
        )
    }
    rollout_rows, _ = (rollout_daily_progress_records(db) if not is_single_ran(program_key) else ([], "disabled"))
    rollout_consumed_by_warehouse_material: dict[tuple[str, str], float] = {}
    for record in rollout_rows:
        material = str(record.get("material type") or record.get("item") or "").strip()
        if not material:
            continue
        material_key = canonical_material_key(material)
        if not material_key:
            continue
        actual = safe_float(record.get("actual"))
        warehouse_key = rollout_warehouse_key(record)
        if warehouse_key:
            key = (warehouse_key, material_key)
            rollout_consumed_by_warehouse_material[key] = rollout_consumed_by_warehouse_material.get(key, 0) + actual

    usage_rows = []
    for balance in balances:
        key = (balance.warehouse_id, balance.product_id)
        total_received = received_totals.get(key, 0)
        total_consumed = consumed_totals.get(key, 0)
        total_adjustment = adjustment_totals.get(key, 0)
        remaining = balance.quantity or 0
        reserved_pending = float(reserved_quantities.get(key, 0) or 0)
        available_stock = max(float(remaining or 0) - reserved_pending, 0)
        display_total = remaining + total_consumed
        denominator = display_total if display_total > 0 else total_received
        usage_percent = round((total_consumed / denominator) * 100, 2) if denominator else 0
        product_name = product_display_name(balance.product)
        material_key = canonical_material_key(product_name or (balance.product.sku if balance.product else ""))
        warehouse_key = normalize_usage_key(balance.warehouse.name if balance.warehouse else "")
        rollout_consumed = rollout_consumed_by_warehouse_material.get((warehouse_key, material_key), 0)
        remaining_after_rollout = display_total - rollout_consumed
        rollout_usage_percent = round((rollout_consumed / display_total) * 100, 2) if display_total else 0
        usage_rows.append(
            {
                "warehouse_id": balance.warehouse_id,
                "warehouse": balance.warehouse.name if balance.warehouse else "",
                "product_id": balance.product_id,
                "sku": balance.product.sku if balance.product else "",
                "part_number": product_part_number(balance.product.sku if balance.product else "", balance.product.part_number if balance.product else ""),
                "product": product_name,
                "unit": balance.product.unit if balance.product else "",
                "total_received": display_total,
                "received_movements": total_received,
                "total_consumed": total_consumed,
                "total_adjustment": total_adjustment,
                "remaining": remaining,
                "wh_remaining": remaining,
                "reserved_pending": reserved_pending,
                "available_stock": available_stock,
                "usage_percent": usage_percent,
                "rollout_consumed_qty": rollout_consumed,
                "remaining_after_rollout": remaining_after_rollout,
                "rollout_usage_percent": rollout_usage_percent,
            }
        )
    return {"success": True, "usage": usage_rows}


def user_can_view_material_return(row: MaterialReturn, viewer: str = "", role: str = "") -> bool:
    role_key = normalize_usage_key(role)
    viewer_key = normalize_usage_key(viewer)
    if role_key in {"admin", "management"}:
        return True
    if role_key == "warehousemanager":
        return warehouse_scope_matches(viewer, row.warehouse.name if row.warehouse else "")
    if role_key == "requester":
        return normalize_usage_key(row.returned_by) == viewer_key or normalize_usage_key(row.created_by) == viewer_key
    return False


@app.get("/api/warehouse/inventory-export.xlsx")
def export_inventory_excel(request: Request, warehouse: str = "", program: str = DEFAULT_PROGRAM, db: Session = Depends(db_session)):
    require_roles(request, "Admin", "Management", "Warehouse Manager")
    program_key = normalize_program(program)
    selected_warehouse = warehouse.strip()
    rows = list_stock_usage(request, program_key, db)["usage"]
    if selected_warehouse:
        rows = [row for row in rows if row["warehouse"] == selected_warehouse]

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Inventory"
    sheet.append(["Warehouse", "Part #", "SKU", "Item", "Unit", "Total Stock", "Issued WH", "Rollout Used", "WH Remaining", "Reserved Pending", "Available", "Usage %"])
    for row in rows:
        sheet.append(
            [
                row["warehouse"],
                row["part_number"],
                row["sku"],
                row["product"],
                row["unit"],
                row["total_received"],
                row["total_consumed"],
                row["rollout_consumed_qty"],
                row["wh_remaining"],
                row["reserved_pending"],
                row["available_stock"],
                row["usage_percent"] / 100,
            ]
        )
    for cell in sheet[1]:
        cell.font = cell.font.copy(bold=True)
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    sheet.column_dimensions["A"].width = 22
    sheet.column_dimensions["B"].width = 20
    sheet.column_dimensions["C"].width = 18
    sheet.column_dimensions["D"].width = 46
    for column in ("E", "F", "G", "H", "I", "J", "K", "L"):
        sheet.column_dimensions[column].width = 16
    for cell in sheet["L"][1:]:
        cell.number_format = "0.0%"

    output = io.BytesIO()
    workbook.save(output)
    output.seek(0)
    suffix = re.sub(r"[^A-Za-z0-9_-]+", "-", selected_warehouse).strip("-") or "all-warehouses"
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="inventory-{suffix}.xlsx"'},
    )


@app.get("/api/warehouse/technician-balances")
def list_technician_balances(request: Request, program: str = DEFAULT_PROGRAM, db: Session = Depends(db_session)):
    require_roles(request, "Admin", "Management", "Warehouse Manager")
    program_key = normalize_program(program)
    query = (
        db.query(TechnicianBalance)
        .options(joinedload(TechnicianBalance.technician), joinedload(TechnicianBalance.product))
        .filter(TechnicianBalance.program == program_key)
        .order_by(TechnicianBalance.technician_id, TechnicianBalance.product_id)
    )
    allowed = allowed_warehouse_ids(request, db, program_key)
    if allowed is not None:
        technician_ids = [row[0] for row in db.query(IssueOrder.technician_id).filter(IssueOrder.program == program_key, IssueOrder.warehouse_id.in_(allowed)).distinct().all()]
        query = query.filter(TechnicianBalance.technician_id.in_(technician_ids))
    rows = query.all()
    return {"success": True, "balances": [technician_balance_to_dict(r) for r in rows]}


@app.get("/api/warehouse/technician-material-usage")
def list_technician_material_usage(request: Request, program: str = DEFAULT_PROGRAM, db: Session = Depends(db_session)):
    require_roles(request, "Admin", "Management", "Warehouse Manager")
    program_key = normalize_program(program)
    rows: dict[tuple[str, int], dict] = {}
    requisitions_query = (
        db.query(MaterialRequisition)
        .options(selectinload(MaterialRequisition.items).joinedload(MaterialRequisitionItem.product))
        .filter(MaterialRequisition.program == program_key)
        .order_by(MaterialRequisition.id.asc())
    )
    allowed = allowed_warehouse_ids(request, db, program_key)
    if allowed is not None:
        requisitions_query = requisitions_query.filter(MaterialRequisition.warehouse_id.in_(allowed))
    requisitions = requisitions_query.all()
    for requisition in requisitions:
        if requisition.status not in {"issued", "signed"}:
            continue
        technician_name = (requisition.receiver_name or "").strip()
        if not technician_name:
            continue
        area = (requisition.site_id or requisition.site_address or "").strip()
        for item in requisition.items:
            if not item.product_id:
                continue
            product = item.product
            key = (area or technician_name, item.product_id)
            current = rows.setdefault(
                key,
                {
                    "technician": technician_name,
                    "area": area,
                    "site_id": requisition.site_id,
                    "site_address": requisition.site_address,
                    "material": material_display_name(item.description or product_display_name(product), product.sku if product else ""),
                    "sku": item.model or (product.sku if product else ""),
                    "mr_issued_qty": 0,
                    "current_app_balance": 0,
                    "last_mr": "",
                    "last_sync": "",
                },
            )
            current["mr_issued_qty"] += item.quantity or 0
            current["last_mr"] = requisition.order_number

    balances = db.query(TechnicianBalance).filter(TechnicianBalance.program == program_key).all()
    for balance in balances:
        technician_name = balance.technician.name if balance.technician else ""
        if not technician_name:
            continue
        product = balance.product
        key = (technician_name, balance.product_id)
        current = rows.setdefault(
            key,
            {
                "technician": technician_name,
                "area": "",
                "site_id": "",
                "site_address": "",
                "material": product_display_name(product),
                "sku": product.sku if product else "",
                "mr_issued_qty": 0,
                "current_app_balance": 0,
                "last_mr": "",
                "last_sync": "",
            },
        )
        current["current_app_balance"] = balance.quantity or 0
        if not current["material"]:
            current["material"] = product_display_name(product)
        if not current["sku"]:
            current["sku"] = product.sku if product else ""

    today = local_today()
    usage = sorted(rows.values(), key=lambda row: (row["technician"], row["material"]))
    for row in usage:
        row["last_sync"] = today
    return {"success": True, "usage": usage}


def normalize_usage_key(value: str) -> str:
    return "".join(ch.lower() for ch in (value or "") if ch.isalnum())


def rollout_warehouse_key(record: dict) -> str:
    text = normalize_usage_key(f"{record.get('city') or ''} {record.get('Area') or ''}")
    if "freezone" in text:
        return normalize_usage_key("Misurata Free Zone")
    if any(token in text for token in ("misurata", "misrata", "misrat")):
        return normalize_usage_key("Misurata Lnet")
    if "tripoli" in text:
        return normalize_usage_key("Tripoli")
    return ""


def canonical_material_key(value: str) -> str:
    key = normalize_usage_key(value)
    if not key:
        return ""

    length_match = re.search(r"(\d+)m", key)
    length = length_match.group(1) if length_match else ""
    if "dropcable" in key and length:
        return f"dropcable{length}m"
    if "distributioncable" in key and length:
        return f"distributioncable{length}m"
    if "corecable" in key and length and ("4core" in key or key.endswith("4")):
        return f"4corecable{length}m"

    aliases = {
        "subbox": ["subbox", "fat2810ss8a"],
        "endbox": ["endbox", "fat2810se8a"],
        "xbox": ["xbox", "ssc2802tx8b"],
        "hubbox": ["hubbox", "fat2811sh4b"],
        "atb": ["atb", "e00atb101"],
        "bigtail": ["bigtail", "pigtail", "l0524vdd"],
        # Check the longer P1_03 SKU first; it starts with the legacy P1 SKU.
        "metalwedgeclamping": ["metalwedgeclamping", "itc3301p103"],
        "plumringhook": ["plumringhook", "itc3301p1"],
        "stypeclamp": ["stypeclamp", "itc3103a1"],
        "polemountingassembly": ["polemountingassembly", "e00dkba04"],
        "plasticcablestoringassembly": ["plasticcablestoringassembly", "itc2102p2"],
    }
    for canonical, values in aliases.items():
        if any(alias in key for alias in values):
            return canonical
    return key


def is_workflow_role(value: str) -> bool:
    return normalize_usage_key(value) in {"approver", "approval", "admin"}


TECHNICIAN_USAGE_ALIASES = {
    "ali": ["علي قراب", "علي"],
    "hamza": ["حمزه بشايره", "حمزة بشايره", "حمزه", "حمزة"],
    "fathoi": ["فتحي", "fathi", "fathoi"],
}


def technician_usage_keys(name: str) -> set[str]:
    base = normalize_usage_key(name)
    keys = {base} if base else set()
    aliases = TECHNICIAN_USAGE_ALIASES.get(base, [])
    for alias in aliases:
        alias_key = normalize_usage_key(alias)
        if alias_key:
            keys.add(alias_key)
    return keys


def area_usage_keys(site_id: str, site_address: str) -> set[str]:
    source = site_id or site_address
    key = normalize_usage_key(source)
    return {key} if key else set()


def material_ledger_product_options(db: Session, program: str, query: str = "", product_id: int = 0) -> list[Product]:
    if product_id:
        product = db.query(Product).filter(Product.id == product_id, Product.program == normalize_program(program)).first()
        return [product] if product else []
    term = str(query or "").strip()
    if not term:
        return []
    needle = f"%{term.lower()}%"
    rows = (
        db.query(Product)
        .filter(
            Product.program == normalize_program(program),
            or_(
                func.lower(Product.sku).like(needle),
                func.lower(Product.name).like(needle),
                func.lower(Product.part_number).like(needle),
                func.lower(Product.qr_code).like(needle),
            )
        )
        .order_by(Product.name.asc(), Product.sku.asc())
        .limit(20)
        .all()
    )
    normalized = normalize_usage_key(term)
    return sorted(
        rows,
        key=lambda row: (
            normalize_usage_key(row.sku) != normalized,
            normalize_usage_key(row.name) != normalized,
            row.name,
            row.sku,
        ),
    )


def material_ledger_event(
    *,
    date: str,
    event_type: str,
    reference: str,
    warehouse: str = "",
    from_wh: str = "",
    to_wh: str = "",
    area: str = "",
    qty: float = 0,
    status: str = "",
    actor: str = "",
    note: str = "",
) -> dict:
    return {
        "date": str(date or ""),
        "type": event_type,
        "reference": reference,
        "warehouse": warehouse,
        "from": from_wh,
        "to": to_wh,
        "area": area,
        "quantity": qty or 0,
        "status": status,
        "actor": actor,
        "note": note,
    }


def initial_stock_reference_rows() -> list[dict]:
    """Return the FTTH opening stock imported from final_WH_tmp.xlsx."""
    global INITIAL_STOCK_REFERENCE_CACHE
    if INITIAL_STOCK_REFERENCE_CACHE is not None:
        return INITIAL_STOCK_REFERENCE_CACHE
    try:
        with open(INITIAL_STOCK_REFERENCE_PATH, "r", encoding="utf-8") as source:
            data = json.load(source)
        INITIAL_STOCK_REFERENCE_CACHE = list(data.get("rows") or [])
    except (OSError, ValueError, TypeError):
        logger.exception("Unable to load initial warehouse stock reference")
        INITIAL_STOCK_REFERENCE_CACHE = []
    return INITIAL_STOCK_REFERENCE_CACHE


@app.get("/api/warehouse/material-ledger")
def material_ledger(request: Request, query: str = "", product_id: int = 0, force: bool = False, program: str = DEFAULT_PROGRAM, db: Session = Depends(db_session)):
    program_key = normalize_program(program)
    products = material_ledger_product_options(db, program_key, query=query, product_id=product_id)
    selected = products[0] if products else None
    matches = [product_to_dict(row) for row in products]
    if selected is None:
        return {
            "success": True,
            "query": query,
            "matches": matches,
            "material": None,
            "summary": {},
            "stock_by_warehouse": [],
            "events": [],
        }

    events: list[dict] = []
    summary = {
        "total_received": 0,
        "mr_issued": 0,
        "tr_out": 0,
        "tr_in": 0,
        "returned": 0,
        "current_stock": 0,
    }

    balances = (
        db.query(StockBalance)
        .options(joinedload(StockBalance.warehouse))
        .filter(StockBalance.product_id == selected.id, StockBalance.program == program_key)
        .order_by(StockBalance.warehouse_id.asc())
        .all()
    )
    allowed = allowed_warehouse_ids(request, db, program_key)
    if allowed is not None:
        balances = [row for row in balances if row.warehouse_id in allowed]
    stock_by_warehouse = [
        {
            "warehouse_id": row.warehouse_id,
            "warehouse": row.warehouse.name if row.warehouse else "",
            "unit": selected.unit or "PCS",
            "current_stock": float(row.quantity or 0),
            "total_received": 0,
            "mr_issued": 0,
            "tr_out": 0,
            "tr_in": 0,
            "returned": 0,
        }
        for row in balances
    ]
    warehouse_rows = {row["warehouse_id"]: row for row in stock_by_warehouse}

    def warehouse_totals(warehouse_id: int | None, warehouse_name: str = "") -> dict | None:
        if warehouse_id is None or (allowed is not None and warehouse_id not in allowed):
            return None
        if warehouse_id not in warehouse_rows:
            warehouse_rows[warehouse_id] = {
                "warehouse_id": warehouse_id,
                "warehouse": warehouse_name,
                "unit": selected.unit or "PCS",
                "current_stock": 0,
                "total_received": 0,
                "mr_issued": 0,
                "tr_out": 0,
                "tr_in": 0,
                "returned": 0,
            }
            stock_by_warehouse.append(warehouse_rows[warehouse_id])
        return warehouse_rows[warehouse_id]

    summary["current_stock"] = sum(float(row.quantity or 0) for row in balances)

    # The legacy system_import orders are retained in the database temporarily,
    # but the dashboard must use final_WH_tmp.xlsx as its opening-stock source.
    if not is_single_ran(program_key):
        warehouses_by_key = {
            normalize_usage_key(row.name): row
            for row in db.query(Warehouse).filter(Warehouse.program == program_key).all()
        }
        selected_sku = normalize_usage_key(selected.sku)
        for reference in initial_stock_reference_rows():
            if normalize_usage_key(reference.get("sku")) != selected_sku:
                continue
            warehouse = warehouses_by_key.get(normalize_usage_key(reference.get("warehouse")))
            if warehouse is None or (allowed is not None and warehouse.id not in allowed):
                continue
            qty = safe_float(reference.get("quantity"))
            summary["total_received"] += qty
            totals = warehouse_totals(warehouse.id, warehouse.name)
            if totals is not None:
                totals["total_received"] += qty
            events.append(
                material_ledger_event(
                    date="Reference",
                    event_type="Initial Stock",
                    reference="final_WH_tmp.xlsx",
                    warehouse=warehouse.name,
                    qty=qty,
                    status="confirmed",
                    actor="Reference",
                    note=reference.get("vendor") or "",
                )
            )

    receipts_query = (
        db.query(ReceiveOrder)
        .options(joinedload(ReceiveOrder.warehouse), selectinload(ReceiveOrder.items))
        .join(ReceiveOrderItem, ReceiveOrderItem.receive_order_id == ReceiveOrder.id)
        .filter(
            ReceiveOrderItem.product_id == selected.id,
            ReceiveOrder.program == program_key,
            ReceiveOrder.created_by != "system_import",
        )
        .order_by(ReceiveOrder.id.asc())
    )
    if allowed is not None:
        receipts_query = receipts_query.filter(ReceiveOrder.warehouse_id.in_(allowed))
    receipts = receipts_query.all()
    for receipt in receipts:
        qty = sum(float(item.quantity or 0) for item in receipt.items if item.product_id == selected.id)
        if receipt.status == "confirmed":
            summary["total_received"] += qty
            totals = warehouse_totals(receipt.warehouse_id, receipt.warehouse.name if receipt.warehouse else "")
            if totals is not None:
                totals["total_received"] += qty
        events.append(
            material_ledger_event(
                date=receipt.receipt_date or (receipt.created_at.date().isoformat() if receipt.created_at else ""),
                event_type="GRN",
                reference=receipt.order_number,
                warehouse=receipt.warehouse.name if receipt.warehouse else "",
                qty=qty,
                status=receipt.status,
                actor=receipt.created_by,
                note=receipt.supplier,
            )
        )

    requisitions_query = (
        db.query(MaterialRequisition)
        .options(joinedload(MaterialRequisition.warehouse), selectinload(MaterialRequisition.items))
        .join(MaterialRequisitionItem, MaterialRequisitionItem.requisition_id == MaterialRequisition.id)
        .filter(MaterialRequisitionItem.product_id == selected.id, MaterialRequisition.program == program_key)
        .order_by(MaterialRequisition.id.asc())
    )
    if allowed is not None:
        requisitions_query = requisitions_query.filter(MaterialRequisition.warehouse_id.in_(allowed))
    requisitions = requisitions_query.all()
    for requisition in requisitions:
        qty = sum(float(item.quantity or 0) for item in requisition.items if item.product_id == selected.id)
        if requisition.status in {"issued", "signed"}:
            summary["mr_issued"] += qty
            totals = warehouse_totals(requisition.warehouse_id, requisition.warehouse.name if requisition.warehouse else "")
            if totals is not None:
                totals["mr_issued"] += qty
            events.append(
                material_ledger_event(
                    date=requisition.creation_date or (requisition.created_at.date().isoformat() if requisition.created_at else ""),
                    event_type="MR",
                    reference=requisition.order_number,
                    warehouse=requisition.warehouse.name if requisition.warehouse else "",
                    area=requisition.site_id or requisition.site_address,
                    qty=qty,
                    status=requisition.status,
                    actor=requisition.requester_name or requisition.created_by,
                    note=requisition.team_leader or requisition.receiver_name,
                )
            )

    transfers_query = (
        db.query(MaterialTransfer)
        .options(
            joinedload(MaterialTransfer.from_warehouse),
            joinedload(MaterialTransfer.to_warehouse),
            selectinload(MaterialTransfer.items),
        )
        .join(MaterialTransferItem, MaterialTransferItem.transfer_id == MaterialTransfer.id)
        .filter(MaterialTransferItem.product_id == selected.id, MaterialTransfer.program == program_key)
        .order_by(MaterialTransfer.id.asc())
    )
    if allowed is not None:
        transfers_query = transfers_query.filter(or_(MaterialTransfer.from_warehouse_id.in_(allowed), MaterialTransfer.to_warehouse_id.in_(allowed)))
    transfers = transfers_query.all()
    for transfer in transfers:
        qty = sum(float(item.quantity or 0) for item in transfer.items if item.product_id == selected.id)
        if transfer.status == "transferred":
            summary["tr_out"] += qty
            summary["tr_in"] += qty
            from_name = transfer.from_warehouse.name if transfer.from_warehouse else ""
            to_name = transfer.to_warehouse.name if transfer.to_warehouse else ""
            from_totals = warehouse_totals(transfer.from_warehouse_id, from_name)
            to_totals = warehouse_totals(transfer.to_warehouse_id, to_name)
            if from_totals is not None:
                from_totals["tr_out"] += qty
            if to_totals is not None:
                to_totals["tr_in"] += qty
            common = {
                "date": transfer.transfer_date or (transfer.created_at.date().isoformat() if transfer.created_at else ""),
                "reference": transfer.transfer_number,
                "from_wh": from_name,
                "to_wh": to_name,
                "qty": qty,
                "status": transfer.status,
                "actor": transfer.requester_name or transfer.created_by,
                "note": transfer.reason or transfer.reference_no,
            }
            events.append(material_ledger_event(event_type="TR Out", warehouse=common["from_wh"], **common))
            events.append(material_ledger_event(event_type="TR In", warehouse=common["to_wh"], **common))

    returns_query = (
        db.query(MaterialReturn)
        .options(joinedload(MaterialReturn.warehouse), selectinload(MaterialReturn.items))
        .join(MaterialReturnItem, MaterialReturnItem.return_id == MaterialReturn.id)
        .filter(MaterialReturnItem.product_id == selected.id, MaterialReturn.program == program_key)
        .order_by(MaterialReturn.id.asc())
    )
    if allowed is not None:
        returns_query = returns_query.filter(MaterialReturn.warehouse_id.in_(allowed))
    returns = returns_query.all()
    for returned in returns:
        qty = sum(float(item.quantity or 0) for item in returned.items if item.product_id == selected.id)
        if returned.status == "confirmed":
            summary["returned"] += qty
            totals = warehouse_totals(returned.warehouse_id, returned.warehouse.name if returned.warehouse else "")
            if totals is not None:
                totals["returned"] += qty
            events.append(
                material_ledger_event(
                    date=returned.return_date or (returned.created_at.date().isoformat() if returned.created_at else ""),
                    event_type="RN",
                    reference=returned.return_number,
                    warehouse=returned.warehouse.name if returned.warehouse else "",
                    area=returned.site_id or returned.site_address,
                    qty=qty,
                    status=returned.status,
                    actor=returned.returned_by or returned.created_by,
                    note=returned.reason,
                )
            )

    events.sort(key=lambda row: (row.get("date") or "", row.get("reference") or "", row.get("type") or ""), reverse=True)
    stock_by_warehouse.sort(key=lambda row: row["warehouse"])
    return {
        "success": True,
        "query": query,
        "matches": matches,
        "material": product_to_dict(selected),
        "summary": summary,
        "stock_by_warehouse": stock_by_warehouse,
        "events": events[:1000],
    }


@app.get("/api/warehouse/rollout-material-usage")
def list_rollout_material_usage(request: Request, db: Session = Depends(db_session), force: bool = False, program: str = DEFAULT_PROGRAM):
    if is_single_ran(program):
        return {"success": True, "usage": [], "rollout_records": 0, "rollout_source": "disabled"}
    program_key = normalize_program(program)
    rollout_rows, rollout_source = rollout_daily_progress_records(db, force=force)
    grouped: dict[tuple[str, str], dict] = {}

    def usage_row(area: str, material_key: str, material: str = "", sku: str = "") -> dict:
        key = (area, material_key)
        if key not in grouped:
            grouped[key] = {
                "area": area,
                "material": material or material_key,
                "sku": sku,
                "mr_issued_qty": 0.0,
                "rollout_used_qty": 0.0,
                "returned_qty": 0.0,
                "warehouses": set(),
            }
        row = grouped[key]
        if material and not row["material"]:
            row["material"] = material
        if sku and not row["sku"]:
            row["sku"] = sku
        return row

    allowed = allowed_warehouse_ids(request, db, program_key)
    requisitions_query = (
        db.query(MaterialRequisition)
        .options(joinedload(MaterialRequisition.warehouse), selectinload(MaterialRequisition.items).joinedload(MaterialRequisitionItem.product))
        .filter(MaterialRequisition.program == program_key, MaterialRequisition.status.in_(["issued", "signed"]))
        .order_by(MaterialRequisition.id.asc())
    )
    if allowed is not None:
        requisitions_query = requisitions_query.filter(MaterialRequisition.warehouse_id.in_(allowed))
    for requisition in requisitions_query.all():
        area = canonical_area_name(requisition.site_id or requisition.site_address)
        if not area:
            continue
        for item in requisition.items:
            qty = float(item.quantity or 0)
            material = product_display_name(item.product) if item.product else str(item.description or "")
            material_key = canonical_material_key(material or (item.product.sku if item.product else ""))
            if qty <= 0 or not material_key:
                continue
            row = usage_row(area, material_key, material, item.product.sku if item.product else item.model)
            row["mr_issued_qty"] += qty
            if requisition.warehouse and requisition.warehouse.name:
                row["warehouses"].add(requisition.warehouse.name)

    returns_query = (
        db.query(MaterialReturn)
        .options(joinedload(MaterialReturn.warehouse), selectinload(MaterialReturn.items).joinedload(MaterialReturnItem.product))
        .filter(MaterialReturn.program == program_key, MaterialReturn.status == "confirmed")
        .order_by(MaterialReturn.id.asc())
    )
    if allowed is not None:
        returns_query = returns_query.filter(MaterialReturn.warehouse_id.in_(allowed))
    for returned in returns_query.all():
        area = canonical_area_name(returned.site_id or returned.site_address)
        if not area:
            continue
        for item in returned.items:
            qty = float(item.quantity or 0)
            material = product_display_name(item.product) if item.product else str(item.description or "")
            material_key = canonical_material_key(material or (item.product.sku if item.product else ""))
            if qty <= 0 or not material_key:
                continue
            row = usage_row(area, material_key, material, item.product.sku if item.product else "")
            row["returned_qty"] += qty
            if returned.warehouse and returned.warehouse.name:
                row["warehouses"].add(returned.warehouse.name)

    for record in rollout_rows:
        area = canonical_area_name(str(record.get("Area") or record.get("area") or "").strip())
        material = str(record.get("material type") or record.get("item") or "").strip()
        if not area or not material:
            continue
        status = normalize_usage_key(str(record.get("status") or record.get("staus") or ""))
        if status and status not in {"done", "completed", "installed"}:
            continue
        actual = safe_float(record.get("actual"))
        material_key = canonical_material_key(material)
        if actual <= 0 or not material_key:
            continue
        usage_row(area, material_key, material)["rollout_used_qty"] += actual

    rows = []
    for row in grouped.values():
        issued = float(row["mr_issued_qty"] or 0)
        installed = float(row["rollout_used_qty"] or 0)
        returned = float(row["returned_qty"] or 0)
        difference = issued - installed - returned
        net_issued = issued - returned
        if difference < 0:
            usage_match = "over_mr"
        elif installed or returned:
            usage_match = "area"
        else:
            usage_match = "none"
        rows.append(
            {
                "area": row["area"],
                "material": row["material"],
                "sku": row["sku"],
                "warehouse": ", ".join(sorted(row["warehouses"])),
                "mr_issued_qty": issued,
                "rollout_used_qty": installed,
                "returned_qty": returned,
                "rollout_actual_qty": installed,
                "remaining_after_rollout": difference,
                "usage_percent": (installed / net_issued * 100) if net_issued > 0 else 0,
                "usage_match": usage_match,
            }
        )
    rows.sort(key=lambda row: (row["area"], row["material"], row["sku"]))
    return {"success": True, "usage": rows, "rollout_records": len(rollout_rows), "rollout_source": rollout_source}


@app.get("/api/warehouse/rollout-material-usage/details")
def list_rollout_material_usage_details(
    request: Request,
    area: str = "",
    material: str = "",
    db: Session = Depends(db_session),
    program: str = DEFAULT_PROGRAM,
):
    require_roles(request, "Admin", "Management", "Requester", "Approval", "Warehouse Manager")
    if is_single_ran(program):
        return {"success": True, "records": [], "total": 0, "source": "disabled"}

    area_key = canonical_area_name(area)
    material_key = canonical_material_key(material)
    if not area_key or not material_key:
        raise HTTPException(status_code=400, detail="Area and material are required")

    rollout_rows, source = rollout_daily_progress_records(db, force=False)
    records = []
    for record in rollout_records_for_session(request, rollout_rows):
        record_area = canonical_area_name(str(record.get("Area") or record.get("area") or "").strip())
        record_material = str(record.get("material type") or record.get("item") or "").strip()
        status = normalize_usage_key(str(record.get("status") or record.get("staus") or ""))
        actual = safe_float(record.get("actual"))
        if (
            record_area != area_key
            or canonical_material_key(record_material) != material_key
            or (status and status not in {"done", "completed", "installed"})
            or actual <= 0
        ):
            continue
        notes = str(record.get("Notes") or record.get("notes") or "").strip()
        hub_match = re.search(r"\bhub\s*:\s*([^|,;]+)", notes, re.IGNORECASE)
        records.append(
            {
                "id": str(record.get("ID") or ""),
                "date": str(record.get("Date") or ""),
                "area": str(record.get("Area") or record.get("area") or ""),
                "xbox": str(record.get("Related to XBOX") or record.get("related_to_xbox") or ""),
                "hub": hub_match.group(1).strip() if hub_match else "",
                "material": record_material,
                "actual": actual,
                "notes": notes,
            }
        )

    records.sort(key=lambda row: (row["date"], row["xbox"], row["hub"], row["id"]))
    return {"success": True, "source": source, "total": sum(row["actual"] for row in records), "records": records}


@app.get("/api/warehouse/movements")
def list_stock_movements(request: Request, limit: int = 50, program: str = DEFAULT_PROGRAM, db: Session = Depends(db_session)):
    program_key = normalize_program(program)
    query = (
        db.query(StockMovement)
        .options(
            joinedload(StockMovement.warehouse),
            joinedload(StockMovement.technician),
            joinedload(StockMovement.product),
        )
        .filter(StockMovement.program == program_key)
        .order_by(StockMovement.id.desc())
    )
    allowed = allowed_warehouse_ids(request, db, program_key)
    if allowed is not None:
        query = query.filter(StockMovement.warehouse_id.in_(allowed))
    rows = query.limit(min(limit, 200)).all()
    return {"success": True, "movements": [movement_to_dict(r) for r in rows]}


@app.get("/api/warehouse/audit-logs")
def list_audit_logs(request: Request, limit: int = 50, program: str = DEFAULT_PROGRAM, db: Session = Depends(db_session)):
    require_roles(request, "Admin")
    program_key = normalize_program(program)
    program_token = f'"program": "{program_key}"'
    if program_key == DEFAULT_PROGRAM:
        audit_filter = or_(AuditLog.details.contains(program_token), ~AuditLog.details.contains('"program"'))
    else:
        audit_filter = AuditLog.details.contains(program_token)
    rows = (
        db.query(AuditLog)
        .filter(audit_filter)
        .order_by(AuditLog.id.desc())
        .limit(min(limit, 200))
        .all()
    )
    return {
        "success": True,
        "logs": [
            {
                "id": row.id,
                "action": row.action,
                "entity_type": row.entity_type,
                "entity_id": row.entity_id,
                "actor": row.actor,
                "details": row.details,
                "created_at": row.created_at.isoformat() if row.created_at else "",
            }
            for row in rows
        ],
    }


@app.get("/api/warehouse/material-requisitions")
def list_material_requisitions(request: Request, limit: int = 50, viewer: str = "", role: str = "", program: str = DEFAULT_PROGRAM, db: Session = Depends(db_session)):
    program_key = normalize_program(program)
    viewer, role = request_scope_viewer(request), current_user(request).role
    rows = (
        db.query(MaterialRequisition)
        .options(joinedload(MaterialRequisition.warehouse), selectinload(MaterialRequisition.items))
        .filter(MaterialRequisition.program == program_key)
        .order_by(MaterialRequisition.id.desc())
        .all()
    )
    visible_rows = [r for r in rows if user_can_view_requisition(r, viewer, role)]
    return {"success": True, "requisitions": [requisition_to_dict(r) for r in visible_rows[: min(limit, 200)]]}


def list_material_requisition_headers(limit: int = 50, db: Session = Depends(db_session), viewer: str = "", role: str = "", program: str = DEFAULT_PROGRAM):
    program_key = normalize_program(program)
    rows = (
        db.query(MaterialRequisition)
        .options(joinedload(MaterialRequisition.warehouse))
        .filter(MaterialRequisition.program == program_key)
        .order_by(MaterialRequisition.id.desc())
        .all()
    )
    visible_rows = [r for r in rows if user_can_view_requisition(r, viewer, role)]
    return {"success": True, "requisitions": [requisition_header_to_dict(r) for r in visible_rows[: min(limit, 200)]]}


@app.get("/api/warehouse/material-requisitions/{requisition_id}")
def get_material_requisition(requisition_id: int, request: Request, viewer: str = "", role: str = "", program: str = DEFAULT_PROGRAM, db: Session = Depends(db_session)):
    program_key = normalize_program(program)
    viewer, role = request_scope_viewer(request), current_user(request).role
    row = (
        db.query(MaterialRequisition)
        .options(
            joinedload(MaterialRequisition.warehouse),
            selectinload(MaterialRequisition.items).joinedload(MaterialRequisitionItem.product),
        )
        .filter(MaterialRequisition.id == requisition_id, MaterialRequisition.program == program_key)
        .first()
    )
    if row is None:
        raise HTTPException(status_code=404, detail="Material requisition not found")
    if not user_can_view_requisition(row, viewer, role):
        raise HTTPException(status_code=403, detail="Not allowed to view this material requisition")
    return {"success": True, "requisition": requisition_to_dict(row)}


@app.get("/api/warehouse/material-requisition-history")
def list_material_requisition_history(
    warehouse: str = "",
    area: str = "",
    technician: str = "",
    requester: str = "",
    status: str = "",
    date_from: str = "",
    date_to: str = "",
    viewer: str = "",
    role: str = "",
    program: str = DEFAULT_PROGRAM,
    db: Session = Depends(db_session),
):
    payload = requisition_history_payload(
        db,
        warehouse=warehouse,
        area=area,
        technician=technician,
        requester=requester,
        status=status,
        date_from=date_from,
        date_to=date_to,
        viewer=viewer,
        role=role,
        program=program,
    )
    return {"success": True, **payload}


@app.get("/api/warehouse/material-requisition-history/export")
def export_material_requisition_history(
    warehouse: str = "",
    area: str = "",
    technician: str = "",
    requester: str = "",
    status: str = "",
    date_from: str = "",
    date_to: str = "",
    viewer: str = "",
    role: str = "",
    program: str = DEFAULT_PROGRAM,
    db: Session = Depends(db_session),
):
    payload = requisition_history_payload(
        db,
        warehouse=warehouse,
        area=area,
        technician=technician,
        requester=requester,
        status=status,
        date_from=date_from,
        date_to=date_to,
        viewer=viewer,
        role=role,
        program=program,
    )
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "MR History"
    sheet.append(["Order", "Date", "Warehouse", "Area", "Site Address", "Requester", "Technician", "Approver", "Status", "Items", "Total Qty", "Materials"])
    for row in payload["rows"]:
        sheet.append(
            [
                row["order_number"],
                row["creation_date"],
                row["warehouse"],
                row["site_id"],
                row["site_address"],
                row["requester_name"],
                row["team_leader"],
                row["receiver_name"],
                row["status"],
                row["item_count"],
                row["total_quantity"],
                row["materials_text"],
            ]
        )
    for col in ("A", "B", "C", "D", "E", "F", "G", "H", "I", "J", "K", "L"):
        sheet.column_dimensions[col].width = 18 if col != "L" else 56
    stream = io.BytesIO()
    workbook.save(stream)
    stream.seek(0)
    filename = f"mr-history-{datetime.now(TRIPOLI_TZ).strftime('%Y%m%d-%H%M%S')}.xlsx"
    headers = {"Content-Disposition": f'attachment; filename="{filename}"'}
    return StreamingResponse(
        stream,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers=headers,
    )


def excel_text(value) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d")
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        return str(int(value)) if float(value).is_integer() else str(value)
    return str(value).strip()


def excel_key(value) -> str:
    return re.sub(r"[^a-z0-9]+", "", excel_text(value).lower())


def excel_qty(value) -> float:
    text = excel_text(value).replace(",", "")
    if not text:
        return 0
    try:
        return float(text)
    except ValueError:
        return 0


def sheet_grid(sheet) -> list[list[str]]:
    return [[excel_text(cell.value) for cell in row] for row in sheet.iter_rows()]


def nearby_sheet_value(grid: list[list[str]], row: int, col: int) -> str:
    label = excel_key(grid[row][col])
    offsets = [1, -1, 2, -2, 3, -3, 4, -4, 0]
    for offset in offsets:
        r = row + (1 if offset == 0 else 0)
        c = col if offset == 0 else col + offset
        if 0 <= r < len(grid) and 0 <= c < len(grid[r]):
            value = grid[r][c].strip()
            key = excel_key(value)
            if value and key != label and not key.endswith("name") and key not in {"date", "siteid", "warehouse", "warehousename"}:
                return value
    return ""


def find_sheet_value(grid: list[list[str]], *labels: str) -> str:
    wanted = {excel_key(label) for label in labels}
    for r, row in enumerate(grid):
        for c, value in enumerate(row):
            key = excel_key(value)
            if key in wanted:
                return nearby_sheet_value(grid, r, c)
    return ""


def product_lookup_maps(db: Session, program: str = DEFAULT_PROGRAM) -> tuple[dict[str, Product], dict[str, Product]]:
    program_key = normalize_program(program)
    exact: dict[str, Product] = {}
    material_keys: dict[str, Product] = {}
    for product in db.query(Product).filter(Product.program == program_key, Product.status == "active").all():
        for value in (product.sku, product.part_number, product_display_name(product), product.name):
            key = excel_key(value)
            if key:
                exact[key] = product
        material_key = canonical_material_key(product_display_name(product))
        if material_key:
            material_keys[material_key] = product
    return exact, material_keys


def find_product_for_import(exact: dict[str, Product], material_keys: dict[str, Product], *values: str) -> Product | None:
    for value in values:
        key = excel_key(value)
        if key in exact:
            return exact[key]
    for value in values:
        key = canonical_material_key(value)
        if key in material_keys:
            return material_keys[key]
    return None


def find_import_header(row: list[str]) -> dict[str, int]:
    mapping: dict[str, int] = {}
    for index, value in enumerate(row):
        key = excel_key(value)
        if key in {"partnbr", "partnumber", "partno", "part"}:
            mapping["part"] = index
        elif key in {"model", "sku", "itemname"}:
            mapping["model"] = index
        elif key in {"description", "descrition", "itemdescription", "itemdescrition", "itemdetail", "material", "materials"}:
            mapping["description"] = index
        elif key in {"uom", "unit"}:
            mapping["uom"] = index
        elif key in {"qty", "quantity"}:
            mapping["quantity"] = index
        elif key in {"remark", "remarks", "comment", "comments"}:
            mapping["remark"] = index
    return mapping


def parse_import_items(db: Session, grid: list[list[str]], program: str = DEFAULT_PROGRAM) -> list[MaterialRequisitionItemIn]:
    header_row = -1
    header: dict[str, int] = {}
    for index, row in enumerate(grid):
        candidate = find_import_header(row)
        if "description" in candidate and "quantity" in candidate and ("part" in candidate or "model" in candidate):
            header_row = index
            header = candidate
            break
    if header_row < 0:
        raise ValueError("Items table was not found")

    exact, material_keys = product_lookup_maps(db, program)
    items: list[MaterialRequisitionItemIn] = []
    empty_count = 0
    for row in grid[header_row + 1 :]:
        part = row[header["part"]].strip() if "part" in header and header["part"] < len(row) else ""
        model = row[header["model"]].strip() if "model" in header and header["model"] < len(row) else ""
        description = row[header["description"]].strip() if header["description"] < len(row) else ""
        qty = excel_qty(row[header["quantity"]] if header["quantity"] < len(row) else "")
        uom = row[header["uom"]].strip() if "uom" in header and header["uom"] < len(row) else "PCS"
        remark = row[header["remark"]].strip() if "remark" in header and header["remark"] < len(row) else ""
        if not part and not model and not description and qty <= 0:
            empty_count += 1
            if empty_count >= 5:
                break
            continue
        empty_count = 0
        if qty <= 0:
            continue
        product = find_product_for_import(exact, material_keys, model, part, description)
        if product is None:
            raise ValueError(f"Material not found: {model or part or description}")
        items.append(
            MaterialRequisitionItemIn(
                product_id=product.id,
                part_nbr=product_part_number(product.sku, product.part_number),
                model=product.sku,
                description=product_display_name(product),
                uom=uom or product.unit or "PCS",
                quantity=qty,
                remark=remark,
            )
        )
    if not items:
        raise ValueError("No valid item rows found")
    return items


def find_warehouse_for_import(db: Session, grid: list[list[str]], default_warehouse_id: int = 0, program: str = DEFAULT_PROGRAM) -> Warehouse:
    program_key = normalize_program(program)
    name = find_sheet_value(grid, "Warehouse name", "Warehouse")
    row = None
    if name:
        row = db.query(Warehouse).filter(Warehouse.program == program_key, func.lower(Warehouse.name) == name.lower()).first()
        if row is None:
            key = normalize_usage_key(name)
            row = next((w for w in db.query(Warehouse).filter(Warehouse.program == program_key).all() if normalize_usage_key(w.name) == key), None)
    if row is None and default_warehouse_id:
        row = db.query(Warehouse).filter(Warehouse.id == default_warehouse_id, Warehouse.program == program_key).first()
    if row is None:
        raise ValueError("Warehouse was not found in the sheet")
    return row


MR_ORDER_PATTERN = re.compile(r"^MR-([A-Z0-9]+)-(\d+)$")


def material_requisition_warehouse_code(warehouse: Warehouse) -> str:
    """Return the stable short code used in MR numbers for a warehouse."""
    key = normalize_usage_key(warehouse.name or "")
    if "freezone" in key:
        return "FZ"
    if "tripoli" in key:
        return "TRI"
    if any(token in key for token in ("misurata", "misrata", "misrat")):
        return "MIS"

    words = re.findall(r"[A-Za-z0-9]+", warehouse.name or "")
    initials = "".join(word[0] for word in words).upper()
    return initials[:4] or f"WH{warehouse.id}"


def is_material_requisition_number_for_warehouse(order_number: str, warehouse: Warehouse) -> bool:
    match = MR_ORDER_PATTERN.fullmatch(str(order_number or "").strip().upper())
    return bool(match and match.group(1) == material_requisition_warehouse_code(warehouse))


def next_material_requisition_number(db: Session, warehouse: Warehouse, program: str = DEFAULT_PROGRAM) -> str:
    """Return the next MR number within the selected warehouse's own sequence."""
    program_key = normalize_program(program)
    warehouse_code = material_requisition_warehouse_code(warehouse)
    highest = 0
    values = (
        db.query(MaterialRequisition.order_number)
        .filter(
            MaterialRequisition.program == program_key,
            MaterialRequisition.warehouse_id == warehouse.id,
        )
        .all()
    )
    for (value,) in values:
        match = MR_ORDER_PATTERN.fullmatch(str(value or "").strip().upper())
        if match and match.group(1) == warehouse_code:
            highest = max(highest, int(match.group(2)))

    candidate = highest + 1
    while (
        db.query(MaterialRequisition.id)
        .filter(
            MaterialRequisition.program == program_key,
            MaterialRequisition.order_number == f"MR-{warehouse_code}-{candidate:04d}",
        )
        .first()
    ):
        candidate += 1
    return f"MR-{warehouse_code}-{candidate:04d}"


def migrate_legacy_material_requisition_numbers() -> None:
    """Give old numeric MRs a warehouse-specific number without changing their data."""
    for program_key, session_factory in all_sessionmakers():
        try:
            with session_factory() as db:
                rows = (
                    db.query(MaterialRequisition)
                    .options(joinedload(MaterialRequisition.warehouse))
                    .filter(MaterialRequisition.program == normalize_program(program_key))
                    .all()
                )
                legacy_rows = [row for row in rows if not is_material_requisition_number_for_warehouse(row.order_number, row.warehouse)]
                if not legacy_rows:
                    continue

                next_sequence: dict[int, int] = {}
                for row in rows:
                    if not is_material_requisition_number_for_warehouse(row.order_number, row.warehouse):
                        continue
                    match = MR_ORDER_PATTERN.fullmatch(row.order_number.strip().upper())
                    next_sequence[row.warehouse_id] = max(next_sequence.get(row.warehouse_id, 0), int(match.group(2)))

                legacy_rows.sort(key=lambda row: (str(row.creation_date or ""), str(row.created_at or ""), row.id))
                changes: list[tuple[MaterialRequisition, str, str]] = []
                for row in legacy_rows:
                    next_sequence[row.warehouse_id] = next_sequence.get(row.warehouse_id, 0) + 1
                    new_order_number = f"MR-{material_requisition_warehouse_code(row.warehouse)}-{next_sequence[row.warehouse_id]:04d}"
                    changes.append((row, str(row.order_number or ""), new_order_number))

                # Temporary values avoid unique-index conflicts if a legacy value
                # happens to look like a new MR number in another warehouse.
                for row, _old_order_number, _new_order_number in changes:
                    row.order_number = f"__mr_renumber_{row.id}__"
                db.flush()

                for row, old_order_number, new_order_number in changes:
                    row.order_number = new_order_number
                    db.query(StockMovement).filter(
                        StockMovement.program == normalize_program(program_key),
                        StockMovement.reference == old_order_number,
                    ).update({StockMovement.reference: new_order_number}, synchronize_session=False)
                    db.query(MaterialScanLog).filter(
                        MaterialScanLog.material_requisition_id == row.id,
                        MaterialScanLog.note == f"MR {old_order_number}",
                    ).update({MaterialScanLog.note: f"MR {new_order_number}"}, synchronize_session=False)
                    db.query(AuditLog).filter(
                        AuditLog.entity_type == "material_requisition",
                        AuditLog.entity_id == old_order_number,
                    ).update({AuditLog.entity_id: new_order_number}, synchronize_session=False)
                    db.add(
                        AuditLog(
                            action="renumber_material_requisition",
                            entity_type="material_requisition",
                            entity_id=new_order_number,
                            actor="system",
                            details=json.dumps(
                                {
                                    "program": normalize_program(program_key),
                                    "warehouse": row.warehouse.name if row.warehouse else "",
                                    "previous_order_number": old_order_number,
                                }
                            ),
                        )
                    )
                db.commit()
                logger.info("%s MR number migration completed for %s records", program_key, len(changes))
        except Exception:
            logger.exception("%s MR number migration failed", program_key)


migrate_legacy_material_requisition_numbers()


def add_previous_mr_numbers_to_notes() -> None:
    """Show the legacy MR number in the request note after a renumbering."""
    for program_key, session_factory in all_sessionmakers():
        try:
            with session_factory() as db:
                audit_rows = (
                    db.query(AuditLog)
                    .filter(
                        AuditLog.action == "renumber_material_requisition",
                        AuditLog.entity_type == "material_requisition",
                    )
                    .all()
                )
                changed = 0
                for audit in audit_rows:
                    try:
                        details = json.loads(audit.details or "{}")
                    except (TypeError, ValueError):
                        continue
                    previous_number = str(details.get("previous_order_number") or "").strip()
                    if not previous_number:
                        continue
                    row = (
                        db.query(MaterialRequisition)
                        .filter(
                            MaterialRequisition.program == normalize_program(program_key),
                            MaterialRequisition.order_number == audit.entity_id,
                        )
                        .first()
                    )
                    if row is None:
                        continue
                    note = f"Previous MR No: {previous_number}"
                    current_comment = (row.requester_comment or "").strip()
                    if note.lower() in current_comment.lower():
                        continue
                    row.requester_comment = f"{current_comment}\n{note}".strip()
                    changed += 1
                if changed:
                    db.commit()
                    logger.info("%s added previous MR numbers to %s request notes", program_key, changed)
        except Exception:
            logger.exception("%s previous MR note migration failed", program_key)


add_previous_mr_numbers_to_notes()


def import_mr_sheet(db: Session, sheet, filename: str, default_warehouse_id: int, actor: str, program: str = DEFAULT_PROGRAM) -> MaterialRequisition:
    program_key = normalize_program(program)
    grid = sheet_grid(sheet)
    warehouse = find_warehouse_for_import(db, grid, default_warehouse_id, program_key)
    items = parse_import_items(db, grid, program_key)
    today = local_today()
    order_number = next_material_requisition_number(db, warehouse, program_key)
    row = MaterialRequisition(
        program=program_key,
        order_number=order_number,
        creation_date=find_sheet_value(grid, "Creation Date", "Date") or today,
        warehouse_id=warehouse.id,
        entity=find_sheet_value(grid, "Entity") or "Rollout",
        project_name=find_sheet_value(grid, "Project Name") or PROGRAM_LABELS[program_key],
        site_id=find_sheet_value(grid, "Site ID", "Area Name") or sheet.title,
        site_address=find_sheet_value(grid, "Site Address", "Site Adress") or "",
        wo_no=find_sheet_value(grid, "WO No") or "",
        product_domain=find_sheet_value(grid, "Product Domain") or "Passive",
        team_leader=find_sheet_value(grid, "Team Leader", "Receiver/TEL") or "",
        receiver_tel=find_sheet_value(grid, "Receiver/TEL") or "",
        request_shipment_time=find_sheet_value(grid, "Request Shipment Time") or today,
        request_arrived_site_time=find_sheet_value(grid, "Request arrived site time") or today,
        requester_name=find_sheet_value(grid, "Requester Name", "Requester") or actor or "Import",
        requester_title=find_sheet_value(grid, "Requester Title") or "Requester",
        requester_date=today,
        requester_comment=f"Imported from {filename} / {sheet.title}",
        receiver_name=find_sheet_value(grid, "Approver Name", "Receiver Name", "Receiver") or "Imported Approval",
        receiver_title=find_sheet_value(grid, "Approver Title", "Receiver Title") or "Approval",
        receiver_date=today,
        status="approved",
        created_by=actor or "Import",
    )
    db.add(row)
    db.flush()
    for index, item in enumerate(items, start=1):
        product = db.get(Product, item.product_id) if item.product_id else None
        db.add(
            MaterialRequisitionItem(
                requisition_id=row.id,
                line_no=index,
                product_id=item.product_id,
                part_nbr=item.part_nbr,
                model=item.model,
                description=item.description,
                vendor=str(product.vendor or "").strip() if program_key == SINGLE_RAN_PROGRAM and product else "",
                uom=item.uom,
                quantity=item.quantity,
                remark=item.remark,
            )
        )
    db.flush()
    issue_material_requisition_row(db, row, actor or "Import")
    log_audit(db, "import_material_requisition_excel", "material_requisition", row.order_number, actor or "Import", {"file": filename, "sheet": sheet.title})
    return row


@app.post("/api/warehouse/material-requisitions/import-excel")
async def import_material_requisition_excels(
    request: Request,
    files: list[UploadFile] = File(...),
    default_warehouse_id: int = Form(0),
    actor: str = Form("Import"),
    program: str = Form(DEFAULT_PROGRAM),
    db: Session = Depends(db_session),
):
    require_roles(request, "Admin", "Management", "Requester", "Approval", "Warehouse Manager")
    require_roles(request, "Admin", "Management", "Warehouse Manager")
    program_key = normalize_program(program)
    actor = request_actor(request)
    if default_warehouse_id:
        require_warehouse_access(request, db, default_warehouse_id, program_key)
    results = []
    imported_rows: list[MaterialRequisition] = []
    for upload in files:
        try:
            content = await upload.read()
            workbook = load_workbook(io.BytesIO(content), data_only=True)
        except Exception as exc:
            results.append({"file": upload.filename, "sheet": "", "success": False, "message": "Could not read Excel file"})
            continue
        for sheet in workbook.worksheets:
            if sheet.sheet_state != "visible":
                continue
            try:
                row = import_mr_sheet(db, sheet, upload.filename or "MR.xlsx", default_warehouse_id, actor, program_key)
                db.commit()
                db.refresh(row)
                imported_rows.append(row)
                results.append({"file": upload.filename, "sheet": sheet.title, "success": True, "order": row.order_number})
            except Exception as exc:
                db.rollback()
                detail = exc.detail if isinstance(exc, HTTPException) else str(exc)
                results.append({"file": upload.filename, "sheet": sheet.title, "success": False, "message": detail})
    clear_warehouse_cache()
    return {
        "success": True,
        "imported": len(imported_rows),
        "failed": len([r for r in results if not r["success"]]),
        "results": results,
        "requisitions": [requisition_to_dict(r) for r in imported_rows],
    }


def list_material_transfer_headers(limit: int = 50, db: Session = Depends(db_session), viewer: str = "", role: str = "", program: str = DEFAULT_PROGRAM):
    program_key = normalize_program(program)
    rows = (
        db.query(MaterialTransfer)
        .options(joinedload(MaterialTransfer.from_warehouse), joinedload(MaterialTransfer.to_warehouse))
        .filter(MaterialTransfer.program == program_key)
        .order_by(MaterialTransfer.id.desc())
        .all()
    )
    visible_rows = [r for r in rows if user_can_view_transfer(r, viewer, role)]
    return {"success": True, "transfers": [transfer_to_dict(r, include_items=False) for r in visible_rows[: min(limit, 200)]]}


@app.get("/api/warehouse/material-transfers")
def list_material_transfers(request: Request, limit: int = 50, viewer: str = "", role: str = "", program: str = DEFAULT_PROGRAM, db: Session = Depends(db_session)):
    program_key = normalize_program(program)
    viewer, role = request_scope_viewer(request), current_user(request).role
    rows = (
        db.query(MaterialTransfer)
        .options(
            joinedload(MaterialTransfer.from_warehouse),
            joinedload(MaterialTransfer.to_warehouse),
            selectinload(MaterialTransfer.items).joinedload(MaterialTransferItem.product),
        )
        .filter(MaterialTransfer.program == program_key)
        .order_by(MaterialTransfer.id.desc())
        .all()
    )
    visible_rows = [r for r in rows if user_can_view_transfer(r, viewer, role)]
    return {"success": True, "transfers": [transfer_to_dict(r) for r in visible_rows[: min(limit, 200)]]}


@app.get("/api/warehouse/material-transfers/{transfer_id}")
def get_material_transfer(transfer_id: int, request: Request, viewer: str = "", role: str = "", program: str = DEFAULT_PROGRAM, db: Session = Depends(db_session)):
    program_key = normalize_program(program)
    viewer, role = request_scope_viewer(request), current_user(request).role
    row = db.query(MaterialTransfer).filter(MaterialTransfer.id == transfer_id, MaterialTransfer.program == program_key).first()
    if row is None:
        raise HTTPException(status_code=404, detail="Material transfer not found")
    if not user_can_view_transfer(row, viewer, role):
        raise HTTPException(status_code=403, detail="Not allowed to view this material transfer")
    return {"success": True, "transfer": transfer_to_dict(row)}


def list_material_return_headers(limit: int = 50, program: str = DEFAULT_PROGRAM, db: Session = Depends(db_session)):
    program_key = normalize_program(program)
    rows = (
        db.query(MaterialReturn)
        .options(joinedload(MaterialReturn.warehouse))
        .filter(MaterialReturn.program == program_key)
        .order_by(MaterialReturn.id.desc())
        .limit(min(limit, 200))
        .all()
    )
    return {"success": True, "returns": [material_return_to_dict(r, include_items=False) for r in rows]}


@app.get("/api/warehouse/material-returns")
def list_material_returns(request: Request, limit: int = 50, program: str = DEFAULT_PROGRAM, db: Session = Depends(db_session)):
    program_key = normalize_program(program)
    viewer, role = request_scope_viewer(request), current_user(request).role
    query = (
        db.query(MaterialReturn)
        .options(
            joinedload(MaterialReturn.warehouse),
            selectinload(MaterialReturn.items).joinedload(MaterialReturnItem.product),
        )
        .filter(MaterialReturn.program == program_key)
        .order_by(MaterialReturn.id.desc())
    )
    allowed = allowed_warehouse_ids(request, db, program_key)
    if allowed is not None:
        query = query.filter(MaterialReturn.warehouse_id.in_(allowed))
    rows = query.limit(min(limit, 200)).all()
    return {"success": True, "returns": [material_return_to_dict(r) for r in rows if user_can_view_material_return(r, viewer, role)]}


@app.get("/api/warehouse/material-returns/{return_id}")
def get_material_return(return_id: int, request: Request, program: str = DEFAULT_PROGRAM, db: Session = Depends(db_session)):
    program_key = normalize_program(program)
    viewer, role = request_scope_viewer(request), current_user(request).role
    row = db.query(MaterialReturn).options(joinedload(MaterialReturn.warehouse)).filter(MaterialReturn.id == return_id, MaterialReturn.program == program_key).first()
    if row is None:
        raise HTTPException(status_code=404, detail="Material return not found")
    if not user_can_view_material_return(row, viewer, role):
        raise HTTPException(status_code=403, detail="Not allowed to view this material return")
    return {"success": True, "return": material_return_to_dict(row)}


@app.get("/api/warehouse/notifications")
def warehouse_notifications(request: Request, user: str = "", program: str = DEFAULT_PROGRAM, db: Session = Depends(db_session)):
    program_key = normalize_program(program)
    user = request_actor(request)
    pending = db.query(MaterialRequisition).filter(MaterialRequisition.program == program_key, MaterialRequisition.status == "pending_approval").all()
    pending_transfers = (
        db.query(MaterialTransfer)
        .options(joinedload(MaterialTransfer.from_warehouse), joinedload(MaterialTransfer.to_warehouse))
        .filter(MaterialTransfer.program == program_key, MaterialTransfer.status == "pending_approval")
        .all()
    )
    pending_returns = (
        db.query(MaterialReturn)
        .options(joinedload(MaterialReturn.warehouse))
        .filter(MaterialReturn.program == program_key, MaterialReturn.status == "pending_warehouse")
        .all()
    )
    user_key = normalize_usage_key(user)
    approval_count = len(pending) if not user_key else sum(1 for row in pending if normalize_usage_key(row.receiver_name) == user_key)
    transfer_approval_count = len(pending_transfers) if not user_key else sum(
        1 for row in pending_transfers if normalize_usage_key(row.approver_name) == user_key
    )
    approved_rows = (
        db.query(MaterialRequisition)
        .options(joinedload(MaterialRequisition.warehouse))
        .filter(MaterialRequisition.program == program_key, MaterialRequisition.status == "approved")
        .all()
    )
    approved_transfer_rows = (
        db.query(MaterialTransfer)
        .options(joinedload(MaterialTransfer.from_warehouse), joinedload(MaterialTransfer.to_warehouse))
        .filter(MaterialTransfer.program == program_key, MaterialTransfer.status == "approved")
        .all()
    )
    approved_count = len(approved_rows) if not user_key else sum(
        1 for row in approved_rows if warehouse_manager_handles_mr(user, row)
    )
    approved_transfer_count = len(approved_transfer_rows) if not user_key else sum(
        1
        for row in approved_transfer_rows
        if warehouse_scope_matches(user, row.from_warehouse.name if row.from_warehouse else "")
        or warehouse_scope_matches(user, row.to_warehouse.name if row.to_warehouse else "")
    )
    pending_return_count = len(pending_returns) if not user_key else sum(
        1 for row in pending_returns if warehouse_scope_matches(user, row.warehouse.name if row.warehouse else "")
    )
    return {
        "success": True,
        "approval_count": approval_count + transfer_approval_count,
        "warehouse_queue_count": approved_count + approved_transfer_count + pending_return_count,
    }


@app.get("/api/warehouse/receive-orders")
def list_receive_orders(request: Request, limit: int = 50, program: str = DEFAULT_PROGRAM, db: Session = Depends(db_session)):
    program_key = normalize_program(program)
    query = (
        db.query(ReceiveOrder)
        .options(joinedload(ReceiveOrder.warehouse), selectinload(ReceiveOrder.items).joinedload(ReceiveOrderItem.product))
        .filter(ReceiveOrder.program == program_key)
        .order_by(ReceiveOrder.id.desc())
    )
    allowed = allowed_warehouse_ids(request, db, program_key)
    if allowed is not None:
        query = query.filter(ReceiveOrder.warehouse_id.in_(allowed))
    rows = query.limit(min(limit, 200)).all()
    return {"success": True, "receipts": [receive_order_to_dict(r) for r in rows]}


def list_receive_order_headers(limit: int = 50, program: str = DEFAULT_PROGRAM, db: Session = Depends(db_session)):
    program_key = normalize_program(program)
    rows = (
        db.query(ReceiveOrder)
        .options(joinedload(ReceiveOrder.warehouse))
        .filter(ReceiveOrder.program == program_key)
        .order_by(ReceiveOrder.id.desc())
        .limit(min(limit, 200))
        .all()
    )
    return {"success": True, "receipts": [receive_order_header_to_dict(r) for r in rows]}


@app.post("/api/warehouse/material-requisitions")
def create_material_requisition(data: MaterialRequisitionIn, request: Request, db: Session = Depends(db_session)):
    user = require_roles(request, "Admin", "Management", "Requester")
    program_key = normalize_program(data.program)
    data.site_id = resolved_site_name(db, program_key, data.site_id)
    warehouse = require_warehouse(db, data.warehouse_id, program_key)
    data.created_by = request_actor(request)
    data.requester_name = data.created_by
    if not data.items:
        raise HTTPException(status_code=400, detail="At least one item is required")
    validate_reservable_stock(db, data.warehouse_id, data.items, program_key)

    order_number = next_material_requisition_number(db, warehouse, program_key)
    row = MaterialRequisition(
        program=program_key,
        order_number=order_number,
        creation_date=data.creation_date,
        warehouse_id=data.warehouse_id,
        entity=data.entity,
        project_name=data.project_name,
        site_id=data.site_id,
        site_address=data.site_address,
        wo_no=data.wo_no,
        product_domain=data.product_domain,
        team_leader=data.team_leader,
        receiver_tel=data.receiver_tel,
        request_shipment_time=data.request_shipment_time,
        request_arrived_site_time=data.request_arrived_site_time,
        requester_name=data.requester_name,
        requester_title=data.requester_title,
        requester_signature=data.requester_signature,
        requester_date=data.requester_date,
        requester_comment=data.requester_comment,
        receiver_name=data.receiver_name.strip() or "Mustafa",
        receiver_title=data.receiver_title,
        receiver_signature=data.receiver_signature,
        receiver_date=data.receiver_date,
        receiver_comment=data.receiver_comment,
        return_reason=data.return_reason,
        status="draft",
        created_by=data.created_by,
    )
    db.add(row)
    db.flush()

    for index, item in enumerate(data.items, start=1):
        product = require_product(db, item.product_id, program_key) if item.product_id else None
        item_vendor = item.vendor.strip()
        if program_key == SINGLE_RAN_PROGRAM:
            if product is None:
                raise HTTPException(status_code=400, detail="Choose a Single RAN material from the search results")
            product_vendor = str(product.vendor or "").strip()
            if not item_vendor:
                raise HTTPException(status_code=400, detail=f"Vendor is required for {product_display_name(product)}")
            if product_vendor and item_vendor.casefold() != product_vendor.casefold():
                raise HTTPException(status_code=400, detail=f"Vendor does not match {product_display_name(product)}")
            item_vendor = product_vendor or item_vendor
        db.add(
            MaterialRequisitionItem(
                requisition_id=row.id,
                line_no=index,
                product_id=item.product_id,
                part_nbr=item.part_nbr or product_part_number(product.sku if product else ""),
                model=item.model or (product.sku if product else ""),
                description=material_display_name(item.description or product_display_name(product), product.sku if product else ""),
                vendor=item_vendor if program_key == SINGLE_RAN_PROGRAM else "",
                uom=item.uom or (product.unit if product else "PCS"),
                quantity=item.quantity,
                remark=item.remark,
            )
        )

    db.flush()
    issue_order = None
    if data.issue_immediately and user.role.strip().lower() == "admin":
        row.status = "approved"
        issue_order = issue_material_requisition_row(db, row)
    else:
        row.status = "pending_approval"

    log_audit(db, "create_material_requisition", "material_requisition", row.order_number, data.created_by, data.model_dump())
    db.commit()
    db.refresh(row)
    if row.status == "pending_approval":
        try:
            notify_mr_created(row, db)
        except Exception:
            logger.exception("MR notification failed after save: %s", row.order_number)
    return {"success": True, "issue_order": issue_order, "requisition": requisition_to_dict(row)}


@app.post("/api/warehouse/material-requisitions/{requisition_id}/resubmit")
def resubmit_material_requisition(requisition_id: int, data: MaterialRequisitionIn, request: Request, db: Session = Depends(db_session)):
    user = require_roles(request, "Admin", "Management", "Requester")
    program_key = normalize_program(data.program)
    data.site_id = resolved_site_name(db, program_key, data.site_id)
    row = (
        db.query(MaterialRequisition)
        .filter(MaterialRequisition.id == requisition_id, MaterialRequisition.program == program_key)
        .with_for_update()
        .first()
    )
    if row is None:
        raise HTTPException(status_code=404, detail="Material requisition not found")
    if user.role.strip().lower() != "admin" and normalize_usage_key(row.created_by) != normalize_usage_key(request_actor(request)):
        raise HTTPException(status_code=403, detail="You can only resubmit your own requisitions")
    if row.status != "returned_for_edit":
        raise HTTPException(status_code=400, detail=f"MR cannot be edited from status {row.status}")
    require_warehouse(db, data.warehouse_id, program_key)
    if not data.items:
        raise HTTPException(status_code=400, detail="At least one item is required")
    validate_reservable_stock(db, data.warehouse_id, data.items, program_key, exclude_requisition_id=row.id)

    row.creation_date = data.creation_date
    row.warehouse_id = data.warehouse_id
    row.entity = data.entity
    row.project_name = data.project_name
    row.site_id = data.site_id
    row.site_address = data.site_address
    row.wo_no = data.wo_no
    row.product_domain = data.product_domain
    row.team_leader = data.team_leader
    row.receiver_tel = data.receiver_tel
    row.request_shipment_time = data.request_shipment_time
    row.request_arrived_site_time = data.request_arrived_site_time
    row.requester_name = request_actor(request)
    row.requester_title = data.requester_title
    row.requester_signature = data.requester_signature
    row.requester_date = data.requester_date
    row.requester_comment = data.requester_comment
    row.receiver_name = data.receiver_name.strip() or "Mustafa"
    row.receiver_title = data.receiver_title
    row.receiver_signature = data.receiver_signature
    row.receiver_date = data.receiver_date
    row.receiver_comment = data.receiver_comment
    row.return_reason = ""
    row.status = "pending_approval"

    db.query(MaterialRequisitionItem).filter(MaterialRequisitionItem.requisition_id == row.id).delete()
    for index, item in enumerate(data.items, start=1):
        product = require_product(db, item.product_id, program_key) if item.product_id else None
        item_vendor = item.vendor.strip()
        if program_key == SINGLE_RAN_PROGRAM:
            if product is None:
                raise HTTPException(status_code=400, detail="Choose a Single RAN material from the search results")
            product_vendor = str(product.vendor or "").strip()
            if not item_vendor:
                raise HTTPException(status_code=400, detail=f"Vendor is required for {product_display_name(product)}")
            if product_vendor and item_vendor.casefold() != product_vendor.casefold():
                raise HTTPException(status_code=400, detail=f"Vendor does not match {product_display_name(product)}")
            item_vendor = product_vendor or item_vendor
        db.add(
            MaterialRequisitionItem(
                requisition_id=row.id,
                line_no=index,
                product_id=item.product_id,
                part_nbr=item.part_nbr or product_part_number(product.sku if product else ""),
                model=item.model or (product.sku if product else ""),
                description=material_display_name(item.description or product_display_name(product), product.sku if product else ""),
                vendor=item_vendor if program_key == SINGLE_RAN_PROGRAM else "",
                uom=item.uom or (product.unit if product else "PCS"),
                quantity=item.quantity,
                remark=item.remark,
            )
        )

    log_audit(db, "resubmit_material_requisition", "material_requisition", row.order_number, request_actor(request), data.model_dump())
    db.commit()
    db.refresh(row)
    try:
        notify_mr_created(row, db)
    except Exception:
        logger.exception("MR notification failed after resubmit: %s", row.order_number)
    return {"success": True, "requisition": requisition_to_dict(row)}


@app.post("/api/warehouse/material-transfers")
def create_material_transfer(data: MaterialTransferIn, request: Request, db: Session = Depends(db_session)):
    user = require_roles(request, "Admin", "Management", "Requester", "Warehouse Manager")
    program_key = normalize_program(data.program)
    if user.role.strip().lower() == "requester":
        require_warehouse(db, data.from_warehouse_id, program_key)
    else:
        require_warehouse_access(request, db, data.from_warehouse_id, program_key)
    require_warehouse(db, data.to_warehouse_id, program_key)
    if data.from_warehouse_id == data.to_warehouse_id:
        raise HTTPException(status_code=400, detail="From and To warehouse must be different")
    if not data.items:
        raise HTTPException(status_code=400, detail="At least one item is required")
    validate_reservable_stock(db, data.from_warehouse_id, data.items, program_key)

    row = MaterialTransfer(
        program=program_key,
        transfer_number=next_number(db, MaterialTransfer, "TR"),
        transfer_date=data.transfer_date or local_today(),
        from_warehouse_id=data.from_warehouse_id,
        to_warehouse_id=data.to_warehouse_id,
        reference_no=data.reference_no.strip(),
        reason=data.reason.strip(),
        requester_name=request_actor(request),
        requester_title=data.requester_title.strip(),
        approver_name=data.approver_name.strip(),
        approver_title=data.approver_title.strip(),
        receiver_name=data.receiver_name.strip(),
        status="pending_approval",
        created_by=request_actor(request),
    )
    db.add(row)
    db.flush()

    for index, item in enumerate(data.items, start=1):
        product = require_product(db, item.product_id, program_key)
        db.add(
            MaterialTransferItem(
                transfer_id=row.id,
                line_no=index,
                product_id=item.product_id,
                part_nbr=product_part_number(product.sku, product.part_number),
                description=product_display_name(product),
                uom=product.unit,
                quantity=item.quantity,
                remark=item.remark.strip(),
            )
        )

    log_audit(db, "create_material_transfer", "material_transfer", row.transfer_number, row.created_by, data.model_dump())
    db.commit()
    db.refresh(row)
    notify_transfer_created(row, db)
    return {"success": True, "transfer_number": row.transfer_number, "transfer": transfer_to_dict(row)}


@app.post("/api/warehouse/material-transfers/{transfer_id}/return-for-edit")
def return_material_transfer_for_edit(transfer_id: int, data: MaterialRequisitionActionIn, request: Request, db: Session = Depends(db_session)):
    require_roles(request, "Admin", "Management", "Approval")
    program_key = normalize_program(data.program)
    row = (
        db.query(MaterialTransfer)
        .filter(MaterialTransfer.id == transfer_id, MaterialTransfer.program == program_key)
        .with_for_update()
        .first()
    )
    if row is None:
        raise HTTPException(status_code=404, detail="Material transfer not found")
    if row.status != "pending_approval":
        raise HTTPException(status_code=400, detail=f"Transfer cannot be returned for edit from status {row.status}")
    comment = data.comment.strip()
    if not comment:
        raise HTTPException(status_code=400, detail="Return reason is required")
    actor = request_actor(request)
    row.approver_name = actor or row.approver_name
    row.approver_title = data.title or row.approver_title
    row.approver_date = local_today()
    row.approver_comment = comment
    row.status = "returned_for_edit"
    log_audit(db, "return_material_transfer_for_edit", "material_transfer", row.transfer_number, actor or "approval", data.model_dump())
    db.commit()
    db.refresh(row)
    notify_transfer_returned_for_edit(row, db)
    return {"success": True, "transfer": transfer_to_dict(row)}


@app.post("/api/warehouse/material-transfers/{transfer_id}/return-by-destination")
def return_material_transfer_by_destination(transfer_id: int, data: MaterialRequisitionActionIn, request: Request, db: Session = Depends(db_session)):
    require_roles(request, "Admin", "Management", "Warehouse Manager")
    program_key = normalize_program(data.program)
    row = (
        db.query(MaterialTransfer)
        .filter(MaterialTransfer.id == transfer_id, MaterialTransfer.program == program_key)
        .with_for_update()
        .first()
    )
    if row is None:
        raise HTTPException(status_code=404, detail="Material transfer not found")
    if row.status != "approved":
        raise HTTPException(status_code=400, detail=f"Transfer cannot be returned from status {row.status}")
    require_warehouse_access(request, db, row.to_warehouse_id, program_key)
    comment = data.comment.strip()
    if not comment:
        raise HTTPException(status_code=400, detail="Return reason is required")

    actor = request_actor(request)
    row.receiver_name = actor or row.receiver_name
    row.receiver_date = local_today()
    row.receiver_comment = comment
    row.status = "returned_for_edit"
    log_audit(
        db,
        "return_material_transfer_by_destination",
        "material_transfer",
        row.transfer_number,
        actor or "warehouse_manager",
        data.model_dump(),
    )
    db.commit()
    db.refresh(row)
    notify_transfer_returned_by_destination(row, db)
    return {"success": True, "transfer": transfer_to_dict(row)}


@app.post("/api/warehouse/material-transfers/{transfer_id}/resubmit")
def resubmit_material_transfer(transfer_id: int, data: MaterialTransferIn, request: Request, db: Session = Depends(db_session)):
    user = require_roles(request, "Admin", "Management", "Requester", "Warehouse Manager")
    program_key = normalize_program(data.program)
    row = (
        db.query(MaterialTransfer)
        .filter(MaterialTransfer.id == transfer_id, MaterialTransfer.program == program_key)
        .with_for_update()
        .first()
    )
    if row is None:
        raise HTTPException(status_code=404, detail="Material transfer not found")
    role_key = normalize_usage_key(user.role)
    if role_key not in {"admin", "management"} and normalize_usage_key(row.created_by) != normalize_usage_key(request_actor(request)):
        raise HTTPException(status_code=403, detail="You can only resubmit your own material transfers")
    if row.status != "returned_for_edit":
        raise HTTPException(status_code=400, detail=f"Transfer cannot be edited from status {row.status}")
    if role_key == "requester":
        require_warehouse(db, data.from_warehouse_id, program_key)
    else:
        require_warehouse_access(request, db, data.from_warehouse_id, program_key)
    require_warehouse(db, data.to_warehouse_id, program_key)
    if data.from_warehouse_id == data.to_warehouse_id:
        raise HTTPException(status_code=400, detail="From and To warehouse must be different")
    if not data.items:
        raise HTTPException(status_code=400, detail="At least one item is required")
    validate_reservable_stock(db, data.from_warehouse_id, data.items, program_key, exclude_transfer_id=row.id)

    validated_items = []
    for item in data.items:
        product = require_product(db, item.product_id, program_key)
        validated_items.append((item, product))

    row.transfer_date = data.transfer_date or local_today()
    row.from_warehouse_id = data.from_warehouse_id
    row.to_warehouse_id = data.to_warehouse_id
    row.reference_no = data.reference_no.strip()
    row.reason = data.reason.strip()
    row.requester_name = request_actor(request)
    row.requester_title = data.requester_title.strip()
    row.approver_name = data.approver_name.strip()
    row.approver_title = data.approver_title.strip()
    row.approver_date = ""
    row.approver_comment = ""
    row.receiver_name = data.receiver_name.strip()
    row.receiver_date = ""
    row.receiver_comment = ""
    row.status = "pending_approval"

    db.query(MaterialTransferItem).filter(MaterialTransferItem.transfer_id == row.id).delete()
    for index, (item, product) in enumerate(validated_items, start=1):
        db.add(
            MaterialTransferItem(
                transfer_id=row.id,
                line_no=index,
                product_id=item.product_id,
                part_nbr=product_part_number(product.sku, product.part_number),
                description=product_display_name(product),
                uom=product.unit,
                quantity=item.quantity,
                remark=item.remark.strip(),
            )
        )

    log_audit(db, "resubmit_material_transfer", "material_transfer", row.transfer_number, request_actor(request), data.model_dump())
    db.commit()
    db.refresh(row)
    notify_transfer_created(row, db)
    return {"success": True, "transfer": transfer_to_dict(row)}


@app.post("/api/warehouse/material-transfers/{transfer_id}/approve")
def approve_material_transfer(transfer_id: int, data: MaterialRequisitionActionIn, request: Request, db: Session = Depends(db_session)):
    require_roles(request, "Admin", "Management", "Approval")
    program_key = normalize_program(data.program)
    row = (
        db.query(MaterialTransfer)
        .filter(MaterialTransfer.id == transfer_id, MaterialTransfer.program == program_key)
        .with_for_update()
        .first()
    )
    if row is None:
        raise HTTPException(status_code=404, detail="Material transfer not found")
    if row.status != "pending_approval":
        raise HTTPException(status_code=400, detail=f"Transfer cannot be approved from status {row.status}")
    validate_reservable_stock(db, row.from_warehouse_id, row.items, program_key, exclude_transfer_id=row.id)
    actor = request_actor(request)
    row.approver_name = actor or row.approver_name
    row.approver_title = data.title or row.approver_title
    row.approver_date = local_today()
    row.approver_comment = data.comment
    row.status = "approved"
    log_audit(db, "approve_material_transfer", "material_transfer", row.transfer_number, actor or "approval", data.model_dump())
    db.commit()
    db.refresh(row)
    notify_transfer_approved(row, db)
    return {"success": True, "transfer": transfer_to_dict(row)}


@app.post("/api/warehouse/material-transfers/{transfer_id}/reject")
def reject_material_transfer(transfer_id: int, data: MaterialRequisitionActionIn, request: Request, db: Session = Depends(db_session)):
    require_roles(request, "Admin", "Management", "Approval")
    program_key = normalize_program(data.program)
    row = (
        db.query(MaterialTransfer)
        .filter(MaterialTransfer.id == transfer_id, MaterialTransfer.program == program_key)
        .with_for_update()
        .first()
    )
    if row is None:
        raise HTTPException(status_code=404, detail="Material transfer not found")
    if row.status != "pending_approval":
        raise HTTPException(status_code=400, detail=f"Transfer cannot be rejected from status {row.status}")
    actor = request_actor(request)
    row.approver_name = actor or row.approver_name
    row.approver_title = data.title or row.approver_title
    row.approver_date = local_today()
    row.approver_comment = data.comment
    row.status = "rejected"
    log_audit(db, "reject_material_transfer", "material_transfer", row.transfer_number, actor or "approval", data.model_dump())
    db.commit()
    db.refresh(row)
    return {"success": True, "transfer": transfer_to_dict(row)}


@app.post("/api/warehouse/material-transfers/{transfer_id}/confirm")
def confirm_material_transfer(transfer_id: int, request: Request, data: MaterialRequisitionActionIn = MaterialRequisitionActionIn(), db: Session = Depends(db_session)):
    require_roles(request, "Admin", "Management", "Warehouse Manager")
    program_key = normalize_program(data.program)
    row = (
        db.query(MaterialTransfer)
        .filter(MaterialTransfer.id == transfer_id, MaterialTransfer.program == program_key)
        .with_for_update()
        .first()
    )
    if row is None:
        raise HTTPException(status_code=404, detail="Material transfer not found")
    if row.status != "approved":
        raise HTTPException(status_code=400, detail=f"Transfer cannot be confirmed from status {row.status}")
    require_warehouse_access(request, db, row.to_warehouse_id, program_key)
    actor = request_actor(request)

    for item in row.items:
        product = require_product(db, item.product_id, program_key)
        from_balance = locked_stock_balance(db, row.from_warehouse_id, item.product_id, program_key)
        if from_balance.quantity < item.quantity:
            raise HTTPException(
                status_code=400,
                detail=f"Insufficient stock for {product_display_name(product)}. Requested {item.quantity}, available {from_balance.quantity}.",
            )

    for item in row.items:
        from_balance = locked_stock_balance(db, row.from_warehouse_id, item.product_id, program_key)
        to_balance = locked_stock_balance(db, row.to_warehouse_id, item.product_id, program_key)
        from_balance.quantity -= item.quantity
        to_balance.quantity += item.quantity
        db.add(
            StockMovement(
                program=program_key,
                movement_type="transfer_out",
                product_id=item.product_id,
                warehouse_id=row.from_warehouse_id,
                quantity=-item.quantity,
                reference=row.transfer_number,
                note=f"Transfer to {row.to_warehouse.name if row.to_warehouse else row.to_warehouse_id}",
                created_by=actor,
            )
        )
        db.add(
            StockMovement(
                program=program_key,
                movement_type="transfer_in",
                product_id=item.product_id,
                warehouse_id=row.to_warehouse_id,
                quantity=item.quantity,
                reference=row.transfer_number,
                note=f"Transfer from {row.from_warehouse.name if row.from_warehouse else row.from_warehouse_id}",
                created_by=actor,
            )
        )

    row.receiver_name = actor
    row.receiver_date = local_today()
    row.receiver_comment = data.comment
    row.status = "transferred"
    log_audit(db, "confirm_material_transfer", "material_transfer", row.transfer_number, actor, data.model_dump())
    db.commit()
    db.refresh(row)
    return {"success": True, "transfer": transfer_to_dict(row)}


@app.post("/api/warehouse/material-returns")
def create_material_return(data: MaterialReturnIn, request: Request, db: Session = Depends(db_session)):
    user = require_roles(request, "Admin", "Management", "Warehouse Manager", "Requester")
    program_key = normalize_program(data.program)
    require_warehouse(db, data.warehouse_id, program_key)
    requester_submission = user.role.strip().lower() == "requester"
    if not requester_submission:
        require_warehouse_access(request, db, data.warehouse_id, program_key)
    data.created_by = request_actor(request)
    if not data.items:
        raise HTTPException(status_code=400, detail="At least one item is required")

    row = MaterialReturn(
        program=program_key,
        return_number=next_number(db, MaterialReturn, "RET"),
        return_date=data.return_date or local_today(),
        site_id=data.site_id.strip(),
        site_address=data.site_address.strip(),
        warehouse_id=data.warehouse_id,
        returned_by=data.returned_by.strip(),
        received_by="" if requester_submission else data.created_by,
        reason=data.reason.strip(),
        status="pending_warehouse" if requester_submission else "confirmed",
        created_by=data.created_by.strip() or "manager",
    )
    db.add(row)
    db.flush()

    for index, item in enumerate(data.items, start=1):
        product = require_product(db, item.product_id, program_key)
        if not requester_submission:
            locked_stock_balance(db, data.warehouse_id, item.product_id, program_key).quantity += item.quantity
        db.add(
            MaterialReturnItem(
                return_id=row.id,
                line_no=index,
                product_id=item.product_id,
                part_nbr=product_part_number(product.sku, product.part_number),
                description=product_display_name(product),
                uom=product.unit,
                quantity=item.quantity,
                condition=item.condition.strip() or "Good",
                remark=item.remark.strip(),
            )
        )
        if not requester_submission:
            db.add(
                StockMovement(
                    program=program_key,
                    movement_type="return_in",
                    product_id=item.product_id,
                    warehouse_id=data.warehouse_id,
                    quantity=item.quantity,
                    reference=row.return_number,
                    note=f"Returned from site {row.site_id or row.site_address}: {item.condition.strip() or 'Good'}",
                    created_by=row.created_by,
                )
            )

    log_audit(
        db,
        "submit_material_return" if requester_submission else "create_material_return",
        "material_return",
        row.return_number,
        row.created_by,
        data.model_dump(),
    )
    db.commit()
    db.refresh(row)
    return {"success": True, "return_number": row.return_number, "return": material_return_to_dict(row)}


@app.post("/api/warehouse/material-returns/{return_id}/approve")
def approve_material_return(return_id: int, data: MaterialRequisitionActionIn, request: Request, db: Session = Depends(db_session)):
    require_roles(request, "Admin", "Management", "Warehouse Manager")
    program_key = normalize_program(data.program)
    row = (
        db.query(MaterialReturn)
        .options(selectinload(MaterialReturn.items), joinedload(MaterialReturn.warehouse))
        .filter(MaterialReturn.id == return_id, MaterialReturn.program == program_key)
        .with_for_update(of=MaterialReturn)
        .first()
    )
    if row is None:
        raise HTTPException(status_code=404, detail="Material return not found")
    if row.status != "pending_warehouse":
        raise HTTPException(status_code=400, detail=f"Return cannot be approved from status {row.status}")
    require_warehouse_access(request, db, row.warehouse_id, program_key)

    actor = request_actor(request)
    for item in row.items:
        require_product(db, item.product_id, program_key)
        locked_stock_balance(db, row.warehouse_id, item.product_id, program_key).quantity += item.quantity
        db.add(
            StockMovement(
                program=program_key,
                movement_type="return_in",
                product_id=item.product_id,
                warehouse_id=row.warehouse_id,
                quantity=item.quantity,
                reference=row.return_number,
                source_item_id=item.id,
                note=f"Approved return from site {row.site_id or row.site_address}: {item.condition or 'Good'}",
                created_by=actor,
            )
        )
    row.received_by = actor
    row.status = "confirmed"
    log_audit(db, "approve_material_return", "material_return", row.return_number, actor, data.model_dump())
    db.commit()
    db.refresh(row)
    return {"success": True, "return": material_return_to_dict(row)}


@app.post("/api/warehouse/material-returns/{return_id}/reject")
def reject_material_return(return_id: int, data: MaterialRequisitionActionIn, request: Request, db: Session = Depends(db_session)):
    require_roles(request, "Admin", "Management", "Warehouse Manager")
    program_key = normalize_program(data.program)
    row = db.query(MaterialReturn).options(joinedload(MaterialReturn.warehouse)).filter(
        MaterialReturn.id == return_id,
        MaterialReturn.program == program_key,
    ).first()
    if row is None:
        raise HTTPException(status_code=404, detail="Material return not found")
    if row.status != "pending_warehouse":
        raise HTTPException(status_code=400, detail=f"Return cannot be rejected from status {row.status}")
    require_warehouse_access(request, db, row.warehouse_id, program_key)
    row.status = "rejected"
    row.received_by = request_actor(request)
    log_audit(db, "reject_material_return", "material_return", row.return_number, request_actor(request), data.model_dump())
    db.commit()
    db.refresh(row)
    return {"success": True, "return": material_return_to_dict(row)}


@app.post("/api/warehouse/material-requisitions/{requisition_id}/signature")
def sign_material_requisition(requisition_id: int, data: MaterialRequisitionSignatureIn, request: Request, db: Session = Depends(db_session)):
    program_key = normalize_program(data.program)
    row = db.query(MaterialRequisition).filter(MaterialRequisition.id == requisition_id, MaterialRequisition.program == program_key).first()
    if row is None:
        raise HTTPException(status_code=404, detail="Material requisition not found")
    actor = request_actor(request)
    if current_user(request).role.strip().lower() == "requester":
        if normalize_usage_key(row.created_by) != normalize_usage_key(actor) and current_user(request).role.strip().lower() != "admin":
            raise HTTPException(status_code=403, detail="You can only sign your own requisition")
        row.requester_name = actor
        row.requester_title = data.title or row.requester_title
        row.requester_date = data.date or row.requester_date
        row.requester_comment = data.comment
        row.requester_signature = data.signature
    else:
        require_roles(request, "Admin", "Approval", "Warehouse Manager")
        require_warehouse_access(request, db, row.warehouse_id, program_key)
        row.receiver_name = actor
        row.receiver_title = data.title or row.receiver_title
        row.receiver_date = data.date or row.receiver_date
        row.receiver_comment = data.comment
        row.receiver_signature = data.signature
    if row.status != "issued" and row.requester_signature and row.receiver_signature:
        row.status = "signed"
    log_audit(db, "sign_material_requisition", "material_requisition", row.order_number, actor, data.model_dump())
    db.commit()
    db.refresh(row)
    return {"success": True, "requisition": requisition_to_dict(row)}


@app.post("/api/warehouse/material-requisitions/{requisition_id}/approve")
def approve_material_requisition(requisition_id: int, data: MaterialRequisitionActionIn, request: Request, db: Session = Depends(db_session)):
    user = require_roles(request, "Admin", "Approval")
    program_key = normalize_program(data.program)
    row = (
        db.query(MaterialRequisition)
        .filter(MaterialRequisition.id == requisition_id, MaterialRequisition.program == program_key)
        .with_for_update()
        .first()
    )
    if row is None:
        raise HTTPException(status_code=404, detail="Material requisition not found")
    if row.status not in {"pending_approval", "draft", "rejected"}:
        raise HTTPException(status_code=400, detail=f"MR cannot be approved from status {row.status}")
    validate_reservable_stock(db, row.warehouse_id, row.items, program_key, exclude_requisition_id=row.id)
    actor = request_actor(request)
    # All Approval users share the approval queue. Record the person who
    # completed the action instead of relying on the draft's placeholder name.
    row.receiver_name = actor
    row.receiver_title = data.title or row.receiver_title
    row.receiver_date = local_today()
    row.receiver_comment = data.comment
    if data.signature:
        row.receiver_signature = data.signature
    row.status = "approved"
    log_audit(db, "approve_material_requisition", "material_requisition", row.order_number, actor or "approver", data.model_dump())
    db.commit()
    db.refresh(row)
    notify_mr_approved(row, db)
    return {"success": True, "requisition": requisition_to_dict(row)}


@app.post("/api/warehouse/material-requisitions/{requisition_id}/reject")
def reject_material_requisition(requisition_id: int, data: MaterialRequisitionActionIn, request: Request, db: Session = Depends(db_session)):
    user = require_roles(request, "Admin", "Approval")
    program_key = normalize_program(data.program)
    row = (
        db.query(MaterialRequisition)
        .filter(MaterialRequisition.id == requisition_id, MaterialRequisition.program == program_key)
        .with_for_update()
        .first()
    )
    if row is None:
        raise HTTPException(status_code=404, detail="Material requisition not found")
    if row.status not in {"pending_approval", "draft"}:
        raise HTTPException(status_code=400, detail=f"MR cannot be rejected from status {row.status}")
    actor = request_actor(request)
    row.receiver_name = actor
    row.receiver_title = data.title or row.receiver_title
    row.receiver_date = local_today()
    row.receiver_comment = data.comment
    row.status = "rejected"
    log_audit(db, "reject_material_requisition", "material_requisition", row.order_number, actor or "approver", data.model_dump())
    db.commit()
    db.refresh(row)
    notify_mr_rejected(row, db)
    return {"success": True, "requisition": requisition_to_dict(row)}


@app.post("/api/warehouse/material-requisitions/{requisition_id}/return-for-edit")
def return_material_requisition_for_edit(requisition_id: int, data: MaterialRequisitionActionIn, request: Request, db: Session = Depends(db_session)):
    user = require_roles(request, "Admin", "Management", "Warehouse Manager", "Approval")
    program_key = normalize_program(data.program)
    row = (
        db.query(MaterialRequisition)
        .filter(MaterialRequisition.id == requisition_id, MaterialRequisition.program == program_key)
        .with_for_update()
        .first()
    )
    if row is None:
        raise HTTPException(status_code=404, detail="Material requisition not found")
    role_key = normalize_usage_key(user.role)
    returning_from_approval = role_key in {"approval", "approver"} and row.status == "pending_approval"
    returning_from_warehouse = row.status == "approved"
    if not returning_from_approval and not returning_from_warehouse:
        raise HTTPException(status_code=400, detail=f"MR cannot be returned for edit from status {row.status}")
    if not returning_from_approval:
        require_warehouse_access(request, db, row.warehouse_id, program_key)
    actor = request_actor(request)
    row.receiver_name = actor or row.receiver_name
    row.receiver_title = data.title or row.receiver_title
    row.receiver_date = local_today()
    row.receiver_comment = data.comment
    row.return_reason = data.comment
    row.status = "returned_for_edit"
    log_audit(db, "return_material_requisition_for_edit", "material_requisition", row.order_number, actor or "reviewer", data.model_dump())
    db.commit()
    db.refresh(row)
    notify_mr_returned_for_edit(row, db)
    return {"success": True, "requisition": requisition_to_dict(row)}


@app.post("/api/warehouse/material-requisitions/{requisition_id}/issue")
def issue_material_requisition(requisition_id: int, request: Request, data: MaterialRequisitionActionIn = MaterialRequisitionActionIn(), db: Session = Depends(db_session)):
    require_roles(request, "Admin", "Management", "Warehouse Manager")
    program_key = normalize_program(data.program)
    row = (
        db.query(MaterialRequisition)
        .filter(MaterialRequisition.id == requisition_id, MaterialRequisition.program == program_key)
        .with_for_update()
        .first()
    )
    if row is None:
        raise HTTPException(status_code=404, detail="Material requisition not found")
    require_warehouse_access(request, db, row.warehouse_id, program_key)
    issue_order = issue_material_requisition_row(db, row, request_actor(request))
    db.commit()
    db.refresh(row)
    notify_mr_issued(row, db)
    return {"success": True, "issue_order": issue_order, "requisition": requisition_to_dict(row)}


@app.post("/api/warehouse/material-requisitions/{requisition_id}/delete")
def delete_material_requisition(requisition_id: int, request: Request, data: MaterialRequisitionActionIn = MaterialRequisitionActionIn(), db: Session = Depends(db_session)):
    require_roles(request, "Admin")
    row = (
        db.query(MaterialRequisition)
        .options(selectinload(MaterialRequisition.items), joinedload(MaterialRequisition.warehouse))
        .filter(MaterialRequisition.id == requisition_id, MaterialRequisition.program == normalize_program(data.program))
        .first()
    )
    if row is None:
        raise HTTPException(status_code=404, detail="Material requisition not found")
    try:
        result = delete_material_requisition_row(db, row, request_actor(request))
        db.commit()
    except Exception:
        db.rollback()
        raise
    return {"success": True, **result}


@app.post("/api/warehouse/receive")
def receive_stock(data: ReceiveIn, request: Request, db: Session = Depends(db_session)):
    require_roles(request, "Admin", "Management", "Warehouse Manager")
    program_key = normalize_program(data.program)
    require_warehouse_access(request, db, data.warehouse_id, program_key)
    data.created_by = request_actor(request)
    if not data.items:
        raise HTTPException(status_code=400, detail="At least one item is required")

    order = ReceiveOrder(
        program=program_key,
        order_number=data.receipt_number.strip() or next_number(db, ReceiveOrder, "GRN"),
        supplier=data.supplier.strip(),
        receipt_date=data.receipt_date,
        warehouse_id=data.warehouse_id,
        created_by=data.created_by,
    )
    db.add(order)
    db.flush()

    for item in data.items:
        product = require_product(db, item.product_id, program_key)
        if program_key == SINGLE_RAN_PROGRAM and data.supplier.strip() and not str(product.vendor or "").strip():
            product.vendor = data.supplier.strip()
        validate_serial_count(product, item.quantity, item.serial_numbers)
        stock_balance(db, data.warehouse_id, item.product_id, program_key).quantity += item.quantity

        serials = item.serial_numbers if product.tracking_type == "serialized" else [""]
        for serial in serials:
            if serial:
                exists = (
                    db.query(ProductSerial)
                    .filter(ProductSerial.program == program_key, ProductSerial.serial_number == serial)
                    .first()
                )
                if exists:
                    raise HTTPException(status_code=400, detail=f"Serial already exists: {serial}")
                db.add(
                    ProductSerial(
                        program=program_key,
                        product_id=item.product_id,
                        serial_number=serial,
                        status="in_warehouse",
                        warehouse_id=data.warehouse_id,
                    )
                )
            db.add(ReceiveOrderItem(receive_order_id=order.id, product_id=item.product_id, quantity=1 if serial else item.quantity, serial_number=serial))
            db.add(
                StockMovement(
                    program=program_key,
                    movement_type="receive",
                    product_id=item.product_id,
                    warehouse_id=data.warehouse_id,
                    quantity=1 if serial else item.quantity,
                    serial_number=serial,
                    reference=order.order_number,
                    created_by=data.created_by,
                )
            )

    log_audit(db, "receive_stock", "receive_order", order.order_number, data.created_by, data.model_dump())
    try:
        db.commit()
    except Exception:
        db.rollback()
        raise
    return {"success": True, "order_number": order.order_number}


@app.post("/api/warehouse/receive-inventory")
def receive_inventory(data: InventoryReceiveIn, request: Request, db: Session = Depends(db_session)):
    require_roles(request, "Admin", "Management", "Warehouse Manager")
    program_key = normalize_program(data.program)
    require_warehouse_access(request, db, data.warehouse_id, program_key)
    data.created_by = request_actor(request)
    sku = data.sku.strip()
    name = material_display_name(data.name.strip(), sku)
    if not sku:
        raise HTTPException(status_code=400, detail="SKU is required")
    if not name:
        raise HTTPException(status_code=400, detail="Material name is required")

    product = db.query(Product).filter(Product.program == program_key, Product.sku == sku).first()
    if product is None:
        product = Product(
            program=program_key,
            sku=sku,
            part_number=product_part_number(sku, data.part_number),
            category=data.category.strip(),
            name=name,
            item_detail=name,
            vendor=data.supplier.strip() if program_key == SINGLE_RAN_PROGRAM else "",
            qr_code=data.qr_code.strip(),
            unit=data.unit.strip() or "PCS",
            tracking_type="bulk",
            min_stock=0,
        )
        db.add(product)
        db.flush()
    else:
        product.part_number = product_part_number(sku, data.part_number or product.part_number)
        product.name = material_display_name(name, sku) or product.name
        product.item_detail = product.item_detail or name
        product.unit = data.unit.strip() or product.unit or "PCS"
        product.qr_code = data.qr_code.strip() or product.qr_code
        product.category = data.category.strip() or product.category
        if program_key == SINGLE_RAN_PROGRAM and data.supplier.strip() and not str(product.vendor or "").strip():
            product.vendor = data.supplier.strip()

    order = ReceiveOrder(
        program=program_key,
        order_number=data.receipt_number.strip() or next_number(db, ReceiveOrder, "GRN"),
        supplier=data.supplier.strip(),
        receipt_date=data.receipt_date,
        warehouse_id=data.warehouse_id,
        created_by=data.created_by,
    )
    db.add(order)
    db.flush()

    stock_balance(db, data.warehouse_id, product.id, program_key).quantity += data.quantity
    db.add(ReceiveOrderItem(receive_order_id=order.id, product_id=product.id, quantity=data.quantity, serial_number=""))
    db.add(
        StockMovement(
            program=program_key,
            movement_type="receive",
            product_id=product.id,
            warehouse_id=data.warehouse_id,
            quantity=data.quantity,
            serial_number="",
            reference=order.order_number,
            created_by=data.created_by,
            note=f"Inventory receive: {data.receipt_number.strip()}",
        )
    )
    log_audit(db, "receive_inventory", "receive_order", order.order_number, data.created_by, data.model_dump())
    try:
        db.commit()
    except Exception as exc:
        db.rollback()
        raise HTTPException(status_code=400, detail="Receipt number already exists or receive failed") from exc
    db.refresh(order)
    db.refresh(product)
    balance = stock_balance(db, data.warehouse_id, product.id, program_key)
    return {
        "success": True,
        "order_number": order.order_number,
        "product": product_to_dict(product),
        "receipt": receive_order_to_dict(order),
        "balance": balance.quantity,
    }


@app.post("/api/warehouse/adjust-inventory")
def adjust_inventory(data: InventoryAdjustmentIn, request: Request, db: Session = Depends(db_session)):
    require_roles(request, "Admin", "Management", "Warehouse Manager")
    program_key = normalize_program(data.program)
    require_warehouse_access(request, db, data.warehouse_id, program_key)
    data.created_by = request_actor(request)
    product = db.query(Product).filter(Product.program == program_key, Product.sku == data.sku.strip()).first()
    if product is None:
        raise HTTPException(status_code=404, detail="SKU not found")

    balance = stock_balance(db, data.warehouse_id, product.id, program_key)
    old_quantity = balance.quantity or 0
    delta = data.quantity - old_quantity
    balance.quantity = data.quantity

    if delta:
        db.add(
            StockMovement(
                program=program_key,
                movement_type="adjustment",
                product_id=product.id,
                warehouse_id=data.warehouse_id,
                quantity=delta,
                serial_number="",
                reference=f"ADJ-{local_today()}",
                created_by=data.created_by,
                note=data.note.strip() or f"Stock adjusted from {old_quantity} to {data.quantity}",
            )
        )
    log_audit(
        db,
        "adjust_inventory",
        "stock_balance",
        f"{data.warehouse_id}:{product.id}",
        data.created_by,
        {"sku": product.sku, "old_quantity": old_quantity, "new_quantity": data.quantity, "delta": delta, "note": data.note},
    )
    db.commit()
    return {
        "success": True,
        "sku": product.sku,
        "product": product_display_name(product),
        "old_quantity": old_quantity,
        "new_quantity": data.quantity,
        "delta": delta,
    }


@app.post("/api/warehouse/issue")
def issue_to_technician(data: IssueIn, request: Request, db: Session = Depends(db_session)):
    require_roles(request, "Admin", "Management", "Warehouse Manager")
    program_key = normalize_program(data.program)
    require_warehouse_access(request, db, data.warehouse_id, program_key)
    data.created_by = request_actor(request)
    require_technician(db, data.technician_id, program_key)
    if not data.items:
        raise HTTPException(status_code=400, detail="At least one item is required")

    order = IssueOrder(
        program=program_key,
        order_number=next_number(db, IssueOrder, "ISS"),
        warehouse_id=data.warehouse_id,
        technician_id=data.technician_id,
        created_by=data.created_by,
    )
    db.add(order)
    db.flush()

    for item in data.items:
        product = require_product(db, item.product_id, program_key)
        validate_serial_count(product, item.quantity, item.serial_numbers)
        balance = stock_balance(db, data.warehouse_id, item.product_id, program_key)
        if balance.quantity < item.quantity:
            warehouse = db.get(Warehouse, data.warehouse_id)
            warehouse_name = warehouse.name if warehouse else str(data.warehouse_id)
            material_name = product_display_name(product) or product.sku or f"product {product.id}"
            raise HTTPException(
                status_code=400,
                detail=(
                    f"Insufficient stock for {material_name} in {warehouse_name}. "
                    f"Requested {item.quantity}, available {balance.quantity}."
                ),
            )
        balance.quantity -= item.quantity
        technician_balance(db, data.technician_id, item.product_id, program_key).quantity += item.quantity

        serials = item.serial_numbers if product.tracking_type == "serialized" else [""]
        for serial in serials:
            if serial:
                serial_row = (
                    db.query(ProductSerial)
                    .filter(
                        ProductSerial.product_id == item.product_id,
                        ProductSerial.program == program_key,
                        ProductSerial.serial_number == serial,
                        ProductSerial.status == "in_warehouse",
                        ProductSerial.warehouse_id == data.warehouse_id,
                    )
                    .first()
                )
                if serial_row is None:
                    raise HTTPException(status_code=400, detail=f"Serial not available in warehouse: {serial}")
                serial_row.status = "with_technician"
                serial_row.warehouse_id = None
                serial_row.technician_id = data.technician_id

            db.add(IssueOrderItem(issue_order_id=order.id, product_id=item.product_id, quantity=1 if serial else item.quantity, serial_number=serial))
            db.add(
                StockMovement(
                    program=program_key,
                    movement_type="issue_to_technician",
                    product_id=item.product_id,
                    warehouse_id=data.warehouse_id,
                    technician_id=data.technician_id,
                    quantity=-(1 if serial else item.quantity),
                    serial_number=serial,
                    reference=order.order_number,
                    created_by=data.created_by,
                )
            )

    log_audit(db, "issue_to_technician", "issue_order", order.order_number, data.created_by, data.model_dump())
    try:
        db.commit()
    except Exception:
        db.rollback()
        raise
    return {"success": True, "order_number": order.order_number}
